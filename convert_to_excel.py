#!/usr/bin/env python3
"""
Script mejorado para convertir SISTEMAS_PROMPTS_CONSOLIDADO.md a formato Excel (.xlsx)
Versión mejorada con mejor formato, navegación y funcionalidades
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_excel_from_markdown(md_file, xlsx_file):
    """Convierte un archivo markdown a formato Excel con formato mejorado"""
    
    # Leer el archivo markdown
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Estilos mejorados con más opciones
    styles = {
        'header_font': Font(bold=True, size=16, color="FFFFFF"),
        'header_fill': PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        'title_font': Font(bold=True, size=13, color="1F4E78"),
        'title_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'subtitle_font': Font(bold=True, size=11, color="2F5597"),
        'subtitle_fill': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),
        'code_font': Font(name="Courier New", size=9, color="006400"),
        'code_fill': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        'xml_font': Font(name="Courier New", size=10, color="000080", bold=True),
        'xml_fill': PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid"),
        'normal_font': Font(size=11, name="Calibri"),
        'bold_font': Font(size=11, name="Calibri", bold=True),
        'italic_font': Font(size=11, name="Calibri", italic=True),
        'warning_font': Font(size=10, name="Calibri", italic=True, color="CC6600"),
        'warning_fill': PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
        'note_font': Font(size=10, name="Calibri", italic=True, color="0066CC"),
        'note_fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
    }
    
    thin_border = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC")
    )
    
    # Dividir por sistemas principales (# título)
    sections = re.split(r'^#\s+([^\n]+)', content, flags=re.MULTILINE)
    
    sheet_names = []
    current_sheet = None
    stats = {'total_sections': 0, 'total_tables': 0, 'total_code_blocks': 0}
    
    # Procesar secciones
    for i, section in enumerate(sections):
        if i == 0:
            # Contenido antes del primer título - crear hoja de introducción
            if section.strip():
                intro_sheet = wb.create_sheet(title="📖 Introducción")
                process_content(intro_sheet, section, styles, thin_border, stats)
            continue
        
        if i % 2 == 1:
            # Es un título - crear nueva hoja
            title = section.strip()
            # Limpiar y acortar nombre de hoja
            sheet_name = clean_sheet_name(title)
            if not sheet_name:
                sheet_name = f"Sheet_{len(wb.worksheets) + 1}"
            
            current_sheet = wb.create_sheet(title=sheet_name)
            sheet_names.append((sheet_name, title))
            stats['total_sections'] += 1
            current_row = 1
            
            # Agregar título principal con estilo mejorado
            cell = current_sheet.cell(row=current_row, column=1, value=title)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
            current_sheet.merge_cells(f'A{current_row}:F{current_row}')
            current_sheet.row_dimensions[current_row].height = 35
            current_row += 1
            
            # Agregar espacio
            current_row += 1
        
        else:
            # Es contenido
            if current_sheet is None:
                continue
            
            final_row = process_content(current_sheet, section, styles, thin_border, stats, int(current_row))
            current_row = final_row
    
    # Ajustar ancho de columnas en todas las hojas
    for sheet in wb.worksheets:
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 60
        sheet.column_dimensions['D'].width = 60
        sheet.column_dimensions['E'].width = 60
        sheet.column_dimensions['F'].width = 30
        
        # Aplicar filtros a la primera fila de datos (si existe)
        if sheet.max_row > 1:
            try:
                sheet.auto_filter.ref = f"A1:F{sheet.max_row}"
            except:
                pass
    
    # Crear hoja de índice mejorada
    create_index_sheet(wb, sheet_names, styles, thin_border, stats)
    
    # Guardar
    wb.save(xlsx_file)
    print(f"✅ Documento Excel mejorado creado exitosamente: {xlsx_file}")
    print(f"📊 Ubicación: {xlsx_file}")
    print(f"📑 Hojas creadas: {len(wb.worksheets)}")
    print(f"📋 Sistemas documentados: {stats['total_sections']}")
    print(f"📊 Tablas procesadas: {stats['total_tables']}")
    print(f"💻 Bloques de código: {stats['total_code_blocks']}")

def clean_sheet_name(name):
    """Limpia el nombre para que sea válido en Excel"""
    # Remover emojis y caracteres especiales
    name = re.sub(r'[^\w\s-]', '', name)
    name = name.replace(' ', '_')
    # Excel limita a 31 caracteres
    return name[:31] if name else None

def process_content(sheet, content, styles, border, stats, start_row=1):
    """Procesa el contenido markdown y lo agrega a la hoja. Retorna el número de fila final."""
    
    lines = content.split('\n')
    current_row = int(start_row)
    in_code_block = False
    code_lines = []
    in_table = False
    table_rows = []
    code_block_language = ""
    
    for line in lines:
        line_stripped = line.strip()
        original_line = line
        indent_level = (len(original_line) - len(original_line.lstrip())) // 2
        
        # Bloque de código
        if line_stripped.startswith('```'):
            if in_code_block:
                # Fin del bloque
                if code_lines:
                    cell = sheet.cell(row=current_row, column=2, value='\n'.join(code_lines))
                    cell.font = styles['code_font']
                    cell.fill = styles['code_fill']
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = border
                    sheet.merge_cells(f'B{current_row}:F{current_row}')
                    sheet.row_dimensions[current_row].height = max(20, len(code_lines) * 1.2)
                    current_row += 1
                    stats['total_code_blocks'] += 1
                code_lines = []
                code_block_language = ""
                in_code_block = False
            else:
                in_code_block = True
                code_block_language = line_stripped[3:].strip()
            continue
        
        if in_code_block:
            code_lines.append(line)
            continue
        
        # Tablas markdown
        if '|' in line_stripped and line_stripped.startswith('|'):
            if not in_table:
                in_table = True
                table_rows = []
            table_rows.append(line_stripped)
            continue
        elif in_table:
            # Procesar tabla
            table_start_row = process_table(sheet, table_rows, current_row, styles, border)
            current_row = table_start_row
            table_rows = []
            in_table = False
            stats['total_tables'] += 1
        
        # Título nivel 2 (##)
        if line_stripped.startswith('## ') and not line_stripped.startswith('###'):
            text = line_stripped[3:].strip()
            current_row += 1
            cell = sheet.cell(row=current_row, column=2, value=text)
            cell.font = styles['title_font']
            cell.fill = styles['title_fill']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
            sheet.merge_cells(f'B{current_row}:F{current_row}')
            sheet.row_dimensions[current_row].height = 25
            current_row += 1
        
        # Título nivel 3 (###)
        elif line_stripped.startswith('### '):
            text = line_stripped[4:].strip()
            current_row += 1
            cell = sheet.cell(row=current_row, column=3, value=text)
            cell.font = styles['subtitle_font']
            cell.fill = styles['subtitle_fill']
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
            sheet.merge_cells(f'C{current_row}:F{current_row}')
            sheet.row_dimensions[current_row].height = 20
            current_row += 1
        
        # Línea horizontal
        elif line_stripped == '---' or (line_stripped.startswith('---') and len(line_stripped) <= 5):
            current_row += 1
            cell = sheet.cell(row=current_row, column=2, value='─' * 100)
            cell.font = Font(size=8, color="CCCCCC")
            sheet.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
        
        # Lista ordenada
        elif re.match(r'^\d+\.\s+', line_stripped):
            text = re.sub(r'^\d+\.\s+', '', line_stripped)
            text = clean_markdown_formatting(text)
            cell = sheet.cell(row=current_row, column=3, value=f"  {text}")
            cell.font = styles['normal_font']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
            cell.border = border
            sheet.merge_cells(f'C{current_row}:F{current_row}')
            current_row += 1
        
        # Lista con viñetas
        elif line_stripped.startswith('- ') or line_stripped.startswith('* '):
            text = line_stripped[2:].strip()
            text = clean_markdown_formatting(text)
            # Detectar nivel de anidación
            indent_text = "  " * indent_level + "• "
            cell = sheet.cell(row=current_row, column=3, value=f"{indent_text}{text}")
            cell.font = styles['normal_font']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=indent_level)
            cell.border = border
            sheet.merge_cells(f'C{current_row}:F{current_row}')
            current_row += 1
        
        # Etiquetas XML (goal, format_rules, etc.)
        elif line_stripped.startswith('<') and line_stripped.endswith('>'):
            current_row += 1
            cell = sheet.cell(row=current_row, column=2, value=line_stripped)
            cell.font = styles['xml_font']
            cell.fill = styles['xml_fill']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border
            sheet.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
        
        # Advertencias y notas especiales
        elif line_stripped.upper().startswith('NEVER') or line_stripped.upper().startswith('AVOID'):
            text = clean_markdown_formatting(line_stripped)
            cell = sheet.cell(row=current_row, column=2, value=f"⚠️ {text}")
            cell.font = styles['warning_font']
            cell.fill = styles['warning_fill']
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = border
            sheet.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
        
        # Párrafo normal
        elif line_stripped:
            text = clean_markdown_formatting(line_stripped)
            if text:
                # Detectar si es una nota o información importante
                if text.upper().startswith('NOTE:') or text.upper().startswith('IMPORTANT:'):
                    cell = sheet.cell(row=current_row, column=2, value=f"ℹ️ {text}")
                    cell.font = styles['note_font']
                    cell.fill = styles['note_fill']
                else:
                    cell = sheet.cell(row=current_row, column=2, value=text)
                    cell.font = styles['normal_font']
                
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = border
                sheet.merge_cells(f'B{current_row}:F{current_row}')
                current_row += 1
        
        # Línea vacía
        else:
            if current_row > 1:  # No agregar espacio al inicio
                # Solo agregar espacio si la última fila tiene contenido
                try:
                    if sheet.cell(row=current_row-1, column=2).value:
                        current_row += 1
                except:
                    pass
    
    return int(current_row)

def clean_markdown_formatting(text):
    """Limpia el formato markdown del texto manteniendo estructura"""
    # Negrita
    text = re.sub(r'\*\*([^\*]+)\*\*', r'\1', text)
    # Cursiva
    text = re.sub(r'(?<!\*)\*([^\*]+)\*(?!\*)', r'\1', text)
    # Enlaces (mantener solo el texto)
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    # Código inline
    text = re.sub(r'`([^`]+)`', r'\1', text)
    # Referencias de citas
    text = re.sub(r'\[\d+\]', '', text)
    return text.strip()

def process_table(sheet, table_rows, start_row, styles, border):
    """Procesa una tabla markdown y la agrega a la hoja"""
    if not table_rows or len(table_rows) < 2:
        return start_row
    
    current_row = start_row + 1
    headers = []
    
    # Procesar encabezados
    if '|' in table_rows[0]:
        headers = [cell.strip() for cell in table_rows[0].split('|')[1:-1]]
        # Agregar encabezados con estilo mejorado
        for col_idx, header in enumerate(headers, start=2):
            if col_idx <= 7:  # Limitar a columnas disponibles
                cell = sheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = styles['bold_font']
                cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
        current_row += 1
    
    # Procesar separador (si existe)
    start_idx = 2 if len(table_rows) > 1 and '---' in table_rows[1] else 1
    
    # Procesar filas de datos
    for row_idx in range(start_idx, len(table_rows)):
        if '|' in table_rows[row_idx]:
            cells = [cell.strip() for cell in table_rows[row_idx].split('|')[1:-1]]
            for col_idx, cell_value in enumerate(cells, start=2):
                if col_idx <= 7:  # Limitar a columnas disponibles
                    cell = sheet.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.font = styles['normal_font']
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.border = border
                    # Alternar color de filas para mejor legibilidad
                    if (current_row - start_row - 1) % 2 == 0:
                        cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
            current_row += 1
    
    return current_row

def create_index_sheet(wb, sheet_names, styles, border, stats):
    """Crea una hoja de índice mejorada con hipervínculos y estadísticas"""
    index_sheet = wb.create_sheet(title="📋 Índice", index=0)
    
    # Título principal
    cell = index_sheet.cell(row=1, column=1, value="📚 Índice de Sistemas de Prompts")
    cell.font = Font(bold=True, size=18, color="FFFFFF")
    cell.fill = styles['header_fill']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    index_sheet.merge_cells('A1:D1')
    index_sheet.row_dimensions[1].height = 40
    
    # Información de generación
    row = 2
    info_text = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell = index_sheet.cell(row=row, column=1, value=info_text)
    cell.font = Font(italic=True, size=9, color="666666")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    index_sheet.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # Encabezados de columna
    headers = ["#", "Sistema", "Descripción", "Acceso"]
    for col_idx, header in enumerate(headers, start=1):
        cell = index_sheet.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12, color="1F4E78")
        cell.fill = styles['title_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1
    
    # Agregar sistemas
    for idx, (sheet_name, title) in enumerate(sheet_names, 1):
        # Número
        cell = index_sheet.cell(row=row, column=1, value=idx)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Título del sistema
        cell = index_sheet.cell(row=row, column=2, value=title)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Descripción
        description = f"Sistema completo de prompts para {title.lower().replace('sistema', '').strip()}"
        cell = index_sheet.cell(row=row, column=3, value=description)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        
        # Hipervínculo a la hoja
        cell = index_sheet.cell(row=row, column=4, value=f"👉 Ir a {sheet_name}")
        cell.font = Font(size=11, color="0563C1", underline='single')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.hyperlink = f"#{sheet_name}!A1"
        cell.border = border
        
        # Alternar color de filas
        if idx % 2 == 0:
            for col in range(1, 5):
                index_sheet.cell(row=row, column=col).fill = PatternFill(
                    start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        row += 1
    
    # Ajustar columnas
    index_sheet.column_dimensions['A'].width = 5
    index_sheet.column_dimensions['B'].width = 50
    index_sheet.column_dimensions['C'].width = 60
    index_sheet.column_dimensions['D'].width = 20
    
    # Agregar estadísticas
    row += 2
    stats_row = row
    
    # Título de estadísticas
    cell = index_sheet.cell(row=row, column=1, value="📊 Estadísticas del Documento")
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.fill = styles['subtitle_fill']
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border
    index_sheet.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Estadísticas
    stats_data = [
        ("Total de sistemas", stats['total_sections']),
        ("Tablas procesadas", stats['total_tables']),
        ("Bloques de código", stats['total_code_blocks']),
        ("Hojas creadas", len(wb.worksheets)),
    ]
    
    for stat_name, stat_value in stats_data:
        cell = index_sheet.cell(row=row, column=1, value=stat_name)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        cell = index_sheet.cell(row=row, column=2, value=stat_value)
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        index_sheet.merge_cells(f'B{row}:D{row}')
        row += 1

if __name__ == '__main__':
    input_file = '01_marketing/SISTEMAS_PROMPTS_CONSOLIDADO.md'
    output_file = '01_marketing/SISTEMAS_PROMPTS_CONSOLIDADO.xlsx'
    
    try:
        create_excel_from_markdown(input_file, output_file)
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {input_file}")
    except Exception as e:
        print(f"❌ Error al convertir: {e}")
        import traceback
        traceback.print_exc()
