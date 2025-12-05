import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def enhance_excel_v3(filename):
    # First run v2 creation to ensure base exists (importing function would be cleaner but for script isolation we'll load or recreate)
    # Assuming V2 exists from previous step. If not, user needs to run V2 first.
    # We will load V2 and append new sheets.
    
    try:
        wb = load_workbook("Ventura_Capital_Herramientas_V2.xlsx")
    except FileNotFoundError:
        print("Error: V2 Excel not found. Please run the previous step first.")
        return

    # --- Styles ---
    header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Segoe UI', size=10)
    center_align = Alignment(horizontal='center')
    
    def format_new_sheet(ws):
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font

    # --- 1. SaaS Cohort Analysis ---
    ws_cohort = wb.create_sheet("Cohort Analysis")
    
    # Example Cohort Data (Retention %)
    cohort_data = {
        "Cohort": ["Jan-24", "Feb-24", "Mar-24", "Apr-24", "May-24", "Jun-24"],
        "New Users": [50, 65, 80, 95, 120, 150],
        "Month 1": ["100%", "100%", "100%", "100%", "100%", "100%"],
        "Month 2": ["95%", "94%", "96%", "95%", "97%", ""],
        "Month 3": ["90%", "88%", "92%", "90%", "", ""],
        "Month 4": ["85%", "82%", "88%", "", "", ""],
        "Month 5": ["80%", "78%", "", "", "", ""],
        "Month 6": ["75%", "", "", "", "", ""]
    }
    
    df_cohort = pd.DataFrame(cohort_data)
    
    for r in dataframe_to_rows(df_cohort, index=False, header=True):
        ws_cohort.append(r)
        
    format_new_sheet(ws_cohort)
    
    # Conditional Formatting for Heatmap effect (simplified)
    # (In a real scenario, we'd add color scales here)

    # --- 2. Staffing Plan ---
    ws_staff = wb.create_sheet("Staffing Plan")
    
    staff_data = {
        "Role": ["CEO", "CTO", "Lead Dev", "Sales Mgr", "Mkt Specialist", "CS Rep", "Total Headcount"],
        "Type": ["Founder", "Founder", "FTE", "FTE", "FTE", "FTE", ""],
        "Salary (Annual)": [120000, 120000, 90000, 80000, 60000, 45000, ""],
        "Q1 2025": [1, 1, 1, 1, 0, 1, 5],
        "Q2 2025": [1, 1, 2, 1, 1, 2, 8],
        "Q3 2025": [1, 1, 3, 2, 2, 3, 12],
        "Q4 2025": [1, 1, 4, 3, 2, 4, 15],
        "Cost Q4": ["$30k", "$30k", "$90k", "$60k", "$30k", "$45k", "$285k"]
    }
    
    df_staff = pd.DataFrame(staff_data)
    
    for r in dataframe_to_rows(df_staff, index=False, header=True):
        ws_staff.append(r)
        
    format_new_sheet(ws_staff)
    
    # Adjust column widths
    for ws in [ws_cohort, ws_staff]:
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 5, 40)

    output_filename = "Ventura_Capital_Herramientas_V3.xlsx"
    wb.save(output_filename)
    print(f"Enhanced Excel V3 created: {output_filename}")

if __name__ == "__main__":
    enhance_excel_v3("Ventura_Capital_Herramientas_V3.xlsx")







