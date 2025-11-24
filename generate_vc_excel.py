import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart

def create_excel_workbook(filename):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    workbook = writer.book

    # --- Styles ---
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Dark Blue
    
    subheader_font = Font(name='Calibri', size=11, bold=True)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') # Light Blue
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))

    def format_sheet(worksheet):
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 5, 60)
            
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

    # --- 1. Due Diligence Checklist (from Anexo C) ---
    dd_data = [
        ["Category", "Item", "Status", "Notes"],
        ["Corporativos", "Artículos de Incorporación (original y enmiendas)", "Pending", ""],
        ["Corporativos", "Estatutos Corporativos (bylaws)", "Pending", ""],
        ["Corporativos", "Registro de Acciones (stock ledger)", "Pending", ""],
        ["Corporativos", "Minutas de Juntas (board minutes)", "Pending", ""],
        ["Corporativos", "Resoluciones de Accionistas", "Pending", ""],
        ["Corporativos", "Registro de Directores y Oficiales", "Pending", ""],
        ["Corporativos", "Permisos y Licencias", "Pending", ""],
        ["Corporativos", "Registro de Marcas", "Pending", ""],
        ["Capitalización", "Cap Table Actualizado", "Pending", ""],
        ["Capitalización", "Acuerdos de Accionistas", "Pending", ""],
        ["Capitalización", "Planes de Opciones (stock option plans)", "Pending", ""],
        ["Capitalización", "Acuerdos de Vesting", "Pending", ""],
        ["Capitalización", "Convertible Notes / Warrants", "Pending", ""],
        ["Financieros", "Estados Financieros Auditados (3 años)", "Pending", ""],
        ["Financieros", "Estados Financieros No Auditados (trimestrales)", "Pending", ""],
        ["Financieros", "Flujo de Caja", "Pending", ""],
        ["Financieros", "Presupuestos y Proyecciones", "Pending", ""],
        ["Financieros", "Informes de Auditoría", "Pending", ""],
        ["Legales", "Contratos de Clientes", "Pending", ""],
        ["Legales", "Contratos de Proveedores", "Pending", ""],
        ["Legales", "Contratos de Empleados & NDAs", "Pending", ""],
        ["Legales", "Acuerdos de Propiedad Intelectual", "Pending", ""],
        ["Técnicos", "Patentes y Marcas Registradas", "Pending", ""],
        ["Técnicos", "Documentación Técnica y Arquitectura", "Pending", ""],
        ["Técnicos", "Seguridad y Compliance", "Pending", ""],
        ["Mercado", "Estudios de Mercado y Competencia", "Pending", ""],
        ["Mercado", "Estrategia de Marketing y Ventas", "Pending", ""],
        ["Mercado", "Métricas (CAC, LTV, Churn)", "Pending", ""]
    ]
    
    df_dd = pd.DataFrame(dd_data[1:], columns=dd_data[0])
    df_dd.to_excel(writer, sheet_name='Due Diligence Checklist', index=False)
    format_sheet(writer.sheets['Due Diligence Checklist'])

    # --- 2. Financial Projections (from Anexo B) ---
    # Base Scenario
    fin_base_data = {
        "Métrica": ["ARR", "MRR", "Usuarios Activos", "CAC", "LTV", "LTV/CAC", "Churn Rate", "Gross Margin", "EBITDA", "Cash Flow"],
        "2024": [139000, 11600, 500, 150, 1350, "9:1", "5%", "85%", -865000, -800000],
        "2025": [1460000, 121500, 2500, 120, 2400, "20:1", "3%", "87%", -350000, -200000],
        "2026": [3600000, 300000, 8000, 100, 4200, "42:1", "2%", "89%", 1600000, 1200000],
        "CAGR": ["+1,200%", "+1,500%", "+300%", "-33%", "+211%", "+367%", "-60%", "+4.7%", "-", "-"]
    }
    df_fin_base = pd.DataFrame(fin_base_data)
    df_fin_base.to_excel(writer, sheet_name='Proyecciones Financieras', startrow=1, index=False)
    
    ws_fin = writer.sheets['Proyecciones Financieras']
    ws_fin['A1'] = "Escenario Base (Realista)"
    ws_fin['A1'].font = Font(bold=True, size=14)
    
    # Add chart for ARR Growth
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Proyección ARR (Escenario Base)"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "Año"
    
    data = Reference(ws_fin, min_col=2, min_row=2, max_row=2, max_col=4) # ARR row
    cats = Reference(ws_fin, min_col=2, min_row=2, max_row=2) # Headers 2024-2026 (using first row values roughly)
    # Actually specific series
    # Let's restructure for chart friendly format in another sheet or below? 
    # Easier: just data table for now.
    
    format_sheet(ws_fin)

    # --- 3. Cap Table (from Anexo B) ---
    cap_data = [
        ["Ronda", "Accionista", "Acciones", "%", "Valor"],
        ["Actual", "Fundadores", 8000000, "80%", 6400000],
        ["Actual", "ESOP", 2000000, "20%", 1600000],
        ["Actual", "Total", 10000000, "100%", 8000000],
        ["Series A", "Fundadores", 8000000, "64%", 6400000],
        ["Series A", "ESOP", 2000000, "16%", 1600000],
        ["Series A", "Series A Investors", 2500000, "20%", 2000000],
        ["Series A", "Total", 12500000, "100%", 10000000],
    ]
    df_cap = pd.DataFrame(cap_data[1:], columns=cap_data[0])
    df_cap.to_excel(writer, sheet_name='Cap Table', index=False)
    format_sheet(writer.sheets['Cap Table'])

    # --- 4. Valuation Analysis (from Anexo B) ---
    val_data = [
        ["Método", "Detalle", "Valuación Est."],
        ["Múltiplos ARR (Base)", "12x ARR 2026 ($3.6M)", "$43.2M"],
        ["DCF", "Valor Presente Flujos + Terminal", "$30.9M"],
        ["Comparables", "Ajustado LATAM (20x -> 11x)", "$39.6M"],
        ["Promedio", "", "$37.9M"]
    ]
    df_val = pd.DataFrame(val_data[1:], columns=val_data[0])
    df_val.to_excel(writer, sheet_name='Valuación', index=False)
    format_sheet(writer.sheets['Valuación'])

    # --- 5. Metrics Tracking (from Anexo C) ---
    metrics_data = [
        ["Periodo", "Métrica", "Objetivo", "Actual", "Status"],
        ["Semanal", "MRR Growth", "5%", "", ""],
        ["Semanal", "New Customers", "10", "", ""],
        ["Mensual", "Churn Rate", "<5%", "", ""],
        ["Mensual", "CAC Payback", "<6 meses", "", ""],
        ["Trimestral", "LTV/CAC", ">10:1", "", ""],
        ["Anual", "Market Share", "1%", "", ""]
    ]
    df_metrics = pd.DataFrame(metrics_data[1:], columns=metrics_data[0])
    df_metrics.to_excel(writer, sheet_name='KPIs Dashboard', index=False)
    format_sheet(writer.sheets['KPIs Dashboard'])

    writer.close()
    print(f"Excel file created: {filename}")

if __name__ == "__main__":
    create_excel_workbook("Ventura_Capital_Herramientas.xlsx")


