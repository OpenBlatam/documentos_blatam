import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_marketing_budget(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Marketing Budget 2025"
    
    # Styles
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid') # Green
    header_font = Font(color='FFFFFF', bold=True)
    currency_fmt = '$#,##0'
    
    # --- Budget Table ---
    headers = ["Channel", "Monthly Budget", "Est. CAC", "Exp. Monthly Customers", "Annual Cost", "Annual Revenue Impact"]
    ws.append(headers)
    
    data = [
        ("LinkedIn Ads (B2B)", 15000, 350, 43, 180000, 645000),
        ("Meta Ads (Retargeting)", 8000, 120, 67, 96000, 240000),
        ("Google Search (Intent)", 10000, 200, 50, 120000, 300000),
        ("Content / SEO (Agency)", 5000, 50, 100, 60000, 180000),
        ("Events & Conferences", 4000, 1000, 4, 48000, 144000),
        ("Influencer / Affiliates", 3000, 100, 30, 36000, 108000),
        ("TOTAL", 45000, 153, 294, 540000, 1617000)
    ]
    
    for row in data:
        ws.append(row)

    # Formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Number Formats
    # iter_rows returns tuples of Cells. 
    # We need to iterate cells, not use index on the tuple if we specified min/max col correctly or just check length
    for row in ws.iter_rows(min_row=2):
        if len(row) > 1: row[1].number_format = currency_fmt
        if len(row) > 2: row[2].number_format = currency_fmt
        if len(row) > 4: row[4].number_format = currency_fmt
        if len(row) > 5: row[5].number_format = currency_fmt

    # --- ROI Summary (Second Table) ---
    ws.cell(row=10, column=1, value="ROI Summary").font = Font(bold=True, size=12)
    
    ws.cell(row=11, column=1, value="Total Annual Spend")
    ws.cell(row=11, column=2, value="=E8") # Ref Total Annual Cost
    ws.cell(row=11, column=2).number_format = currency_fmt
    
    ws.cell(row=12, column=1, value="Total Est. Revenue")
    ws.cell(row=12, column=2, value="=F8") # Ref Total Revenue
    ws.cell(row=12, column=2).number_format = currency_fmt
    
    ws.cell(row=13, column=1, value="Marketing ROI (ROAS)")
    ws.cell(row=13, column=2, value="=F8/E8")
    ws.cell(row=13, column=2).number_format = '0.0x'

    # Adjust Widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    wb.save(filename)
    print(f"Marketing Budget created: {filename}")

if __name__ == "__main__":
    create_marketing_budget("Ventura_Capital_Marketing_Budget_ROI.xlsx")

