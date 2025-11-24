import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_dd_tracker(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "DD Request Tracker"
    
    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    headers = ["ID", "Category", "Request Item", "Status", "Owner", "Data Room Location", "Notes"]
    ws.append(headers)
    
    data = [
        ("F-01", "Financial", "3-Year Projections", "Completed", "CEO", "01_Financials/Master_Model_V5.xlsx", ""),
        ("F-02", "Financial", "Bank Statements (Last 12m)", "Pending", "Finance", "-", "Waiting for accountant"),
        ("L-01", "Legal", "Cap Table", "Completed", "CEO", "01_Financials/Herramientas_V3.xlsx", ""),
        ("L-02", "Legal", "Incorporation Docs", "Completed", "Legal", "02_Legal/Documentacion_V2.docx", ""),
        ("T-01", "Tech", "Architecture Diagram", "Completed", "CTO", "05_Technical/Tech_Brief.docx", ""),
        ("T-02", "Tech", "Pentest Report", "In Progress", "CTO", "-", "Vendor scheduled for next week"),
        ("C-01", "Commercial", "Top 10 Customers Contracts", "Pending", "Sales", "-", "Redacting sensitive info"),
        ("H-01", "HR", "Key Employee Agreements", "Completed", "HR", "09_Talent/Key_Hires_JDs.docx", "")
    ]
    
    for row in data:
        ws.append(row)
        
    # Formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['F'].width = 40

    wb.save(filename)
    print(f"DD Tracker created: {filename}")

if __name__ == "__main__":
    create_dd_tracker("Ventura_Capital_DD_Request_Tracker.xlsx")


