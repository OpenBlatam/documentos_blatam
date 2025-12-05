#!/usr/bin/env python3
"""
Script ULTIMATE para convertir documentos con análisis avanzado, comparativas y dashboard HTML
"""

import os
import sys
import re
from pathlib import Path
from datetime import datetime
import json
from collections import Counter, defaultdict
import math

try:
    import numpy as np
except ImportError:
    print("Instalando numpy...")
    os.system("pip install numpy")
    import numpy as np

# Importar el convertidor mejorado anterior
sys.path.insert(0, str(Path(__file__).parent))
from convert_docs_to_formats_improved import AdvancedDocumentConverter

try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')
    import seaborn as sns
    sns.set_style("whitegrid")
    plt.rcParams['figure.figsize'] = (14, 10)
    plt.rcParams['font.size'] = 11
    plt.rcParams['font.family'] = 'DejaVu Sans'
except ImportError:
    print("Instalando matplotlib y seaborn...")
    os.system("pip install matplotlib seaborn")
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')
    import seaborn as sns

try:
    from docx import Document
    from docx.shared import Inches, Pt, RGBColor
    from docx.enum.text import WD_ALIGN_PARAGRAPH
except ImportError:
    os.system("pip install python-docx")
    from docx import Document
    from docx.shared import Inches, Pt, RGBColor

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference, ScatterChart
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

class UltimateDocumentConverter(AdvancedDocumentConverter):
    def __init__(self, output_dir="converted_docs"):
        super().__init__(output_dir)
        self.all_documents_stats = []
        
    def calculate_reading_time(self, word_count):
        """Calcula tiempo estimado de lectura (palabras por minuto = 200)"""
        minutes = word_count / 200
        hours = int(minutes // 60)
        mins = int(minutes % 60)
        return {"minutes": minutes, "hours": hours, "remaining_minutes": mins}
    
    def analyze_document_structure(self, sections):
        """Analiza la estructura jerárquica del documento"""
        structure = {
            "max_depth": max([s["level"] for s in sections] + [0]),
            "avg_section_length": sum(len(s["content"]) for s in sections) / max(len(sections), 1),
            "sections_by_level": Counter([s["level"] for s in sections]),
            "largest_section": max(sections, key=lambda x: len(x["content"])) if sections else None,
            "sections_with_code": [s for s in sections if s["code_blocks"] > 0],
            "sections_with_links": [s for s in sections if s["links"] > 0],
        }
        return structure
    
    def create_comparative_analysis(self, all_stats):
        """Crea análisis comparativo entre todos los documentos"""
        if len(all_stats) < 2:
            return None
        
        comparison = {
            "documents": [s["name"] for s in all_stats],
            "word_counts": [s["stats"]["total_words"] for s in all_stats],
            "section_counts": [s["stats"]["sections"] for s in all_stats],
            "readability_scores": [s["stats"]["readability_score"] for s in all_stats],
            "complexity_scores": [s["stats"]["complexity_score"] for s in all_stats],
            "code_blocks": [s["stats"]["code_blocks"] for s in all_stats],
        }
        return comparison
    
    def create_comparative_charts(self, comparison, output_name):
        """Crea gráficas comparativas entre documentos"""
        charts = []
        
        if not comparison:
            return charts
        
        # 1. Comparación de tamaño (palabras)
        fig, ax = plt.subplots(figsize=(12, 7))
        bars = ax.bar(comparison["documents"], comparison["word_counts"], 
                     color=['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6'][:len(comparison["documents"])],
                     edgecolor='black', linewidth=1.5)
        ax.set_title('Comparación de Tamaño de Documentos', fontsize=16, fontweight='bold', pad=20)
        ax.set_ylabel('Número de Palabras', fontsize=12, fontweight='bold')
        ax.set_xlabel('Documento', fontsize=12, fontweight='bold')
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        
        for bar, val in zip(bars, comparison["word_counts"]):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:,}', ha='center', va='bottom', fontweight='bold')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        chart_path = self.temp_dir / f"{output_name}_comparison_size.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        charts.append(("Comparación de Tamaño", chart_path))
        
        # 2. Comparación de métricas de calidad
        fig, ax = plt.subplots(figsize=(12, 7))
        x = np.arange(len(comparison["documents"]))
        width = 0.35
        
        bars1 = ax.bar(x - width/2, comparison["readability_scores"], width, 
                      label='Legibilidad', color='#2ecc71', edgecolor='black')
        bars2 = ax.bar(x + width/2, comparison["complexity_scores"], width,
                      label='Complejidad', color='#e74c3c', edgecolor='black')
        
        ax.set_title('Comparación de Métricas de Calidad', fontsize=16, fontweight='bold', pad=20)
        ax.set_ylabel('Score (0-100)', fontsize=12, fontweight='bold')
        ax.set_xlabel('Documento', fontsize=12, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(comparison["documents"], rotation=45, ha='right')
        ax.legend(fontsize=11)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_ylim(0, 100)
        
        plt.tight_layout()
        chart_path = self.temp_dir / f"{output_name}_comparison_quality.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        charts.append(("Comparación de Calidad", chart_path))
        
        # 3. Heatmap de comparación
        fig, ax = plt.subplots(figsize=(10, 6))
        metrics_data = {
            'Palabras': [w/max(comparison["word_counts"])*100 for w in comparison["word_counts"]],
            'Secciones': [s/max(comparison["section_counts"])*100 for s in comparison["section_counts"]],
            'Legibilidad': comparison["readability_scores"],
            'Complejidad': comparison["complexity_scores"],
            'Código': [c/max(comparison["code_blocks"])*100 if max(comparison["code_blocks"]) > 0 else [0]*len(comparison["code_blocks"]) for c in comparison["code_blocks"]],
        }
        
        data_matrix = np.array([
            metrics_data['Palabras'],
            metrics_data['Secciones'],
            metrics_data['Legibilidad'],
            metrics_data['Complejidad'],
            metrics_data['Código']
        ])
        
        im = ax.imshow(data_matrix, cmap='YlOrRd', aspect='auto')
        ax.set_xticks(np.arange(len(comparison["documents"])))
        ax.set_yticks(np.arange(len(metrics_data)))
        ax.set_xticklabels(comparison["documents"], rotation=45, ha='right')
        ax.set_yticklabels(list(metrics_data.keys()))
        ax.set_title('Heatmap Comparativo de Métricas', fontsize=16, fontweight='bold', pad=20)
        
        # Agregar valores en las celdas
        for i in range(len(metrics_data)):
            for j in range(len(comparison["documents"])):
                text = ax.text(j, i, f'{data_matrix[i, j]:.0f}',
                             ha="center", va="center", color="black", fontweight='bold')
        
        plt.colorbar(im, ax=ax)
        plt.tight_layout()
        chart_path = self.temp_dir / f"{output_name}_comparison_heatmap.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        charts.append(("Heatmap Comparativo", chart_path))
        
        return charts
    
    def create_dashboard_html(self, all_docs_data, output_name="Dashboard"):
        """Crea un dashboard HTML interactivo"""
        html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Dashboard de Análisis de Documentos</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@3.9.1/dist/chart.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            padding: 30px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }}
        h1 {{
            color: #2c3e50;
            text-align: center;
            margin-bottom: 10px;
            font-size: 2.5em;
        }}
        .subtitle {{
            text-align: center;
            color: #7f8c8d;
            margin-bottom: 30px;
            font-size: 1.1em;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }}
        .stat-card h3 {{
            font-size: 0.9em;
            opacity: 0.9;
            margin-bottom: 10px;
        }}
        .stat-card .value {{
            font-size: 2em;
            font-weight: bold;
        }}
        .charts-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(500px, 1fr));
            gap: 30px;
            margin-top: 30px;
        }}
        .chart-container {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .chart-container h3 {{
            color: #2c3e50;
            margin-bottom: 15px;
            text-align: center;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background: #667eea;
            color: white;
            font-weight: bold;
        }}
        tr:hover {{
            background: #f5f5f5;
        }}
        .document-section {{
            margin-top: 40px;
            padding: 20px;
            background: #f8f9fa;
            border-radius: 10px;
        }}
        .document-section h2 {{
            color: #2c3e50;
            margin-bottom: 20px;
            border-bottom: 3px solid #667eea;
            padding-bottom: 10px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Dashboard de Análisis de Documentos</h1>
        <p class="subtitle">Generado el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        
        <div class="stats-grid">
"""
        
        # Agregar estadísticas generales
        total_words = sum(d["stats"]["total_words"] for d in all_docs_data)
        total_sections = sum(d["stats"]["sections"] for d in all_docs_data)
        avg_readability = sum(d["stats"]["readability_score"] for d in all_docs_data) / len(all_docs_data)
        avg_complexity = sum(d["stats"]["complexity_score"] for d in all_docs_data) / len(all_docs_data)
        
        html_content += f"""
            <div class="stat-card">
                <h3>Total de Documentos</h3>
                <div class="value">{len(all_docs_data)}</div>
            </div>
            <div class="stat-card">
                <h3>Total de Palabras</h3>
                <div class="value">{total_words:,}</div>
            </div>
            <div class="stat-card">
                <h3>Total de Secciones</h3>
                <div class="value">{total_sections}</div>
            </div>
            <div class="stat-card">
                <h3>Legibilidad Promedio</h3>
                <div class="value">{avg_readability:.0f}/100</div>
            </div>
            <div class="stat-card">
                <h3>Complejidad Promedio</h3>
                <div class="value">{avg_complexity:.0f}/100</div>
            </div>
        </div>
        
        <div class="charts-grid">
            <div class="chart-container">
                <h3>Comparación de Tamaño</h3>
                <canvas id="sizeChart"></canvas>
            </div>
            <div class="chart-container">
                <h3>Métricas de Calidad</h3>
                <canvas id="qualityChart"></canvas>
            </div>
        </div>
"""
        
        # Tabla comparativa
        html_content += """
        <div class="document-section">
            <h2>📋 Tabla Comparativa de Documentos</h2>
            <table>
                <thead>
                    <tr>
                        <th>Documento</th>
                        <th>Palabras</th>
                        <th>Secciones</th>
                        <th>Legibilidad</th>
                        <th>Complejidad</th>
                        <th>Bloques Código</th>
                        <th>Enlaces</th>
                        <th>Tiempo Lectura</th>
                    </tr>
                </thead>
                <tbody>
"""
        
        for doc in all_docs_data:
            reading_time = self.calculate_reading_time(doc["stats"]["total_words"])
            time_str = f"{reading_time['hours']}h {reading_time['remaining_minutes']}m" if reading_time['hours'] > 0 else f"{reading_time['remaining_minutes']}m"
            
            html_content += f"""
                    <tr>
                        <td><strong>{doc['name']}</strong></td>
                        <td>{doc['stats']['total_words']:,}</td>
                        <td>{doc['stats']['sections']}</td>
                        <td>{doc['stats']['readability_score']}/100</td>
                        <td>{doc['stats']['complexity_score']}/100</td>
                        <td>{doc['stats']['code_blocks']}</td>
                        <td>{doc['stats']['links']}</td>
                        <td>{time_str}</td>
                    </tr>
"""
        
        html_content += """
                </tbody>
            </table>
        </div>
"""
        
        # Secciones individuales por documento
        for doc in all_docs_data:
            reading_time = self.calculate_reading_time(doc["stats"]["total_words"])
            time_str = f"{reading_time['hours']}h {reading_time['remaining_minutes']}m" if reading_time['hours'] > 0 else f"{reading_time['remaining_minutes']}m"
            
            html_content += f"""
        <div class="document-section">
            <h2>📄 {doc['name']}</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>Palabras</h3>
                    <div class="value">{doc['stats']['total_words']:,}</div>
                </div>
                <div class="stat-card">
                    <h3>Secciones</h3>
                    <div class="value">{doc['stats']['sections']}</div>
                </div>
                <div class="stat-card">
                    <h3>Legibilidad</h3>
                    <div class="value">{doc['stats']['readability_score']}/100</div>
                </div>
                <div class="stat-card">
                    <h3>Complejidad</h3>
                    <div class="value">{doc['stats']['complexity_score']}/100</div>
                </div>
                <div class="stat-card">
                    <h3>Tiempo Lectura</h3>
                    <div class="value">{time_str}</div>
                </div>
            </div>
        </div>
"""
        
        # JavaScript para gráficas
        doc_names = [d['name'] for d in all_docs_data]
        word_counts = [d['stats']['total_words'] for d in all_docs_data]
        readability_scores = [d['stats']['readability_score'] for d in all_docs_data]
        complexity_scores = [d['stats']['complexity_score'] for d in all_docs_data]
        
        html_content += f"""
        <script>
            // Gráfica de tamaño
            const sizeCtx = document.getElementById('sizeChart').getContext('2d');
            new Chart(sizeCtx, {{
                type: 'bar',
                data: {{
                    labels: {json.dumps(doc_names)},
                    datasets: [{{
                        label: 'Número de Palabras',
                        data: {json.dumps(word_counts)},
                        backgroundColor: ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6'],
                        borderColor: ['#2980b9', '#27ae60', '#c0392b', '#d68910', '#8e44ad'],
                        borderWidth: 2
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{
                        legend: {{ display: false }},
                        title: {{ display: false }}
                    }},
                    scales: {{
                        y: {{
                            beginAtZero: true
                        }}
                    }}
                }}
            }});
            
            // Gráfica de calidad
            const qualityCtx = document.getElementById('qualityChart').getContext('2d');
            new Chart(qualityCtx, {{
                type: 'bar',
                data: {{
                    labels: {json.dumps(doc_names)},
                    datasets: [{{
                        label: 'Legibilidad',
                        data: {json.dumps(readability_scores)},
                        backgroundColor: '#2ecc71',
                        borderColor: '#27ae60',
                        borderWidth: 2
                    }}, {{
                        label: 'Complejidad',
                        data: {json.dumps(complexity_scores)},
                        backgroundColor: '#e74c3c',
                        borderColor: '#c0392b',
                        borderWidth: 2
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{
                        title: {{ display: false }}
                    }},
                    scales: {{
                        y: {{
                            beginAtZero: true,
                            max: 100
                        }}
                    }}
                }}
            }});
        </script>
    </div>
</body>
</html>
"""
        
        output_path = self.output_dir / f"{output_name}.html"
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✓ Dashboard HTML guardado: {output_path}")
        return output_path
    
    def convert_all_documents_ultimate(self, documents):
        """Convierte todos los documentos con análisis comparativo"""
        print("=" * 70)
        print("CONVERSIÓN ULTIMATE CON ANÁLISIS COMPARATIVO")
        print("=" * 70)
        print()
        
        all_docs_data = []
        results = []
        
        # Procesar cada documento
        for doc in documents:
            if not doc["path"].exists():
                print(f"⚠ Archivo no encontrado: {doc['path']}")
                continue
            
            print(f"\n📄 Procesando: {doc['name']}")
            print("-" * 70)
            
            try:
                content = self.read_markdown(doc["path"])
                sections = self.parse_markdown_sections(content)
                stats = self.create_advanced_statistics(content, sections)
                
                # Calcular tiempo de lectura
                reading_time = self.calculate_reading_time(stats["total_words"])
                stats["reading_time"] = reading_time
                
                # Analizar estructura
                structure = self.analyze_document_structure(sections)
                stats["structure"] = structure
                
                all_docs_data.append({
                    "name": doc["name"],
                    "path": str(doc["path"]),
                    "stats": stats,
                    "sections": sections
                })
                
                # Crear gráficas individuales
                charts = self.create_professional_charts(stats, sections, doc["name"])
                
                # Convertir a formatos
                word_path = self.convert_to_word_improved(doc["path"], doc["name"])
                excel_path = self.convert_to_excel_improved(doc["path"], doc["name"])
                pdf_path = self.convert_to_pdf_improved(doc["path"], doc["name"])
                
                results.append({
                    "document": doc["name"],
                    "word": str(word_path),
                    "excel": str(excel_path),
                    "pdf": str(pdf_path)
                })
                
            except Exception as e:
                print(f"❌ Error procesando {doc['name']}: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # Crear análisis comparativo
        if len(all_docs_data) >= 2:
            print("\n" + "=" * 70)
            print("CREANDO ANÁLISIS COMPARATIVO")
            print("=" * 70)
            
            comparison = self.create_comparative_analysis(all_docs_data)
            if comparison:
                comp_charts = self.create_comparative_charts(comparison, "COMPARATIVE")
                
                # Crear documento comparativo en Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Comparativa"
                
                # Encabezados
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                headers = ["Documento", "Palabras", "Secciones", "Legibilidad", "Complejidad", "Código", "Tiempo Lectura"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Datos
                for row, doc in enumerate(all_docs_data, 2):
                    reading_time = doc["stats"]["reading_time"]
                    time_str = f"{reading_time['hours']}h {reading_time['remaining_minutes']}m" if reading_time['hours'] > 0 else f"{reading_time['remaining_minutes']}m"
                    
                    ws.cell(row=row, column=1).value = doc["name"]
                    ws.cell(row=row, column=2).value = doc["stats"]["total_words"]
                    ws.cell(row=row, column=3).value = doc["stats"]["sections"]
                    ws.cell(row=row, column=4).value = doc["stats"]["readability_score"]
                    ws.cell(row=row, column=5).value = doc["stats"]["complexity_score"]
                    ws.cell(row=row, column=6).value = doc["stats"]["code_blocks"]
                    ws.cell(row=row, column=7).value = time_str
                
                # Ajustar columnas
                for col in range(1, 8):
                    ws.column_dimensions[chr(64 + col)].width = 20
                
                comp_excel_path = self.output_dir / "COMPARATIVE_ANALYSIS.xlsx"
                wb.save(str(comp_excel_path))
                print(f"✓ Excel comparativo guardado: {comp_excel_path}")
        
        # Crear dashboard HTML
        print("\n" + "=" * 70)
        print("CREANDO DASHBOARD HTML")
        print("=" * 70)
        dashboard_path = self.create_dashboard_html(all_docs_data, "Dashboard_Analisis")
        
        # Resumen final
        print("\n" + "=" * 70)
        print("RESUMEN DE CONVERSIÓN ULTIMATE")
        print("=" * 70)
        print(f"\nDocumentos procesados: {len(results)}")
        print(f"Directorio de salida: {self.output_dir}")
        print("\nArchivos generados:")
        for result in results:
            print(f"\n  📄 {result['document']}:")
            print(f"     - Word: {result['word']}")
            print(f"     - Excel: {result['excel']}")
            print(f"     - PDF: {result['pdf']}")
        
        if len(all_docs_data) >= 2:
            print(f"\n  📊 Análisis Comparativo:")
            print(f"     - Excel: {comp_excel_path}")
            print(f"     - Dashboard HTML: {dashboard_path}")
        
        print("\n✅ Conversión ULTIMATE completada!")
        
        return results

def main():
    """Función principal"""
    base_dir = Path("/Users/adan/Documents/documentos_blatam")
    production_dir = base_dir / "truthgpt_collected/integration_code/production_code"
    
    documents = [
        {
            "path": base_dir / "airflow_automation_prompt.md",
            "name": "Automation_Expert_Prompt"
        },
        {
            "path": production_dir / "ARCHITECTURE_IMPROVEMENTS.md",
            "name": "Architecture_Improvements"
        },
        {
            "path": production_dir / "REFACTORING_PLAN.md",
            "name": "Refactoring_Plan"
        }
    ]
    
    converter = UltimateDocumentConverter()
    converter.convert_all_documents_ultimate(documents)

if __name__ == "__main__":
    main()

