import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def upgrade_excel_v5(filename):
    # Load V4
    try:
        wb = load_workbook("Ventura_Capital_Master_Model.xlsx")
    except FileNotFoundError:
        print("Error: V4 Master Model not found in root.")
        return

    # --- Styles ---
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center')

    # --- 1. Sensitivity Analysis ---
    if "Sensitivity" in wb.sheetnames:
        del wb["Sensitivity"]
    ws_sens = wb.create_sheet("Sensitivity Analysis")
    
    ws_sens['B2'] = "EBITDA Sensitivity to Growth & Churn (Year 3)"
    ws_sens['B2'].font = Font(bold=True, size=14)
    
    # Simulation Data (Mock Data Table)
    # Vertical: Churn Rate
    # Horizontal: Growth Rate
    
    data = [
        ["Churn \\ Growth", "10%", "15%", "20%", "25%", "30%"],
        ["3%", "$1.2M", "$1.5M", "$1.9M", "$2.4M", "$3.0M"],
        ["5% (Base)", "$0.8M", "$1.1M", "$1.6M", "$2.1M", "$2.6M"],
        ["7%", "$0.5M", "$0.8M", "$1.2M", "$1.7M", "$2.2M"],
        ["10%", "$0.1M", "$0.4M", "$0.7M", "$1.1M", "$1.6M"]
    ]
    
    start_row = 4
    start_col = 2
    
    for r_idx, row_data in enumerate(data):
        for c_idx, val in enumerate(row_data):
            cell = ws_sens.cell(row=start_row + r_idx, column=start_col + c_idx)
            cell.value = val
            cell.alignment = center_align
            
            if r_idx == 0 or c_idx == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # Highlight Base Case
            if val == "$1.6M" and r_idx == 2:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # --- 2. Use of Funds Waterfall ---
    # Simple table for waterfall chart creation
    ws_funds = wb.create_sheet("Use of Funds Detail")
    
    funds_data = [
        ["Category", "Amount", "% of Total"],
        ["Product Development (R&D)", 800000, 0.40],
        ["Sales & Marketing", 700000, 0.35],
        ["Operations & G&A", 300000, 0.15],
        ["Contingency", 200000, 0.10],
        ["Total Series A", 2000000, 1.00]
    ]
    
    for r in funds_data:
        ws_funds.append(r)
        
    # Format Funds Table
    for cell in ws_funds[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Currency Format
    for row in ws_funds.iter_rows(min_row=2, max_col=2):
        row[1].number_format = '$#,##0'
    for row in ws_funds.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].number_format = '0%'

    new_filename = "Ventura_Capital_Master_Model_V5.xlsx"
    wb.save(new_filename)
    print(f"Upgraded Excel V5 created: {new_filename}")

if __name__ == "__main__":
    upgrade_excel_v5("Ventura_Capital_Master_Model_V5.xlsx")

