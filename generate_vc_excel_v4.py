import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

def create_master_model(filename):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    workbook = writer.book

    # --- Styles ---
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Light Yellow for inputs
    calc_font = Font(name='Calibri', size=11, color='000000')
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), 
                         right=Side(style='thin', color='D9D9D9'), 
                         top=Side(style='thin', color='D9D9D9'), 
                         bottom=Side(style='thin', color='D9D9D9'))

    def format_sheet(ws):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
    
    # --- 1. Assumptions (Control Sheet) ---
    ws_assump = workbook.create_sheet("Assumptions")
    ws_assump.column_dimensions['A'].width = 30
    ws_assump.column_dimensions['B'].width = 15
    
    assumptions = [
        ("Metric", "Value", "Description"),
        ("Starting Customers", 500, "Current active users"),
        ("Monthly Growth Rate (Year 1)", 0.20, "20% MoM Growth"),
        ("Monthly Growth Rate (Year 2)", 0.15, "15% MoM Growth"),
        ("Monthly Growth Rate (Year 3)", 0.10, "10% MoM Growth"),
        ("Churn Rate", 0.05, "Monthly Churn"),
        ("ARPU (Average Rev Per User)", 29, "Weighted Avg Price"),
        ("CAC", 150, "Cost per Acq"),
        ("Gross Margin", 0.85, "SaaS Margin"),
        ("Fixed Costs (Monthly)", 50000, "Salaries, Rent, Server")
    ]
    
    for i, row in enumerate(assumptions, 1):
        ws_assump.cell(row=i, column=1, value=row[0])
        ws_assump.cell(row=i, column=2, value=row[1])
        ws_assump.cell(row=i, column=3, value=row[2])
        
        if i > 1: # Inputs
            ws_assump.cell(row=i, column=2).fill = input_fill
            if "Rate" in row[0] or "Margin" in row[0]:
                ws_assump.cell(row=i, column=2).number_format = '0.0%'
            else:
                ws_assump.cell(row=i, column=2).number_format = '#,##0'
                
    format_sheet(ws_assump)
    
    # --- 2. Monthly Model (Calculations) ---
    ws_model = workbook.create_sheet("Monthly Model")
    
    # Headers (Months 1-36)
    headers = ["Metric"] + [f"Month {i}" for i in range(1, 37)]
    ws_model.append(headers)
    
    metrics = ["Total Customers", "New Customers", "Churned Customers", "Revenue (MRR)", "COGS", "Gross Profit", "OpEx", "EBITDA"]
    
    # We will use formulas referring to Assumptions sheet
    # Simplified logic for script: Pre-calculate or write formulas
    # Writing formulas for Excel interactivity
    
    # Row indices
    r_cust = 2
    r_new = 3
    r_churn = 4
    r_rev = 5
    r_cogs = 6
    r_gp = 7
    r_opex = 8
    r_ebitda = 9
    
    # Initial setup
    ws_model.cell(row=r_cust, column=1, value="Total Customers")
    ws_model.cell(row=r_new, column=1, value="New Customers")
    ws_model.cell(row=r_churn, column=1, value="Churned Customers")
    ws_model.cell(row=r_rev, column=1, value="Revenue (MRR)")
    ws_model.cell(row=r_cogs, column=1, value="COGS")
    ws_model.cell(row=r_gp, column=1, value="Gross Profit")
    ws_model.cell(row=r_opex, column=1, value="OpEx")
    ws_model.cell(row=r_ebitda, column=1, value="EBITDA")

    # Month 1 Formulas
    ws_model.cell(row=r_cust, column=2, value="=Assumptions!B2") # Start
    ws_model.cell(row=r_rev, column=2, value=f"=B{r_cust}*Assumptions!B7")
    ws_model.cell(row=r_cogs, column=2, value=f"=B{r_rev}*(1-Assumptions!B9)")
    ws_model.cell(row=r_gp, column=2, value=f"=B{r_rev}-B{r_cogs}")
    ws_model.cell(row=r_opex, column=2, value="=Assumptions!B10")
    ws_model.cell(row=r_ebitda, column=2, value=f"=B{r_gp}-B{r_opex}")
    
    # Loop for Month 2-36
    for col in range(3, 38):
        prev_col = get_column_letter(col-1)
        curr_col = get_column_letter(col)
        
        # Growth Rate logic (Year 1, 2, 3)
        growth_cell = "Assumptions!B3" if col <= 13 else ("Assumptions!B4" if col <= 25 else "Assumptions!B5")
        
        # New Cust = Prev Total * Growth
        ws_model.cell(row=r_new, column=col, value=f"={prev_col}{r_cust}*{growth_cell}")
        
        # Churn = Prev Total * Churn Rate
        ws_model.cell(row=r_churn, column=col, value=f"={prev_col}{r_cust}*Assumptions!B6")
        
        # Total = Prev + New - Churn
        ws_model.cell(row=r_cust, column=col, value=f"={prev_col}{r_cust}+{curr_col}{r_new}-{curr_col}{r_churn}")
        
        # Revenue
        ws_model.cell(row=r_rev, column=col, value=f"={curr_col}{r_cust}*Assumptions!B7")
        
        # COGS
        ws_model.cell(row=r_cogs, column=col, value=f"={curr_col}{r_rev}*(1-Assumptions!B9)")
        
        # GP
        ws_model.cell(row=r_gp, column=col, value=f"={curr_col}{r_rev}-{curr_col}{r_cogs}")
        
        # OpEx (Simple growth 5% every 6 months?) Let's keep fixed for simplicity + 2% monthly inflation
        ws_model.cell(row=r_opex, column=col, value=f"={prev_col}{r_opex}*1.02")
        
        # EBITDA
        ws_model.cell(row=r_ebitda, column=col, value=f"={curr_col}{r_gp}-{curr_col}{r_opex}")

    format_sheet(ws_model)
    
    # Charts
    ws_charts = workbook.create_sheet("Visuals")
    chart_rev = LineChart()
    chart_rev.title = "MRR Growth Projection (36 Months)"
    chart_rev.y_axis.title = "$"
    chart_rev.x_axis.title = "Month"
    
    data = Reference(ws_model, min_col=2, min_row=5, max_col=37, max_row=5) # Revenue Row
    chart_rev.add_data(data, titles_from_data=False)
    ws_charts.add_chart(chart_rev, "B2")
    
    chart_ebitda = BarChart()
    chart_ebitda.title = "EBITDA Evolution"
    data_ebitda = Reference(ws_model, min_col=2, min_row=9, max_col=37, max_row=9)
    chart_ebitda.add_data(data_ebitda)
    ws_charts.add_chart(chart_ebitda, "B18")

    writer.close()
    print(f"Master Model created: {filename}")

if __name__ == "__main__":
    create_master_model("Ventura_Capital_Master_Model.xlsx")


