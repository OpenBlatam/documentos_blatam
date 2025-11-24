import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_roadmap_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Roadmap 2025"
    
    # Styles
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    
    # Columns
    headers = ["Quarter", "Theme", "Feature / Initiative", "Status", "Dev Effort (Weeks)", "Strategic Value"]
    ws.append(headers)
    
    data = [
        ("Q1 2025", "Market Expansion", "Portuguese Language Model V1", "In Progress", 8, "Unlock Brazil Market"),
        ("Q1 2025", "Growth", "Viral Watermark Removal (Paid)", "Planned", 2, "Increase Conversion"),
        ("Q2 2025", "Product Moat", "Proprietary Video Gen Alpha", "R&D", 12, "Differentiation vs Copy.ai"),
        ("Q2 2025", "Enterprise", "SSO & Role Based Access", "Planned", 4, "Unblock Enterprise Deals"),
        ("Q3 2025", "Integration", "Shopify / TiendaNube Plugin", "Backlog", 6, "Distribution Channel"),
        ("Q3 2025", "AI Core", "Context Window Expansion (32k)", "Backlog", 8, "Better Long-form Content"),
        ("Q4 2025", "Platform", "Mobile App iOS/Android", "Concept", 16, "Usage Frequency")
    ]
    
    for row in data:
        ws.append(row)
        
    # Formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
    # Conditional Formatting (Status)
    # Simplified for script: just setting column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25

    wb.save(filename)
    print(f"Roadmap Excel created: {filename}")

if __name__ == "__main__":
    create_roadmap_excel("Ventura_Capital_Product_Roadmap.xlsx")


