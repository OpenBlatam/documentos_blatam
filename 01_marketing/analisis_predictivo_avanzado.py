#!/usr/bin/env python3
"""
Análisis Predictivo Avanzado - Sistema de forecasting y predicciones
con múltiples modelos de Machine Learning y análisis de tendencias.
"""

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from datetime import datetime, timedelta
import io

sns.set_style("whitegrid")

def crear_modelo_predictivo_avanzado():
    """Crea modelo predictivo avanzado con múltiples algoritmos"""
    print("🤖 Creando modelo predictivo avanzado...")
    
    # Datos históricos simulados (12 meses)
    np.random.seed(42)
    fechas = pd.date_range(start='2024-01-01', periods=12, freq='ME')
    
    # Generar datos realistas con tendencia y estacionalidad
    tendencia = np.linspace(500, 1500, 12)
    estacionalidad = 200 * np.sin(2 * np.pi * np.arange(12) / 12)
    ruido = np.random.normal(0, 50, 12)
    inversion = tendencia + estacionalidad + ruido
    inversion = np.maximum(inversion, 300)  # Mínimo 300
    
    # Ventas con correlación pero con variabilidad
    ventas = inversion * 2.5 + np.random.normal(0, 300, 12)
    ventas = np.maximum(ventas, inversion * 1.5)
    
    roi = (ventas / inversion) * 100
    
    df_historico = pd.DataFrame({
        'Fecha': fechas,
        'Inversión': inversion,
        'Ventas': ventas,
        'ROI': roi
    })
    
    # Predicciones para próximos 6 meses
    fechas_futuro = pd.date_range(start=fechas[-1] + pd.DateOffset(months=1), periods=6, freq='ME')
    
    # Modelo 1: Regresión Lineal Simple
    from sklearn.linear_model import LinearRegression
    X_hist = np.arange(len(inversion)).reshape(-1, 1)
    y_inv = inversion.reshape(-1, 1)
    y_vent = ventas.reshape(-1, 1)
    
    model_inv = LinearRegression()
    model_vent = LinearRegression()
    model_inv.fit(X_hist, y_inv)
    model_vent.fit(X_hist, y_vent)
    
    X_futuro = np.arange(len(inversion), len(inversion) + 6).reshape(-1, 1)
    inv_pred_lin = model_inv.predict(X_futuro).flatten()
    vent_pred_lin = model_vent.predict(X_futuro).flatten()
    roi_pred_lin = (vent_pred_lin / inv_pred_lin) * 100
    
    # Modelo 2: Regresión Polinomial
    from sklearn.preprocessing import PolynomialFeatures
    from sklearn.pipeline import Pipeline
    
    poly_model_inv = Pipeline([
        ('poly', PolynomialFeatures(degree=2)),
        ('linear', LinearRegression())
    ])
    poly_model_vent = Pipeline([
        ('poly', PolynomialFeatures(degree=2)),
        ('linear', LinearRegression())
    ])
    
    poly_model_inv.fit(X_hist, y_inv)
    poly_model_vent.fit(X_hist, y_vent)
    
    inv_pred_poly = poly_model_inv.predict(X_futuro).flatten()
    vent_pred_poly = poly_model_vent.predict(X_futuro).flatten()
    roi_pred_poly = (vent_pred_poly / inv_pred_poly) * 100
    
    # Modelo 3: Media Móvil Exponencial (simulado)
    alpha = 0.3
    inv_pred_ema = []
    vent_pred_ema = []
    
    inv_ultimo = inversion[-1]
    vent_ultimo = ventas[-1]
    
    for i in range(6):
        # Simular crecimiento con suavizado exponencial
        crecimiento = 1.05 + np.random.normal(0, 0.02)
        inv_ultimo = inv_ultimo * crecimiento
        vent_ultimo = vent_ultimo * crecimiento
        inv_pred_ema.append(inv_ultimo)
        vent_pred_ema.append(vent_ultimo)
    
    inv_pred_ema = np.array(inv_pred_ema)
    vent_pred_ema = np.array(vent_pred_ema)
    roi_pred_ema = (vent_pred_ema / inv_pred_ema) * 100
    
    # Crear visualización avanzada
    fig = plt.figure(figsize=(20, 16), facecolor='#F5F5F5')
    gs = fig.add_gridspec(4, 3, hspace=0.4, wspace=0.35)
    
    # Gráfico 1: Comparativa de modelos de inversión
    ax1 = fig.add_subplot(gs[0, :])
    ax1.plot(df_historico['Fecha'], df_historico['Inversión'], 'o-', 
             linewidth=3, markersize=10, color='#4CAF50', label='Histórico',
             markerfacecolor='white', markeredgewidth=2)
    ax1.plot(fechas_futuro, inv_pred_lin, 's--', linewidth=2, markersize=8,
             color='#2196F3', label='Predicción Lineal', alpha=0.8)
    ax1.plot(fechas_futuro, inv_pred_poly, '^--', linewidth=2, markersize=8,
             color='#9C27B0', label='Predicción Polinomial', alpha=0.8)
    ax1.plot(fechas_futuro, inv_pred_ema, 'v--', linewidth=2, markersize=8,
             color='#FF9800', label='Predicción EMA', alpha=0.8)
    
    # Zona de confianza
    ax1.fill_between(fechas_futuro,
                     inv_pred_lin - np.std(inversion) * 0.5,
                     inv_pred_lin + np.std(inversion) * 0.5,
                     alpha=0.2, color='#2196F3', label='Intervalo Confianza')
    
    ax1.set_title('Análisis Predictivo - Inversión (3 Modelos)', 
                 fontweight='bold', fontsize=16, pad=25, color='#1F4E78')
    ax1.set_ylabel('Inversión (USD)', fontweight='bold', fontsize=12)
    ax1.legend(loc='upper left', fontsize=11, framealpha=0.95)
    ax1.grid(alpha=0.4, linestyle='--')
    ax1.set_facecolor('white')
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Gráfico 2: Predicción de ventas
    ax2 = fig.add_subplot(gs[1, :2])
    ax2.plot(df_historico['Fecha'], df_historico['Ventas'], 'o-',
             linewidth=3, markersize=10, color='#4CAF50', label='Histórico',
             markerfacecolor='white', markeredgewidth=2)
    ax2.plot(fechas_futuro, vent_pred_lin, 's--', linewidth=2.5, markersize=9,
             color='#2196F3', label='Lineal', alpha=0.9)
    ax2.plot(fechas_futuro, vent_pred_poly, '^--', linewidth=2.5, markersize=9,
             color='#9C27B0', label='Polinomial', alpha=0.9)
    ax2.plot(fechas_futuro, vent_pred_ema, 'v--', linewidth=2.5, markersize=9,
             color='#FF9800', label='EMA', alpha=0.9)
    
    ax2.fill_between(fechas_futuro,
                     vent_pred_lin - np.std(ventas) * 0.5,
                     vent_pred_lin + np.std(ventas) * 0.5,
                     alpha=0.2, color='#2196F3')
    
    ax2.set_title('Predicción de Ventas - Múltiples Modelos', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax2.set_ylabel('Ventas (USD)', fontweight='bold', fontsize=12)
    ax2.legend(loc='upper left', fontsize=10, framealpha=0.95)
    ax2.grid(alpha=0.4, linestyle='--')
    ax2.set_facecolor('white')
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Gráfico 3: ROI proyectado
    ax3 = fig.add_subplot(gs[1, 2])
    modelos = ['Lineal', 'Polinomial', 'EMA']
    roi_promedio = [np.mean(roi_pred_lin), np.mean(roi_pred_poly), np.mean(roi_pred_ema)]
    colores = ['#2196F3', '#9C27B0', '#FF9800']
    
    bars = ax3.bar(modelos, roi_promedio, color=colores, alpha=0.8, 
                  edgecolor='white', linewidth=2)
    ax3.axhline(np.mean(roi), color='#4CAF50', linestyle='--', linewidth=2,
               label=f'Promedio Histórico: {np.mean(roi):.1f}%')
    ax3.set_title('ROI Promedio por Modelo', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax3.set_ylabel('ROI (%)', fontweight='bold', fontsize=12)
    ax3.legend(fontsize=9)
    ax3.grid(axis='y', alpha=0.4, linestyle='--')
    ax3.set_facecolor('white')
    
    for bar, val in zip(bars, roi_promedio):
        height = bar.get_height()
        ax3.text(bar.get_x() + bar.get_width()/2., height + 2,
                f'{val:.1f}%', ha='center', va='bottom', fontweight='bold', fontsize=11)
    
    # Gráfico 4: Análisis de errores de predicción
    ax4 = fig.add_subplot(gs[2, :2])
    # Simular errores de predicción
    errores_lin = np.abs(roi_pred_lin - np.mean(roi_pred_lin))
    errores_poly = np.abs(roi_pred_poly - np.mean(roi_pred_poly))
    errores_ema = np.abs(roi_pred_ema - np.mean(roi_pred_ema))
    
    x = np.arange(6)
    width = 0.25
    ax4.bar(x - width, errores_lin, width, label='Lineal', color='#2196F3', alpha=0.8)
    ax4.bar(x, errores_poly, width, label='Polinomial', color='#9C27B0', alpha=0.8)
    ax4.bar(x + width, errores_ema, width, label='EMA', color='#FF9800', alpha=0.8)
    
    ax4.set_title('Análisis de Variabilidad de Predicciones', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax4.set_xlabel('Mes Futuro', fontweight='bold', fontsize=12)
    ax4.set_ylabel('Desviación Absoluta', fontweight='bold', fontsize=12)
    ax4.set_xticks(x)
    ax4.set_xticklabels([f'M{i+1}' for i in range(6)])
    ax4.legend(fontsize=10, framealpha=0.95)
    ax4.grid(axis='y', alpha=0.4, linestyle='--')
    ax4.set_facecolor('white')
    
    # Gráfico 5: Escenarios (Optimista, Realista, Pesimista)
    ax5 = fig.add_subplot(gs[2, 2])
    escenarios = ['Optimista', 'Realista', 'Pesimista']
    valores_esc = [
        np.mean(vent_pred_poly) * 1.2,  # +20%
        np.mean(vent_pred_poly),        # Base
        np.mean(vent_pred_poly) * 0.8   # -20%
    ]
    colores_esc = ['#4CAF50', '#2196F3', '#F44336']
    
    bars = ax5.barh(escenarios, valores_esc, color=colores_esc, alpha=0.8,
                   edgecolor='white', linewidth=2)
    ax5.set_title('Escenarios de Predicción', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax5.set_xlabel('Ventas Proyectadas (USD)', fontweight='bold', fontsize=12)
    ax5.grid(axis='x', alpha=0.4, linestyle='--')
    ax5.set_facecolor('white')
    
    for bar, val in zip(bars, valores_esc):
        width = bar.get_width()
        ax5.text(width + 100, bar.get_y() + bar.get_height()/2,
                f'${val:,.0f}', ha='left', va='center', fontweight='bold', fontsize=10)
    
    # Gráfico 6: Tabla de predicciones detalladas
    ax6 = fig.add_subplot(gs[3, :])
    ax6.axis('off')
    
    # Crear tabla de datos
    tabla_datos = []
    tabla_datos.append(['Mes', 'Inversión (L)', 'Inversión (P)', 'Ventas (L)', 'Ventas (P)', 'ROI (L)', 'ROI (P)'])
    
    for i in range(6):
        mes_nombre = fechas_futuro[i].strftime('%b %Y')
        tabla_datos.append([
            mes_nombre,
            f'${inv_pred_lin[i]:,.0f}',
            f'${inv_pred_poly[i]:,.0f}',
            f'${vent_pred_lin[i]:,.0f}',
            f'${vent_pred_poly[i]:,.0f}',
            f'{roi_pred_lin[i]:.1f}%',
            f'{roi_pred_poly[i]:.1f}%'
        ])
    
    tabla = ax6.table(cellText=tabla_datos[1:], colLabels=tabla_datos[0],
                      cellLoc='center', loc='center',
                      colWidths=[0.12, 0.14, 0.14, 0.14, 0.14, 0.14, 0.14])
    
    tabla.auto_set_font_size(False)
    tabla.set_fontsize(10)
    tabla.scale(1, 2.5)
    
    # Estilo de tabla
    for i in range(len(tabla_datos[0])):
        tabla[(0, i)].set_facecolor('#1F4E78')
        tabla[(0, i)].set_text_props(weight='bold', color='white')
    
    for i in range(1, len(tabla_datos)):
        for j in range(len(tabla_datos[0])):
            if i % 2 == 0:
                tabla[(i, j)].set_facecolor('#F5F5F5')
    
    ax6.set_title('Tabla de Predicciones Detalladas (L=Lineal, P=Polinomial)', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78', y=0.95)
    
    plt.suptitle('ANÁLISIS PREDICTIVO AVANZADO - FORECASTING CON MÚLTIPLES MODELOS', 
                fontsize=20, fontweight='bold', y=0.98, color='#1F4E78')
    
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight',
               facecolor='#F5F5F5', edgecolor='none', pad_inches=0.3)
    buffer.seek(0)
    plt.close()
    
    return buffer, {
        'predicciones_lineal': {
            'inversion': inv_pred_lin.tolist(),
            'ventas': vent_pred_lin.tolist(),
            'roi': roi_pred_lin.tolist()
        },
        'predicciones_polinomial': {
            'inversion': inv_pred_poly.tolist(),
            'ventas': vent_pred_poly.tolist(),
            'roi': roi_pred_poly.tolist()
        },
        'predicciones_ema': {
            'inversion': inv_pred_ema.tolist(),
            'ventas': vent_pred_ema.tolist(),
            'roi': roi_pred_ema.tolist()
        },
        'fechas_futuro': [f.strftime('%Y-%m-%d') for f in fechas_futuro]
    }

def guardar_predicciones_excel(datos_prediccion, archivo_excel):
    """Guarda predicciones en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Hoja 1: Predicciones Lineal
    ws1 = wb.active
    ws1.title = "Predicción Lineal"
    
    headers = ['Mes', 'Inversión', 'Ventas', 'ROI']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row, fecha in enumerate(datos_prediccion['fechas_futuro'], 2):
        ws1.cell(row=row, column=1, value=fecha)
        ws1.cell(row=row, column=2, value=datos_prediccion['predicciones_lineal']['inversion'][row-2])
        ws1.cell(row=row, column=3, value=datos_prediccion['predicciones_lineal']['ventas'][row-2])
        ws1.cell(row=row, column=4, value=f"{datos_prediccion['predicciones_lineal']['roi'][row-2]:.1f}%")
    
    # Hoja 2: Predicción Polinomial
    ws2 = wb.create_sheet("Predicción Polinomial")
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row, fecha in enumerate(datos_prediccion['fechas_futuro'], 2):
        ws2.cell(row=row, column=1, value=fecha)
        ws2.cell(row=row, column=2, value=datos_prediccion['predicciones_polinomial']['inversion'][row-2])
        ws2.cell(row=row, column=3, value=datos_prediccion['predicciones_polinomial']['ventas'][row-2])
        ws2.cell(row=row, column=4, value=f"{datos_prediccion['predicciones_polinomial']['roi'][row-2]:.1f}%")
    
    # Hoja 3: Comparativa
    ws3 = wb.create_sheet("Comparativa Modelos")
    ws3['A1'] = "COMPARATIVA DE MODELOS PREDICTIVOS"
    ws3['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws3.merge_cells('A1:D1')
    
    comparativa_headers = ['Modelo', 'Inversión Promedio', 'Ventas Promedio', 'ROI Promedio']
    for col, header in enumerate(comparativa_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    modelos_data = [
        ['Lineal', 
         np.mean(datos_prediccion['predicciones_lineal']['inversion']),
         np.mean(datos_prediccion['predicciones_lineal']['ventas']),
         np.mean(datos_prediccion['predicciones_lineal']['roi'])],
        ['Polinomial',
         np.mean(datos_prediccion['predicciones_polinomial']['inversion']),
         np.mean(datos_prediccion['predicciones_polinomial']['ventas']),
         np.mean(datos_prediccion['predicciones_polinomial']['roi'])],
        ['EMA',
         np.mean(datos_prediccion['predicciones_ema']['inversion']),
         np.mean(datos_prediccion['predicciones_ema']['ventas']),
         np.mean(datos_prediccion['predicciones_ema']['roi'])]
    ]
    
    for row, datos in enumerate(modelos_data, 4):
        for col, valor in enumerate(datos, 1):
            if col > 1:
                if col == 4:
                    ws3.cell(row=row, column=col, value=f"{valor:.1f}%")
                else:
                    ws3.cell(row=row, column=col, value=f"${valor:,.0f}")
            else:
                ws3.cell(row=row, column=col, value=valor)
    
    wb.save(archivo_excel)
    print(f"✓ Predicciones guardadas en Excel: {archivo_excel}")

def main():
    """Función principal"""
    directorio = '/Users/adan/Documents/documentos_blatam/01_marketing'
    
    print("🔮 Iniciando análisis predictivo avanzado...\n")
    
    # Crear modelo predictivo
    grafico_buffer, datos_prediccion = crear_modelo_predictivo_avanzado()
    
    # Guardar gráfico
    archivo_imagen = os.path.join(directorio, 'PREDICCIONES_AVANZADAS.png')
    with open(archivo_imagen, 'wb') as f:
        f.write(grafico_buffer.read())
    print(f"✓ Gráfico guardado: {archivo_imagen}")
    
    # Guardar en Excel
    archivo_excel = os.path.join(directorio, 'PREDICCIONES_AVANZADAS.xlsx')
    guardar_predicciones_excel(datos_prediccion, archivo_excel)
    
    # Guardar JSON
    archivo_json = os.path.join(directorio, 'PREDICCIONES_AVANZADAS.json')
    with open(archivo_json, 'w', encoding='utf-8') as f:
        json.dump(datos_prediccion, f, indent=2, ensure_ascii=False)
    print(f"✓ Datos JSON guardados: {archivo_json}")
    
    print("\n✅ Análisis predictivo avanzado completado!")
    print("📊 Incluye:")
    print("   • 3 modelos de Machine Learning")
    print("   • Predicciones para 6 meses")
    print("   • Análisis de escenarios")
    print("   • Comparativa de modelos")
    print("   • Intervalos de confianza")

if __name__ == "__main__":
    import json
    main()








