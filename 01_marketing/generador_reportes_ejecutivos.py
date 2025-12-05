#!/usr/bin/env python3
"""
Generador de Reportes Ejecutivos - Crea reportes ejecutivos automáticos
con resúmenes, métricas clave y visualizaciones.
"""

import os
import json
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

sns.set_style("whitegrid")

class GeneradorReportesEjecutivos:
    """Genera reportes ejecutivos automáticos"""
    
    def __init__(self):
        self.directorio = '/Users/adan/Documents/documentos_blatam/01_marketing'
        self.directorio_reportes = os.path.join(self.directorio, 'Reportes_Ejecutivos')
        os.makedirs(self.directorio_reportes, exist_ok=True)
    
    def generar_datos_ejemplo(self):
        """Genera datos de ejemplo para el reporte"""
        # Simular datos de los últimos 12 meses
        fechas = pd.date_range(start='2024-01-01', periods=12, freq='ME')
        
        np.random.seed(42)
        datos = {
            'Mes': [f.strftime('%b %Y') for f in fechas],
            'Inversión': np.random.uniform(5000, 15000, 12),
            'Ventas': np.random.uniform(15000, 40000, 12),
            'Clientes_Nuevos': np.random.randint(50, 200, 12),
            'Tasa_Conversion': np.random.uniform(2.5, 5.5, 12),
            'ROI': np.random.uniform(150, 350, 12),
            'CAC': np.random.uniform(20, 80, 12),
            'LTV': np.random.uniform(200, 600, 12)
        }
        
        df = pd.DataFrame(datos)
        df['ROI_Calculado'] = (df['Ventas'] / df['Inversión']) * 100
        df['Margen'] = df['Ventas'] - df['Inversión']
        
        return df
    
    def calcular_metricas_clave(self, df):
        """Calcula métricas clave del negocio"""
        metricas = {
            'total_inversion': df['Inversión'].sum(),
            'total_ventas': df['Ventas'].sum(),
            'promedio_ventas_mensual': df['Ventas'].mean(),
            'total_clientes': df['Clientes_Nuevos'].sum(),
            'roi_promedio': df['ROI_Calculado'].mean(),
            'cac_promedio': df['CAC'].mean(),
            'ltv_promedio': df['LTV'].mean(),
            'margen_total': df['Margen'].sum(),
            'tasa_conversion_promedio': df['Tasa_Conversion'].mean(),
            'crecimiento_ventas': ((df['Ventas'].iloc[-1] - df['Ventas'].iloc[0]) / df['Ventas'].iloc[0]) * 100,
            'mejor_mes_ventas': df.loc[df['Ventas'].idxmax(), 'Mes'],
            'peor_mes_ventas': df.loc[df['Ventas'].idxmin(), 'Mes']
        }
        
        return metricas
    
    def crear_visualizacion_ejecutiva(self, df, metricas):
        """Crea visualización ejecutiva"""
        fig = plt.figure(figsize=(20, 14), facecolor='#F5F5F5')
        gs = fig.add_gridspec(3, 3, hspace=0.4, wspace=0.35)
        
        # Gráfico 1: Ventas vs Inversión
        ax1 = fig.add_subplot(gs[0, :2])
        meses = df['Mes']
        x = range(len(meses))
        ax1.plot(x, df['Ventas'], 'o-', linewidth=3, markersize=10, 
                color='#4CAF50', label='Ventas', markerfacecolor='white', markeredgewidth=2)
        ax1.plot(x, df['Inversión'], 's-', linewidth=3, markersize=10,
                color='#2196F3', label='Inversión', markerfacecolor='white', markeredgewidth=2)
        ax1.fill_between(x, df['Ventas'], alpha=0.3, color='#4CAF50')
        ax1.fill_between(x, df['Inversión'], alpha=0.3, color='#2196F3')
        ax1.set_title('Evolución de Ventas e Inversión', fontweight='bold', 
                     fontsize=16, pad=25, color='#1F4E78')
        ax1.set_ylabel('Monto (USD)', fontweight='bold', fontsize=12)
        ax1.set_xticks(x)
        ax1.set_xticklabels(meses, rotation=45, ha='right')
        ax1.legend(fontsize=11, framealpha=0.95)
        ax1.grid(alpha=0.4, linestyle='--')
        ax1.set_facecolor('white')
        
        # Gráfico 2: ROI
        ax2 = fig.add_subplot(gs[0, 2])
        ax2.barh(['ROI Promedio'], [metricas['roi_promedio']], 
                color='#4CAF50', alpha=0.8, edgecolor='white', linewidth=2)
        ax2.set_title('ROI Promedio', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
        ax2.set_xlabel('ROI (%)', fontweight='bold', fontsize=12)
        ax2.grid(axis='x', alpha=0.4, linestyle='--')
        ax2.set_facecolor('white')
        ax2.text(metricas['roi_promedio'], 0, f"{metricas['roi_promedio']:.1f}%",
                ha='left', va='center', fontweight='bold', fontsize=14,
                bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
        
        # Gráfico 3: Clientes Nuevos
        ax3 = fig.add_subplot(gs[1, 0])
        ax3.bar(x, df['Clientes_Nuevos'], color='#9C27B0', alpha=0.8,
               edgecolor='white', linewidth=1.5)
        ax3.set_title('Clientes Nuevos por Mes', fontweight='bold', 
                     fontsize=14, pad=20, color='#1F4E78')
        ax3.set_ylabel('Clientes', fontweight='bold', fontsize=12)
        ax3.set_xticks(x)
        ax3.set_xticklabels(meses, rotation=45, ha='right')
        ax3.grid(axis='y', alpha=0.4, linestyle='--')
        ax3.set_facecolor('white')
        
        # Gráfico 4: Tasa de Conversión
        ax4 = fig.add_subplot(gs[1, 1])
        ax4.plot(x, df['Tasa_Conversion'], 'o-', linewidth=2.5, markersize=8,
                color='#FF9800', markerfacecolor='white', markeredgewidth=2)
        ax4.axhline(metricas['tasa_conversion_promedio'], color='red', 
                   linestyle='--', linewidth=2, label='Promedio')
        ax4.set_title('Tasa de Conversión', fontweight='bold', 
                     fontsize=14, pad=20, color='#1F4E78')
        ax4.set_ylabel('Tasa (%)', fontweight='bold', fontsize=12)
        ax4.set_xticks(x)
        ax4.set_xticklabels(meses, rotation=45, ha='right')
        ax4.legend(fontsize=10)
        ax4.grid(alpha=0.4, linestyle='--')
        ax4.set_facecolor('white')
        
        # Gráfico 5: CAC vs LTV
        ax5 = fig.add_subplot(gs[1, 2])
        ax5.bar(['CAC', 'LTV'], [metricas['cac_promedio'], metricas['ltv_promedio']],
               color=['#F44336', '#4CAF50'], alpha=0.8, edgecolor='white', linewidth=2)
        ax5.set_title('CAC vs LTV', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
        ax5.set_ylabel('Valor (USD)', fontweight='bold', fontsize=12)
        ax5.grid(axis='y', alpha=0.4, linestyle='--')
        ax5.set_facecolor('white')
        for i, (label, val) in enumerate(zip(['CAC', 'LTV'], 
                                             [metricas['cac_promedio'], metricas['ltv_promedio']])):
            ax5.text(i, val + 10, f'${val:.0f}', ha='center', va='bottom',
                    fontweight='bold', fontsize=11)
        
        # Gráfico 6: Métricas clave (tabla visual)
        ax6 = fig.add_subplot(gs[2, :])
        ax6.axis('off')
        
        metricas_texto = [
            ['Métrica', 'Valor'],
            ['Total Inversión', f"${metricas['total_inversion']:,.0f}"],
            ['Total Ventas', f"${metricas['total_ventas']:,.0f}"],
            ['Margen Total', f"${metricas['margen_total']:,.0f}"],
            ['Promedio Ventas Mensual', f"${metricas['promedio_ventas_mensual']:,.0f}"],
            ['Total Clientes Nuevos', f"{metricas['total_clientes']:,}"],
            ['ROI Promedio', f"{metricas['roi_promedio']:.1f}%"],
            ['Crecimiento Ventas', f"{metricas['crecimiento_ventas']:.1f}%"],
            ['Mejor Mes', metricas['mejor_mes_ventas']],
            ['Peor Mes', metricas['peor_mes_ventas']]
        ]
        
        tabla = ax6.table(cellText=metricas_texto[1:], colLabels=metricas_texto[0],
                         cellLoc='center', loc='center',
                         colWidths=[0.4, 0.3])
        tabla.auto_set_font_size(False)
        tabla.set_fontsize(11)
        tabla.scale(1, 2.5)
        
        for i in range(len(metricas_texto[0])):
            tabla[(0, i)].set_facecolor('#1F4E78')
            tabla[(0, i)].set_text_props(weight='bold', color='white')
        
        for i in range(1, len(metricas_texto)):
            for j in range(len(metricas_texto[0])):
                if i % 2 == 0:
                    tabla[(i, j)].set_facecolor('#F5F5F5')
        
        ax6.set_title('Métricas Clave del Período', fontweight='bold', 
                     fontsize=16, pad=20, color='#1F4E78', y=0.95)
        
        plt.suptitle('REPORTE EJECUTIVO - RESUMEN ANUAL', 
                    fontsize=20, fontweight='bold', y=0.98, color='#1F4E78')
        
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight',
                   facecolor='#F5F5F5', edgecolor='none', pad_inches=0.3)
        buffer.seek(0)
        plt.close()
        
        return buffer
    
    def crear_excel_ejecutivo(self, df, metricas, archivo_excel):
        """Crea Excel ejecutivo"""
        wb = Workbook()
        
        # Hoja 1: Resumen Ejecutivo
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        
        # Título
        ws1['A1'] = "REPORTE EJECUTIVO"
        ws1['A1'].font = Font(bold=True, size=18, color="1F4E78")
        ws1.merge_cells('A1:B1')
        
        ws1['A2'] = f"Período: {df['Mes'].iloc[0]} - {df['Mes'].iloc[-1]}"
        ws1['A2'].font = Font(size=12)
        ws1.merge_cells('A2:B2')
        
        # Métricas clave
        row = 4
        ws1[f'A{row}'] = "MÉTRICAS CLAVE"
        ws1[f'A{row}'].font = Font(bold=True, size=14, color="1F4E78")
        
        metricas_display = [
            ('Total Inversión', f"${metricas['total_inversion']:,.0f}"),
            ('Total Ventas', f"${metricas['total_ventas']:,.0f}"),
            ('Margen Total', f"${metricas['margen_total']:,.0f}"),
            ('Promedio Ventas Mensual', f"${metricas['promedio_ventas_mensual']:,.0f}"),
            ('Total Clientes Nuevos', f"{metricas['total_clientes']:,}"),
            ('ROI Promedio', f"{metricas['roi_promedio']:.1f}%"),
            ('Crecimiento Ventas', f"{metricas['crecimiento_ventas']:.1f}%"),
            ('CAC Promedio', f"${metricas['cac_promedio']:.2f}"),
            ('LTV Promedio', f"${metricas['ltv_promedio']:.2f}")
        ]
        
        for i, (label, valor) in enumerate(metricas_display, start=row+1):
            ws1[f'A{i}'] = label
            ws1[f'B{i}'] = valor
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'B{i}'].font = Font(size=11)
        
        # Hoja 2: Datos Mensuales
        ws2 = wb.create_sheet("Datos Mensuales")
        for col, header in enumerate(df.columns, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(archivo_excel)
        print(f"✓ Excel ejecutivo guardado: {archivo_excel}")

def main():
    """Función principal"""
    generador = GeneradorReportesEjecutivos()
    
    print("📊 Generando reporte ejecutivo...\n")
    
    # Generar datos
    df = generador.generar_datos_ejemplo()
    metricas = generador.calcular_metricas_clave(df)
    
    # Crear visualización
    grafico_buffer = generador.crear_visualizacion_ejecutiva(df, metricas)
    
    # Guardar gráfico
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_imagen = os.path.join(
        generador.directorio_reportes,
        f'REPORTE_EJECUTIVO_{timestamp}.png'
    )
    with open(archivo_imagen, 'wb') as f:
        f.write(grafico_buffer.read())
    print(f"✓ Gráfico guardado: {archivo_imagen}")
    
    # Crear Excel
    archivo_excel = os.path.join(
        generador.directorio_reportes,
        f'REPORTE_EJECUTIVO_{timestamp}.xlsx'
    )
    generador.crear_excel_ejecutivo(df, metricas, archivo_excel)
    
    print(f"\n✅ Reporte ejecutivo completado!")
    print(f"📁 Archivos guardados en: {generador.directorio_reportes}")

if __name__ == "__main__":
    import numpy as np
    main()








