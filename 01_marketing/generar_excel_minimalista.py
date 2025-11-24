#!/usr/bin/env python3
"""
Script para generar archivo Excel (.xlsx) con diseño minimalista y elegante
Variante 3: Minimalista y Elegante
Paleta: Negro (#0f172a), Blanco, Dorado (#fbbf24)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Colores de la paleta minimalista
COLOR_NEGRO = "0f172a"
COLOR_BLANCO = "FFFFFF"
COLOR_DORADO = "fbbf24"

def crear_excel_minimalista():
    """Crea un archivo Excel profesional con diseño minimalista y elegante"""
    
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== HOJA 1: PORTADA ==========
    ws_portada = wb.create_sheet("Portada", 0)
    
    # Configurar columnas
    ws_portada.column_dimensions['A'].width = 2
    ws_portada.column_dimensions['B'].width = 50
    ws_portada.column_dimensions['C'].width = 50
    ws_portada.column_dimensions['D'].width = 2
    
    # Configurar filas
    for i in range(1, 50):
        ws_portada.row_dimensions[i].height = 20
    
    ws_portada.row_dimensions[10].height = 40
    ws_portada.row_dimensions[12].height = 50
    ws_portada.row_dimensions[15].height = 30
    
    # Título principal
    cell_titulo = ws_portada['B12']
    cell_titulo.value = "Sistema de Creación de Contenido"
    cell_titulo.font = Font(name='Times New Roman', size=32, bold=True, color=COLOR_NEGRO)
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtítulo
    cell_subtitulo = ws_portada['B15']
    cell_subtitulo.value = "Variante 3: Minimalista y Elegante"
    cell_subtitulo.font = Font(name='Times New Roman', size=18, italic=True, color=COLOR_DORADO)
    cell_subtitulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Información del documento
    cell_info = ws_portada['B20']
    cell_info.value = f"Versión: 1.0 | Fecha: {datetime.now().strftime('%d de %B de %Y')}"
    cell_info.font = Font(name='Times New Roman', size=12, color=COLOR_NEGRO)
    cell_info.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fondo blanco para toda la hoja
    fill_blanco = PatternFill(start_color=COLOR_BLANCO, end_color=COLOR_BLANCO, fill_type='solid')
    for row in ws_portada.iter_rows(min_row=1, max_row=50, min_col=1, max_col=4):
        for cell in row:
            cell.fill = fill_blanco
    
    # ========== HOJA 2: RESUMEN EJECUTIVO ==========
    ws_resumen = wb.create_sheet("Resumen Ejecutivo", 1)
    
    # Configurar columnas
    ws_resumen.column_dimensions['A'].width = 3
    ws_resumen.column_dimensions['B'].width = 25
    ws_resumen.column_dimensions['C'].width = 15
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 20
    
    # Título de la hoja
    ws_resumen['B2'] = "Resumen Ejecutivo"
    ws_resumen['B2'].font = Font(name='Times New Roman', size=24, bold=True, color=COLOR_NEGRO)
    ws_resumen['B2'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Línea decorativa dorada
    ws_resumen['B3'] = "─" * 50
    ws_resumen['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # KPIs principales
    ws_resumen['B5'] = "KPIs Principales"
    ws_resumen['B5'].font = Font(name='Times New Roman', size=16, bold=True, color=COLOR_NEGRO)
    
    # Encabezados de tabla
    headers = ['Métrica', 'Valor', 'Objetivo', 'Estado']
    for col_idx, header in enumerate(headers, start=2):
        cell = ws_resumen.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=12, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Datos de ejemplo
    datos_kpis = [
        ['Contenidos Creados', 45, 50, '90%'],
        ['Tasa de Engagement', '8.5%', '10%', '85%'],
        ['Alcance Total', '125K', '150K', '83%'],
        ['Conversiones', 23, 30, '77%'],
    ]
    
    for row_idx, datos in enumerate(datos_kpis, start=8):
        for col_idx, valor in enumerate(datos, start=2):
            cell = ws_resumen.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=11, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            # Alternar fondo para legibilidad
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # ========== HOJA 3: ESTRUCTURA DEL SISTEMA ==========
    ws_estructura = wb.create_sheet("Estructura del Sistema", 2)
    
    ws_estructura.column_dimensions['A'].width = 3
    ws_estructura.column_dimensions['B'].width = 40
    ws_estructura.column_dimensions['C'].width = 60
    
    # Título
    ws_estructura['B2'] = "Estructura del Sistema de Creación de Contenido"
    ws_estructura['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_estructura['B3'] = "─" * 50
    ws_estructura['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Secciones principales
    secciones = [
        ['Goal', 'Definición del objetivo del Content Creator'],
        ['Format Rules', 'Reglas de formato para contenido bien estructurado'],
        ['Restrictions', 'Restricciones y límites del sistema'],
        ['Query Type', 'Tipos de consulta soportados'],
        ['Planning Rules', 'Reglas para planificación de contenido'],
        ['Output', 'Especificaciones de salida'],
        ['Personalization', 'Personalización del sistema'],
    ]
    
    row = 5
    for seccion, descripcion in secciones:
        # Título de sección
        cell_titulo = ws_estructura.cell(row=row, column=2)
        cell_titulo.value = seccion
        cell_titulo.font = Font(name='Times New Roman', size=14, bold=True, color=COLOR_NEGRO)
        cell_titulo.alignment = Alignment(horizontal='left', vertical='center')
        
        # Descripción
        cell_desc = ws_estructura.cell(row=row, column=3)
        cell_desc.value = descripcion
        cell_desc.font = Font(name='Times New Roman', size=11, color=COLOR_NEGRO)
        cell_desc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Borde inferior delgado
        cell_titulo.border = Border(bottom=Side(style='thin', color=COLOR_DORADO))
        cell_desc.border = Border(bottom=Side(style='thin', color=COLOR_DORADO))
        
        row += 3
    
    # ========== HOJA 4: TIPOS DE CONTENIDO ==========
    ws_tipos = wb.create_sheet("Tipos de Contenido", 3)
    
    ws_tipos.column_dimensions['A'].width = 3
    ws_tipos.column_dimensions['B'].width = 25
    ws_tipos.column_dimensions['C'].width = 70
    
    # Título
    ws_tipos['B2'] = "Tipos de Contenido Soportados"
    ws_tipos['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_tipos['B3'] = "─" * 50
    ws_tipos['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_tipos = ['Tipo', 'Descripción']
    for col_idx, header in enumerate(headers_tipos, start=2):
        cell = ws_tipos.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=12, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Tipos de contenido
    tipos_contenido = [
        ['Blog Articles', 'Artículos largos y detallados con secciones claras, markdown y headings'],
        ['Social Media Posts', 'Contenido conciso y atractivo optimizado para plataformas específicas'],
        ['Email Marketing', 'Mensajes claros con llamados a la acción fuertes'],
        ['Landing Pages', 'Copy persuasivo y enfocado en conversión'],
        ['Copywriting', 'Lenguaje persuasivo con propuestas de valor claras'],
        ['Product Descriptions', 'Descripciones detalladas con características y beneficios'],
        ['SEO Content', 'Contenido optimizado con keywords naturales'],
        ['Creative Writing', 'Escritura creativa siguiendo dirección del usuario'],
        ['Technical Documentation', 'Documentación técnica con ejemplos de código'],
        ['Content Strategy', 'Estrategia de contenido basada en brand guidelines'],
    ]
    
    for row_idx, (tipo, descripcion) in enumerate(tipos_contenido, start=6):
        # Tipo
        cell_tipo = ws_tipos.cell(row=row_idx, column=2)
        cell_tipo.value = tipo
        cell_tipo.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_NEGRO)
        cell_tipo.alignment = Alignment(horizontal='left', vertical='center')
        cell_tipo.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
        
        # Descripción
        cell_desc = ws_tipos.cell(row=row_idx, column=3)
        cell_desc.value = descripcion
        cell_desc.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
        cell_desc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_desc.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
        
        # Alternar fondo
        if row_idx % 2 == 0:
            cell_tipo.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
            cell_desc.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # ========== HOJA 5: MÉTRICAS Y ANÁLISIS ==========
    ws_metricas = wb.create_sheet("Métricas y Análisis", 4)
    
    ws_metricas.column_dimensions['A'].width = 3
    ws_metricas.column_dimensions['B'].width = 20
    ws_metricas.column_dimensions['C'].width = 15
    ws_metricas.column_dimensions['D'].width = 15
    ws_metricas.column_dimensions['E'].width = 15
    ws_metricas.column_dimensions['F'].width = 15
    
    # Título
    ws_metricas['B2'] = "Análisis de Métricas"
    ws_metricas['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_metricas['B3'] = "─" * 50
    ws_metricas['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Datos de ejemplo para gráfico
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio']
    contenidos = [35, 42, 38, 45, 48, 52]
    engagement = [7.2, 8.1, 7.8, 8.5, 9.1, 9.5]
    
    # Tabla de datos
    headers_metricas = ['Mes', 'Contenidos', 'Engagement %', 'Alcance', 'Conversiones', 'ROI %']
    for col_idx, header in enumerate(headers_metricas, start=2):
        cell = ws_metricas.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Datos
    for row_idx, mes in enumerate(meses, start=6):
        ws_metricas.cell(row=row_idx, column=2).value = mes
        ws_metricas.cell(row=row_idx, column=3).value = contenidos[row_idx - 6]
        ws_metricas.cell(row=row_idx, column=4).value = engagement[row_idx - 6]
        ws_metricas.cell(row=row_idx, column=5).value = contenidos[row_idx - 6] * 2500  # Alcance estimado
        ws_metricas.cell(row=row_idx, column=6).value = int(contenidos[row_idx - 6] * 0.5)  # Conversiones
        # ROI calculado
        ws_metricas.cell(row=row_idx, column=7).value = f"=((E{row_idx}*10-B{row_idx}*50)/B{row_idx}/50)*100"
        
        for col_idx in range(2, 8):
            cell = ws_metricas.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # Formato condicional para Engagement %
    engagement_range = f"D6:D{5+len(meses)}"
    ws_metricas.conditional_formatting.add(engagement_range,
        ColorScaleRule(start_type='num', start_value=7, start_color='FFE5E5',
                      mid_type='num', mid_value=8.5, mid_color='FFF4E5',
                      end_type='num', end_value=10, end_color='E5F5E5'))
    
    # Formato condicional para ROI
    roi_range = f"G6:G{5+len(meses)}"
    ws_metricas.conditional_formatting.add(roi_range,
        FormulaRule(formula=['G6>0'], stopIfTrue=True, fill=PatternFill(start_color='E5F5E5', end_color='E5F5E5', fill_type='solid')))
    
    # Gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Contenidos Creados por Mes"
    chart.y_axis.title = 'Cantidad'
    chart.x_axis.title = 'Mes'
    chart.height = 10
    chart.width = 15
    
    data = Reference(ws_metricas, min_col=3, min_row=5, max_row=5+len(meses))
    cats = Reference(ws_metricas, min_col=2, min_row=6, max_row=5+len(meses))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Estilo del gráfico (colores minimalistas)
    chart.graph_style = None
    ws_metricas.add_chart(chart, "B18")
    
    # Gráfico de líneas para Engagement
    chart2 = LineChart()
    chart2.title = "Tendencia de Engagement"
    chart2.y_axis.title = 'Engagement %'
    chart2.x_axis.title = 'Mes'
    chart2.height = 10
    chart2.width = 15
    
    data2 = Reference(ws_metricas, min_col=4, min_row=5, max_row=5+len(meses))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws_metricas.add_chart(chart2, "R18")
    
    # ========== HOJA 6: GLOSARIO ==========
    ws_glosario = wb.create_sheet("Glosario", 5)
    
    ws_glosario.column_dimensions['A'].width = 3
    ws_glosario.column_dimensions['B'].width = 30
    ws_glosario.column_dimensions['C'].width = 70
    
    # Título
    ws_glosario['B2'] = "Glosario de Términos"
    ws_glosario['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_glosario['B3'] = "─" * 50
    ws_glosario['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Términos
    terminos = [
        ['Content Creator', 'Profesional especializado en crear contenido estratégico y de alto rendimiento'],
        ['Brand Voice', 'Tono y estilo único de comunicación de una marca'],
        ['Content Brief', 'Documento que especifica los requisitos y objetivos del contenido'],
        ['Engagement', 'Interacción del público con el contenido (likes, comentarios, shares)'],
        ['Call-to-Action (CTA)', 'Invitación clara a realizar una acción específica'],
        ['SEO', 'Optimización para motores de búsqueda'],
        ['UGC', 'Contenido generado por usuarios'],
        ['Landing Page', 'Página web diseñada para convertir visitantes en clientes'],
    ]
    
    row = 5
    for termino, definicion in terminos:
        # Término
        cell_termino = ws_glosario.cell(row=row, column=2)
        cell_termino.value = termino
        cell_termino.font = Font(name='Times New Roman', size=12, bold=True, color=COLOR_NEGRO)
        cell_termino.alignment = Alignment(horizontal='left', vertical='top')
        
        # Definición
        cell_def = ws_glosario.cell(row=row, column=3)
        cell_def.value = definicion
        cell_def.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
        cell_def.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row += 2
    
    # ========== HOJA 7: FÓRMULAS AVANZADAS ==========
    ws_formulas = wb.create_sheet("Fórmulas Avanzadas", 6)
    
    ws_formulas.column_dimensions['A'].width = 3
    ws_formulas.column_dimensions['B'].width = 30
    ws_formulas.column_dimensions['C'].width = 50
    ws_formulas.column_dimensions['D'].width = 20
    
    # Título
    ws_formulas['B2'] = "Fórmulas Avanzadas y Ejemplos"
    ws_formulas['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_formulas['B3'] = "─" * 50
    ws_formulas['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_formulas = ['Categoría', 'Fórmula', 'Descripción', 'Ejemplo']
    for col_idx, header in enumerate(headers_formulas, start=2):
        cell = ws_formulas.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Fórmulas
    formulas_data = [
        ['Búsqueda', 'BUSCARV(valor;matriz;columna;0)', 'Busca un valor en la primera columna', '=BUSCARV(A2;Datos!A:B;2;0)'],
        ['Búsqueda', 'INDICE(matriz;fila;columna)', 'Devuelve valor de matriz', '=INDICE(A1:C10;5;2)'],
        ['Condicional', 'SUMAR.SI(rango;criterio;suma)', 'Suma con condición', '=SUMAR.SI(B:B;">100";C:C)'],
        ['Condicional', 'CONTAR.SI(rango;criterio)', 'Cuenta con condición', '=CONTAR.SI(A:A;"Aprobado")'],
        ['Lógica', 'SI(condición;verdadero;falso)', 'Lógica condicional', '=SI(A1>100;"Alto";"Bajo")'],
        ['Lógica', 'SI.ERROR(valor;valor_si_error)', 'Manejo de errores', '=SI.ERROR(A1/B1;"Error")'],
        ['Texto', 'CONCATENAR(texto1;texto2)', 'Une textos', '=CONCATENAR(A1;" ";B1)'],
        ['Texto', 'EXTRAE(texto;inicio;longitud)', 'Extrae parte del texto', '=EXTRAE(A1;1;5)'],
        ['Fecha', 'AÑO(fecha)', 'Extrae año', '=AÑO(HOY())'],
        ['Fecha', 'MES(fecha)', 'Extrae mes', '=MES(HOY())'],
        ['Estadística', 'PROMEDIO.SI(rango;criterio)', 'Promedio condicional', '=PROMEDIO.SI(A:A;">50")'],
        ['Estadística', 'DESVEST(rango)', 'Desviación estándar', '=DESVEST(A1:A100)'],
    ]
    
    for row_idx, (categoria, formula, descripcion, ejemplo) in enumerate(formulas_data, start=6):
        for col_idx, valor in enumerate([categoria, formula, descripcion, ejemplo], start=2):
            cell = ws_formulas.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # ========== HOJA 8: VALIDACIÓN DE DATOS ==========
    ws_validacion = wb.create_sheet("Validación de Datos", 7)
    
    ws_validacion.column_dimensions['A'].width = 3
    ws_validacion.column_dimensions['B'].width = 25
    ws_validacion.column_dimensions['C'].width = 15
    ws_validacion.column_dimensions['D'].width = 40
    
    # Título
    ws_validacion['B2'] = "Validación de Datos"
    ws_validacion['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_validacion['B3'] = "─" * 50
    ws_validacion['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_valid = ['Campo', 'Tipo Validación', 'Valores Permitidos', 'Descripción']
    for col_idx, header in enumerate(headers_valid, start=2):
        cell = ws_validacion.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Datos de validación
    validaciones = [
        ['Tipo de Contenido', 'Lista', 'Blog, Social Media, Email, Landing', 'Selección de tipo'],
        ['Estado', 'Lista', 'Borrador, Revisión, Publicado', 'Estado del contenido'],
        ['Prioridad', 'Lista', 'Alta, Media, Baja', 'Nivel de prioridad'],
        ['Fecha Publicación', 'Fecha', '>=HOY()', 'Fecha futura o hoy'],
        ['Engagement %', 'Decimal', '0-100', 'Porcentaje de engagement'],
        ['Alcance', 'Entero', '>=0', 'Número positivo'],
    ]
    
    for row_idx, (campo, tipo, valores, descripcion) in enumerate(validaciones, start=6):
        for col_idx, valor in enumerate([campo, tipo, valores, descripcion], start=2):
            cell = ws_validacion.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # Agregar ejemplo práctico de validación
    ws_validacion['B15'] = "Ejemplo Práctico"
    ws_validacion['B15'].font = Font(name='Times New Roman', size=14, bold=True, color=COLOR_NEGRO)
    
    # Crear validación de datos para ejemplo
    dv_tipo = DataValidation(type="list", formula1='"Blog,Social Media,Email,Landing"', allow_blank=True)
    dv_tipo.error = 'Selecciona un tipo válido'
    dv_tipo.errorTitle = 'Error de validación'
    ws_validacion.add_data_validation(dv_tipo)
    dv_tipo.add(f"B17")
    
    ws_validacion['B16'] = "Tipo:"
    ws_validacion['B16'].font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_NEGRO)
    ws_validacion['B17'] = "Selecciona..."
    ws_validacion['B17'].fill = PatternFill(start_color='FFF9E5', end_color='FFF9E5', fill_type='solid')
    
    dv_estado = DataValidation(type="list", formula1='"Borrador,Revisión,Publicado"', allow_blank=True)
    dv_estado.error = 'Selecciona un estado válido'
    dv_estado.errorTitle = 'Error de validación'
    ws_validacion.add_data_validation(dv_estado)
    dv_estado.add(f"C17")
    
    ws_validacion['C16'] = "Estado:"
    ws_validacion['C16'].font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_NEGRO)
    ws_validacion['C17'] = "Selecciona..."
    ws_validacion['C17'].fill = PatternFill(start_color='FFF9E5', end_color='FFF9E5', fill_type='solid')
    
    # ========== HOJA 9: CASOS DE USO ==========
    ws_casos = wb.create_sheet("Casos de Uso", 8)
    
    ws_casos.column_dimensions['A'].width = 3
    ws_casos.column_dimensions['B'].width = 25
    ws_casos.column_dimensions['C'].width = 70
    
    # Título
    ws_casos['B2'] = "Casos de Uso Prácticos"
    ws_casos['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_casos['B3'] = "─" * 50
    ws_casos['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    casos_uso = [
        ['Caso 1: Blog Article', 'Crear un artículo de blog de 2000 palabras sobre "Tendencias de Marketing Digital 2025" con secciones claras, citas y llamados a la acción'],
        ['Caso 2: Social Media', 'Generar 10 publicaciones para Instagram sobre un nuevo producto, optimizadas para engagement y con hashtags relevantes'],
        ['Caso 3: Email Marketing', 'Crear una secuencia de 5 emails de bienvenida para nuevos suscriptores con CTAs estratégicos'],
        ['Caso 4: Landing Page', 'Desarrollar copy persuasivo para landing page de producto SaaS con enfoque en conversión'],
        ['Caso 5: Product Description', 'Escribir descripciones detalladas de productos e-commerce con características, beneficios y puntos de venta'],
        ['Caso 6: SEO Content', 'Crear contenido optimizado para SEO sobre "Cómo mejorar el SEO técnico" con keywords naturales'],
        ['Caso 7: Technical Doc', 'Documentar API REST con ejemplos de código, endpoints y casos de uso'],
        ['Caso 8: Content Strategy', 'Desarrollar estrategia de contenido trimestral basada en brand guidelines y objetivos de negocio'],
    ]
    
    row = 5
    for caso, descripcion in casos_uso:
        # Caso
        cell_caso = ws_casos.cell(row=row, column=2)
        cell_caso.value = caso
        cell_caso.font = Font(name='Times New Roman', size=12, bold=True, color=COLOR_NEGRO)
        cell_caso.alignment = Alignment(horizontal='left', vertical='top')
        
        # Descripción
        cell_desc = ws_casos.cell(row=row, column=3)
        cell_desc.value = descripcion
        cell_desc.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
        cell_desc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row += 2
    
    # ========== HOJA 10: DASHBOARD INTERACTIVO ==========
    ws_dashboard = wb.create_sheet("Dashboard", 9)
    
    ws_dashboard.column_dimensions['A'].width = 3
    ws_dashboard.column_dimensions['B'].width = 20
    ws_dashboard.column_dimensions['C'].width = 20
    ws_dashboard.column_dimensions['D'].width = 20
    ws_dashboard.column_dimensions['E'].width = 20
    
    # Título
    ws_dashboard['B2'] = "Dashboard de Métricas"
    ws_dashboard['B2'].font = Font(name='Times New Roman', size=24, bold=True, color=COLOR_NEGRO)
    
    # KPIs destacados
    kpis = [
        ['Total Contenidos', '=SUM(\'Métricas y Análisis\'!C6:C11)'],
        ['Promedio Engagement', '=AVERAGE(\'Métricas y Análisis\'!D6:D11)'],
        ['Total Alcance', '=SUM(\'Métricas y Análisis\'!E6:E11)'],
        ['Total Conversiones', '=SUM(\'Métricas y Análisis\'!F6:F11)'],
    ]
    
    row = 5
    for kpi_nombre, kpi_formula in kpis:
        # Nombre KPI
        cell_nombre = ws_dashboard.cell(row=row, column=2)
        cell_nombre.value = kpi_nombre
        cell_nombre.font = Font(name='Times New Roman', size=12, color=COLOR_NEGRO)
        cell_nombre.alignment = Alignment(horizontal='left', vertical='center')
        
        # Valor KPI
        cell_valor = ws_dashboard.cell(row=row, column=3)
        cell_valor.value = kpi_formula
        cell_valor.font = Font(name='Times New Roman', size=16, bold=True, color=COLOR_DORADO)
        cell_valor.alignment = Alignment(horizontal='left', vertical='center')
        cell_valor.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        cell_valor.border = Border(
            left=Side(style='medium', color=COLOR_DORADO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
        
        row += 2
    
    # ========== HOJA 11: CHECKLIST DE CALIDAD ==========
    ws_checklist = wb.create_sheet("Checklist de Calidad", 10)
    
    ws_checklist.column_dimensions['A'].width = 3
    ws_checklist.column_dimensions['B'].width = 50
    ws_checklist.column_dimensions['C'].width = 15
    ws_checklist.column_dimensions['D'].width = 30
    
    # Título
    ws_checklist['B2'] = "Checklist de Calidad"
    ws_checklist['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_checklist['B3'] = "─" * 50
    ws_checklist['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_check = ['Item', 'Estado', 'Notas']
    for col_idx, header in enumerate(headers_check, start=2):
        cell = ws_checklist.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Items del checklist
    checklist_items = [
        ['El contenido comienza con un gancho o resumen', 'Pendiente', ''],
        ['No hay encabezados al inicio del contenido', 'Pendiente', ''],
        ['Secciones principales usan encabezados nivel 2', 'Pendiente', ''],
        ['Listas son planas, sin anidamiento', 'Pendiente', ''],
        ['Tablas usadas para comparaciones', 'Pendiente', ''],
        ['Citas incluidas correctamente', 'Pendiente', ''],
        ['Expresiones matemáticas en LaTeX', 'Pendiente', ''],
        ['Fuentes citadas con formato correcto', 'Pendiente', ''],
        ['No hay lenguaje de moralización', 'Pendiente', ''],
        ['No hay contenido con derechos de autor', 'Pendiente', ''],
        ['No hay emojis en el cuerpo', 'Pendiente', ''],
        ['Contenido termina con CTA o próximos pasos', 'Pendiente', ''],
        ['Tono apropiado para audiencia objetivo', 'Pendiente', ''],
        ['Optimizado para plataforma objetivo', 'Pendiente', ''],
        ['Formato consistente en todo el documento', 'Pendiente', ''],
    ]
    
    for row_idx, (item, estado, notas) in enumerate(checklist_items, start=6):
        for col_idx, valor in enumerate([item, estado, notas], start=2):
            cell = ws_checklist.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # ========== HOJA 12: CALENDARIO EDITORIAL ==========
    ws_calendario = wb.create_sheet("Calendario Editorial", 11)
    
    ws_calendario.column_dimensions['A'].width = 3
    ws_calendario.column_dimensions['B'].width = 12
    ws_calendario.column_dimensions['C'].width = 30
    ws_calendario.column_dimensions['D'].width = 20
    ws_calendario.column_dimensions['E'].width = 15
    ws_calendario.column_dimensions['F'].width = 15
    ws_calendario.column_dimensions['G'].width = 20
    
    # Título
    ws_calendario['B2'] = "Calendario Editorial"
    ws_calendario['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_calendario['B3'] = "─" * 50
    ws_calendario['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_cal = ['Fecha', 'Tipo', 'Título', 'Plataforma', 'Estado', 'Autor', 'Notas']
    for col_idx, header in enumerate(headers_cal, start=2):
        cell = ws_calendario.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Datos de ejemplo
    calendario_data = [
        ['2025-01-15', 'Blog', 'Tendencias Marketing 2025', 'Website', 'Publicado', 'Ana García', 'Alto engagement'],
        ['2025-01-18', 'Social Media', 'Lanzamiento Producto X', 'Instagram', 'Programado', 'Carlos López', 'Incluir hashtags'],
        ['2025-01-20', 'Email', 'Newsletter Semanal', 'Email', 'Borrador', 'María Ruiz', 'Revisar CTA'],
        ['2025-01-22', 'Blog', 'Guía SEO Completa', 'Website', 'Revisión', 'Ana García', 'Optimizar keywords'],
        ['2025-01-25', 'Social Media', 'Testimonial Cliente', 'LinkedIn', 'Programado', 'Carlos López', 'Incluir imagen'],
    ]
    
    for row_idx, datos in enumerate(calendario_data, start=6):
        for col_idx, valor in enumerate(datos, start=2):
            cell = ws_calendario.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # Formato condicional para Estado
    estado_range = f"F6:F{5+len(calendario_data)}"
    ws_calendario.conditional_formatting.add(estado_range,
        FormulaRule(formula=['F6="Publicado"'], stopIfTrue=True, fill=PatternFill(start_color='E5F5E5', end_color='E5F5E5', fill_type='solid')))
    ws_calendario.conditional_formatting.add(estado_range,
        FormulaRule(formula=['F6="Programado"'], stopIfTrue=True, fill=PatternFill(start_color='FFF4E5', end_color='FFF4E5', fill_type='solid')))
    ws_calendario.conditional_formatting.add(estado_range,
        FormulaRule(formula=['F6="Borrador"'], stopIfTrue=True, fill=PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')))
    
    # ========== HOJA 13: ANÁLISIS DE COMPETENCIA ==========
    ws_competencia = wb.create_sheet("Análisis Competencia", 12)
    
    ws_competencia.column_dimensions['A'].width = 3
    ws_competencia.column_dimensions['B'].width = 20
    ws_competencia.column_dimensions['C'].width = 15
    ws_competencia.column_dimensions['D'].width = 15
    ws_competencia.column_dimensions['E'].width = 15
    ws_competencia.column_dimensions['F'].width = 20
    
    # Título
    ws_competencia['B2'] = "Análisis de Competencia"
    ws_competencia['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_competencia['B3'] = "─" * 50
    ws_competencia['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    # Encabezados
    headers_comp = ['Competidor', 'Frecuencia Post', 'Engagement %', 'Alcance Promedio', 'Tipo Contenido', 'Fortalezas']
    for col_idx, header in enumerate(headers_comp, start=2):
        cell = ws_competencia.cell(row=5, column=col_idx)
        cell.value = header
        cell.font = Font(name='Times New Roman', size=11, bold=True, color=COLOR_BLANCO)
        cell.fill = PatternFill(start_color=COLOR_NEGRO, end_color=COLOR_NEGRO, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color=COLOR_NEGRO),
            right=Side(style='thin', color=COLOR_NEGRO),
            top=Side(style='thin', color=COLOR_NEGRO),
            bottom=Side(style='thin', color=COLOR_NEGRO)
        )
    
    # Datos de ejemplo
    competencia_data = [
        ['Competidor A', 'Diario', '8.5%', '50K', 'Educativo', 'Contenido de calidad'],
        ['Competidor B', '3x/semana', '6.2%', '35K', 'Promocional', 'Alta frecuencia'],
        ['Competidor C', '2x/semana', '9.1%', '75K', 'Mixto', 'Alto engagement'],
        ['Competidor D', 'Diario', '5.8%', '28K', 'Noticias', 'Actualidad constante'],
    ]
    
    for row_idx, datos in enumerate(competencia_data, start=6):
        for col_idx, valor in enumerate(datos, start=2):
            cell = ws_competencia.cell(row=row_idx, column=col_idx)
            cell.value = valor
            cell.font = Font(name='Times New Roman', size=10, color=COLOR_NEGRO)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin', color=COLOR_NEGRO),
                right=Side(style='thin', color=COLOR_NEGRO),
                top=Side(style='thin', color=COLOR_NEGRO),
                bottom=Side(style='thin', color=COLOR_NEGRO)
            )
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
    
    # Gráfico comparativo
    chart_comp = BarChart()
    chart_comp.type = "col"
    chart_comp.style = 10
    chart_comp.title = "Comparación de Engagement"
    chart_comp.y_axis.title = 'Engagement %'
    chart_comp.height = 8
    chart_comp.width = 12
    
    data_comp = Reference(ws_competencia, min_col=4, min_row=5, max_row=5+len(competencia_data))
    cats_comp = Reference(ws_competencia, min_col=2, min_row=6, max_row=5+len(competencia_data))
    chart_comp.add_data(data_comp, titles_from_data=True)
    chart_comp.set_categories(cats_comp)
    ws_competencia.add_chart(chart_comp, "B15")
    
    # ========== HOJA 14: PLANTILLAS DE CONTENIDO ==========
    ws_plantillas = wb.create_sheet("Plantillas", 13)
    
    ws_plantillas.column_dimensions['A'].width = 3
    ws_plantillas.column_dimensions['B'].width = 25
    ws_plantillas.column_dimensions['C'].width = 70
    
    # Título
    ws_plantillas['B2'] = "Plantillas de Contenido"
    ws_plantillas['B2'].font = Font(name='Times New Roman', size=20, bold=True, color=COLOR_NEGRO)
    
    # Línea decorativa
    ws_plantillas['B3'] = "─" * 50
    ws_plantillas['B3'].font = Font(name='Times New Roman', size=12, color=COLOR_DORADO)
    
    plantillas = [
        ['Blog Post', 'Título atractivo\n\n[Gancho inicial - 2-3 oraciones]\n\n## Sección Principal\n\n[Contenido detallado]\n\n## Otra Sección\n\n[Contenido]\n\n[Conclusión con CTA]'],
        ['Social Media', '[Gancho]\n\n[Contenido principal]\n\n[CTA]\n\n#hashtag1 #hashtag2'],
        ['Email', 'Asunto: [Tema]\n\nHola [Nombre],\n\n[Gancho]\n\n[Cuerpo del mensaje]\n\n[CTA]\n\nSaludos,\n[Firma]'],
        ['Landing Page', '[Headline]\n\n[Subheadline]\n\n[Beneficios clave]\n\n[Características]\n\n[Testimonios]\n\n[CTA principal]'],
    ]
    
    row = 5
    for plantilla, estructura in plantillas:
        # Nombre plantilla
        cell_nombre = ws_plantillas.cell(row=row, column=2)
        cell_nombre.value = plantilla
        cell_nombre.font = Font(name='Times New Roman', size=12, bold=True, color=COLOR_NEGRO)
        cell_nombre.alignment = Alignment(horizontal='left', vertical='top')
        
        # Estructura
        cell_estructura = ws_plantillas.cell(row=row, column=3)
        cell_estructura.value = estructura
        cell_estructura.font = Font(name='Courier New', size=9, color=COLOR_NEGRO)
        cell_estructura.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row += 4
    
    # ========== CONFIGURACIÓN GENERAL ==========
    # Configurar encabezados y pies de página para todas las hojas
    for sheet in wb.worksheets:
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_margins.left = 0.7
        sheet.page_margins.right = 0.7
        sheet.page_margins.top = 0.75
        sheet.page_margins.bottom = 0.75
        sheet.page_margins.header = 0.3
        sheet.page_margins.footer = 0.3
        
        # Encabezado
        sheet.oddHeader.center.text = "&[File]"
        
        # Pie de página
        sheet.oddFooter.center.text = "&P de &N"
    
    # Guardar archivo
    nombre_archivo = "Sistema_Creacion_Contenido_Minimalista.xlsx"
    ruta_completa = os.path.join(os.path.dirname(__file__), nombre_archivo)
    wb.save(ruta_completa)
    print(f"✅ Archivo Excel creado exitosamente: {ruta_completa}")
    
    return ruta_completa

if __name__ == "__main__":
    crear_excel_minimalista()

