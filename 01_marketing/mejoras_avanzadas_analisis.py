#!/usr/bin/env python3
"""
Script con mejoras avanzadas: análisis estadísticos, machine learning básico,
predicciones, y visualizaciones mejoradas.
"""

import os
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import io

sns.set_style("whitegrid")
sns.set_palette("husl")

def crear_analisis_estadistico_avanzado():
    """Crea análisis estadístico avanzado con predicciones"""
    print("📊 Creando análisis estadístico avanzado...")
    
    # Datos simulados más realistas
    np.random.seed(42)
    meses = pd.date_range(start='2024-01-01', periods=12, freq='M')
    inversion_base = np.array([500, 600, 700, 700, 800, 900, 1000, 1100, 1200, 1300, 1400, 1500])
    ruido = np.random.normal(0, 50, 12)
    inversion = inversion_base + ruido
    
    # Ventas con correlación positiva pero con variabilidad
    ventas = inversion * 3 + np.random.normal(0, 200, 12)
    ventas = np.maximum(ventas, inversion * 1.5)  # Mínimo 150% ROI
    
    roi = (ventas / inversion) * 100
    
    df = pd.DataFrame({
        'Mes': meses,
        'Inversión': inversion,
        'Ventas': ventas,
        'ROI': roi
    })
    
    # Análisis estadístico
    stats_dict = {
        'Inversión': {
            'Media': np.mean(inversion),
            'Mediana': np.median(inversion),
            'Desviación Estándar': np.std(inversion),
            'Mínimo': np.min(inversion),
            'Máximo': np.max(inversion),
            'Coeficiente Variación': (np.std(inversion) / np.mean(inversion)) * 100
        },
        'Ventas': {
            'Media': np.mean(ventas),
            'Mediana': np.median(ventas),
            'Desviación Estándar': np.std(ventas),
            'Mínimo': np.min(ventas),
            'Máximo': np.max(ventas),
            'Coeficiente Variación': (np.std(ventas) / np.mean(ventas)) * 100
        },
        'ROI': {
            'Media': np.mean(roi),
            'Mediana': np.median(roi),
            'Desviación Estándar': np.std(roi),
            'Mínimo': np.min(roi),
            'Máximo': np.max(roi),
            'Coeficiente Variación': (np.std(roi) / np.mean(roi)) * 100
        }
    }
    
    # Correlación
    correlacion = np.corrcoef(inversion, ventas)[0, 1]
    
    # Regresión lineal para predicción
    X = np.arange(len(inversion)).reshape(-1, 1)
    y_ventas = ventas.reshape(-1, 1)
    
    model = LinearRegression()
    model.fit(X, y_ventas)
    
    # Predicción próximos 3 meses
    X_future = np.arange(len(inversion), len(inversion) + 3).reshape(-1, 1)
    ventas_pred = model.predict(X_future)
    
    # Regresión polinomial para mejor ajuste
    poly_features = PolynomialFeatures(degree=2)
    X_poly = poly_features.fit_transform(X)
    model_poly = LinearRegression()
    model_poly.fit(X_poly, y_ventas)
    X_future_poly = poly_features.transform(X_future)
    ventas_pred_poly = model_poly.predict(X_future_poly)
    
    # Crear visualización avanzada
    fig = plt.figure(figsize=(20, 14), facecolor='#F5F5F5')
    gs = fig.add_gridspec(4, 4, hspace=0.4, wspace=0.35)
    
    # Gráfico 1: Serie temporal con predicción
    ax1 = fig.add_subplot(gs[0, :2])
    ax1.plot(df['Mes'], df['Ventas'], 'o-', linewidth=3, markersize=10, 
             color='#4CAF50', label='Ventas Real', markerfacecolor='white', markeredgewidth=2)
    ax1.plot(df['Mes'], df['Inversión'], 's--', linewidth=2, markersize=8, 
             color='#F44336', label='Inversión', alpha=0.7)
    
    # Predicciones
    meses_futuros = pd.date_range(start=meses[-1] + pd.DateOffset(months=1), periods=3, freq='M')
    ax1.plot(meses_futuros, ventas_pred.flatten(), '^:', linewidth=2, markersize=10, 
             color='#2196F3', label='Predicción Lineal', alpha=0.8)
    ax1.plot(meses_futuros, ventas_pred_poly.flatten(), 'v:', linewidth=2, markersize=10, 
             color='#9C27B0', label='Predicción Polinomial', alpha=0.8)
    
    ax1.fill_between(meses_futuros, 
                     ventas_pred.flatten() - np.std(ventas) * 0.5,
                     ventas_pred.flatten() + np.std(ventas) * 0.5,
                     alpha=0.2, color='#2196F3', label='Intervalo Confianza')
    
    ax1.set_title('Análisis Predictivo - Ventas e Inversión', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax1.set_ylabel('USD', fontweight='bold', fontsize=12)
    ax1.legend(loc='upper left', fontsize=10, framealpha=0.95)
    ax1.grid(alpha=0.4, linestyle='--')
    ax1.set_facecolor('white')
    
    # Gráfico 2: Distribución de ROI
    ax2 = fig.add_subplot(gs[0, 2:])
    ax2.hist(roi, bins=8, color='#4CAF50', alpha=0.7, edgecolor='white', linewidth=2)
    ax2.axvline(np.mean(roi), color='#F44336', linestyle='--', linewidth=3, label=f'Media: {np.mean(roi):.1f}%')
    ax2.axvline(np.median(roi), color='#2196F3', linestyle='--', linewidth=3, label=f'Mediana: {np.median(roi):.1f}%')
    ax2.set_title('Distribución de ROI', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax2.set_xlabel('ROI (%)', fontweight='bold', fontsize=12)
    ax2.set_ylabel('Frecuencia', fontweight='bold', fontsize=12)
    ax2.legend(fontsize=10)
    ax2.grid(axis='y', alpha=0.4, linestyle='--')
    ax2.set_facecolor('white')
    
    # Gráfico 3: Box plot comparativo
    ax3 = fig.add_subplot(gs[1, :2])
    data_box = [inversion, ventas]
    bp = ax3.boxplot(data_box, labels=['Inversión', 'Ventas'], patch_artist=True,
                     boxprops=dict(facecolor='#2196F3', alpha=0.7),
                     medianprops=dict(color='red', linewidth=2),
                     whiskerprops=dict(linewidth=2),
                     capprops=dict(linewidth=2))
    bp['boxes'][1].set_facecolor('#4CAF50')
    ax3.set_title('Análisis de Distribución - Box Plot', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax3.set_ylabel('USD', fontweight='bold', fontsize=12)
    ax3.grid(axis='y', alpha=0.4, linestyle='--')
    ax3.set_facecolor('white')
    
    # Gráfico 4: Correlación y regresión
    ax4 = fig.add_subplot(gs[1, 2:])
    ax4.scatter(inversion, ventas, s=200, alpha=0.7, color='#4CAF50', 
               edgecolors='white', linewidths=2, zorder=3)
    
    # Línea de regresión
    z = np.polyfit(inversion, ventas, 1)
    p = np.poly1d(z)
    ax4.plot(inversion, p(inversion), "r--", alpha=0.8, linewidth=3, 
            label=f'Regresión (R²={correlacion**2:.3f})', zorder=2)
    
    # Intervalo de confianza
    y_pred = p(inversion)
    se = np.sqrt(np.sum((ventas - y_pred)**2) / (len(ventas) - 2))
    ax4.fill_between(sorted(inversion), 
                     p(sorted(inversion)) - 1.96*se,
                     p(sorted(inversion)) + 1.96*se,
                     alpha=0.2, color='red', label='95% Confianza')
    
    ax4.set_xlabel('Inversión (USD)', fontweight='bold', fontsize=12)
    ax4.set_ylabel('Ventas (USD)', fontweight='bold', fontsize=12)
    ax4.set_title(f'Correlación Inversión-Ventas (r={correlacion:.3f})', 
                 fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax4.legend(loc='upper left', fontsize=10, framealpha=0.95)
    ax4.grid(alpha=0.4, linestyle='--')
    ax4.set_facecolor('white')
    
    # Gráfico 5: Tendencias estacionales
    ax5 = fig.add_subplot(gs[2, :2])
    meses_nombres = [m.strftime('%b') for m in meses]
    ax5.plot(meses_nombres, roi, 'o-', linewidth=3, markersize=12, 
            color='#9C27B0', markerfacecolor='white', markeredgewidth=2.5)
    ax5.fill_between(meses_nombres, roi, alpha=0.3, color='#9C27B0')
    ax5.axhline(np.mean(roi), color='#F44336', linestyle='--', linewidth=2, 
               label=f'Media: {np.mean(roi):.1f}%')
    ax5.set_title('Tendencia de ROI Mensual', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax5.set_ylabel('ROI (%)', fontweight='bold', fontsize=12)
    ax5.set_xlabel('Mes', fontweight='bold', fontsize=12)
    ax5.legend(fontsize=10)
    ax5.grid(alpha=0.4, linestyle='--')
    ax5.set_facecolor('white')
    plt.setp(ax5.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Gráfico 6: Análisis de varianza
    ax6 = fig.add_subplot(gs[2, 2:])
    categorias = ['Q1', 'Q2', 'Q3', 'Q4']
    q1_roi = np.mean(roi[:3])
    q2_roi = np.mean(roi[3:6])
    q3_roi = np.mean(roi[6:9])
    q4_roi = np.mean(roi[9:12])
    valores_trim = [q1_roi, q2_roi, q3_roi, q4_roi]
    
    bars = ax6.bar(categorias, valores_trim, color=['#FF9800', '#2196F3', '#4CAF50', '#9C27B0'],
                   alpha=0.8, edgecolor='white', linewidth=2)
    ax6.axhline(np.mean(roi), color='#F44336', linestyle='--', linewidth=2, 
               label=f'Media General: {np.mean(roi):.1f}%')
    ax6.set_title('ROI por Trimestre', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    ax6.set_ylabel('ROI Promedio (%)', fontweight='bold', fontsize=12)
    ax6.legend(fontsize=10)
    ax6.grid(axis='y', alpha=0.4, linestyle='--')
    ax6.set_facecolor('white')
    for bar, val in zip(bars, valores_trim):
        height = bar.get_height()
        ax6.text(bar.get_x() + bar.get_width()/2., height + 5,
                f'{val:.1f}%', ha='center', va='bottom', fontweight='bold', fontsize=11)
    
    # Gráfico 7: Heatmap de correlaciones
    ax7 = fig.add_subplot(gs[3, :2])
    df_corr = df[['Inversión', 'Ventas', 'ROI']].corr()
    im = ax7.imshow(df_corr, cmap='RdYlGn', aspect='auto', vmin=-1, vmax=1)
    ax7.set_xticks(range(len(df_corr.columns)))
    ax7.set_yticks(range(len(df_corr.columns)))
    ax7.set_xticklabels(df_corr.columns, fontsize=11, fontweight='bold')
    ax7.set_yticklabels(df_corr.columns, fontsize=11, fontweight='bold')
    ax7.set_title('Matriz de Correlación', fontweight='bold', fontsize=14, pad=20, color='#1F4E78')
    cbar = plt.colorbar(im, ax=ax7, fraction=0.046, pad=0.04)
    cbar.set_label('Correlación', fontweight='bold', fontsize=11)
    for i in range(len(df_corr.columns)):
        for j in range(len(df_corr.columns)):
            text = ax7.text(j, i, f'{df_corr.iloc[i, j]:.2f}',
                          ha="center", va="center", color="black" if abs(df_corr.iloc[i, j]) < 0.5 else "white",
                          fontweight='bold', fontsize=11)
    
    # Gráfico 8: Resumen estadístico
    ax8 = fig.add_subplot(gs[3, 2:])
    ax8.axis('off')
    
    stats_text = f"""
    ANÁLISIS ESTADÍSTICO COMPLETO
    
    INVERSIÓN:
    • Media: ${np.mean(inversion):,.0f}
    • Mediana: ${np.median(inversion):,.0f}
    • Desv. Estándar: ${np.std(inversion):,.0f}
    • CV: {(np.std(inversion)/np.mean(inversion)*100):.1f}%
    
    VENTAS:
    • Media: ${np.mean(ventas):,.0f}
    • Mediana: ${np.median(ventas):,.0f}
    • Desv. Estándar: ${np.std(ventas):,.0f}
    • CV: {(np.std(ventas)/np.mean(ventas)*100):.1f}%
    
    ROI:
    • Media: {np.mean(roi):.1f}%
    • Mediana: {np.median(roi):.1f}%
    • Desv. Estándar: {np.std(roi):.1f}%
    • Rango: {np.min(roi):.1f}% - {np.max(roi):.1f}%
    
    CORRELACIÓN:
    • Inversión-Ventas: {correlacion:.3f}
    • R²: {correlacion**2:.3f}
    
    PREDICCIÓN (Próximos 3 meses):
    • Ventas esperadas: ${np.mean(ventas_pred):,.0f}
    • ROI esperado: {np.mean(roi):.1f}%
    """
    
    ax8.text(0.1, 0.5, stats_text, fontsize=11, fontfamily='monospace',
            verticalalignment='center', bbox=dict(boxstyle='round', 
            facecolor='white', alpha=0.9, edgecolor='#1F4E78', linewidth=2))
    
    plt.suptitle('ANÁLISIS ESTADÍSTICO AVANZADO Y PREDICTIVO', 
                fontsize=20, fontweight='bold', y=0.98, color='#1F4E78')
    
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight', 
               facecolor='#F5F5F5', edgecolor='none', pad_inches=0.3)
    buffer.seek(0)
    plt.close()
    
    return buffer, stats_dict

def guardar_analisis_excel(stats_dict, archivo_excel):
    """Guarda análisis estadístico en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Análisis Estadístico"
    
    # Título
    ws['A1'] = "ANÁLISIS ESTADÍSTICO AVANZADO"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells('A1:D1')
    
    row = 3
    for metrica, valores in stats_dict.items():
        ws.cell(row=row, column=1, value=metrica).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        for stat, valor in valores.items():
            ws.cell(row=row, column=1, value=stat).font = Font(bold=True)
            if isinstance(valor, float):
                ws.cell(row=row, column=2, value=f"{valor:.2f}")
            else:
                ws.cell(row=row, column=2, value=valor)
            row += 1
        row += 1
    
    wb.save(archivo_excel)
    print(f"✓ Análisis guardado en: {archivo_excel}")

def main():
    """Función principal"""
    directorio = '/Users/adan/Documents/documentos_blatam/01_marketing'
    
    print("🔬 Iniciando análisis estadístico avanzado...\n")
    
    # Crear análisis
    grafico_buffer, stats_dict = crear_analisis_estadistico_avanzado()
    
    # Guardar gráfico
    archivo_imagen = os.path.join(directorio, 'ANALISIS_ESTADISTICO_AVANZADO.png')
    with open(archivo_imagen, 'wb') as f:
        f.write(grafico_buffer.read())
    print(f"✓ Gráfico guardado: {archivo_imagen}")
    
    # Guardar en Excel
    archivo_excel = os.path.join(directorio, 'ANALISIS_ESTADISTICO_AVANZADO.xlsx')
    guardar_analisis_excel(stats_dict, archivo_excel)
    
    print("\n✅ Análisis estadístico avanzado completado!")
    print("📊 Incluye:")
    print("   • Análisis descriptivo completo")
    print("   • Correlaciones y regresiones")
    print("   • Predicciones con ML")
    print("   • Análisis de distribución")
    print("   • Intervalos de confianza")

if __name__ == "__main__":
    main()



