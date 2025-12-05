"""
ExcelProcessor - Procesador profesional para documentos Excel
"""

import os
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from ..config import ConfigManager
from ..logger import LoggerManager

class ExcelProcessor:
    """Procesador avanzado para documentos Excel"""
    
    def __init__(self, config: ConfigManager, logger: LoggerManager):
        self.config = config
        self.logger = logger.get_logger()
        self.excel_config = config.config.excel
    
    def convert(self, parsed_content: Dict[str, Any],
                base_name: str, output_dir: str) -> str:
        """Convierte contenido parseado a Excel"""
        wb = Workbook()
        
        # Eliminar hoja por defecto si hay contenido estructurado
        if 'sections' in parsed_content and len(parsed_content['sections']) > 0:
            wb.remove(wb.active)
        
        # Procesar secciones
        for i, section in enumerate(parsed_content.get('sections', []), 1):
            ws = wb.create_sheet(title=section.get('title', f'Hoja{i}')[:31])
            self._process_sheet(ws, section)
        
        # Si no hay secciones, usar contenido general
        if len(wb.worksheets) == 0:
            ws = wb.active
            ws.title = base_name[:31]
            self._process_general_content(ws, parsed_content)
        
        # Guardar
        output_file = os.path.join(output_dir, f"{base_name}.xlsx")
        wb.save(output_file)
        
        self.logger.info(f"Documento Excel creado: {output_file}")
        return output_file
    
    def _process_sheet(self, ws, section: Dict[str, Any]):
        """Procesa una hoja de Excel"""
        row = 1
        
        # Título
        if 'title' in section:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = section['title']
            cell.font = Font(bold=True, size=self.excel_config.header_font_size + 2,
                           color=self.excel_config.header_text_color)
            cell.fill = PatternFill(start_color=self.excel_config.header_bg_color,
                                   end_color=self.excel_config.header_bg_color,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 2
        
        # Tablas
        if 'tables' in section:
            for table_data in section['tables']:
                self._add_table(ws, table_data, row)
                row += len(table_data) + 2
        
        # Contenido de texto
        if 'content' in section:
            self._add_text_content(ws, section['content'], row)
        
        # Ajustar ancho de columnas
        if self.excel_config.auto_width:
            self._auto_adjust_column_width(ws)
        
        # Formato condicional
        if self.excel_config.conditional_formatting:
            self._apply_conditional_formatting(ws)
    
    def _process_general_content(self, ws, parsed_content: Dict[str, Any]):
        """Procesa contenido general"""
        row = 1
        
        if 'title' in parsed_content:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = parsed_content['title']
            cell.font = Font(bold=True, size=self.excel_config.header_font_size + 2)
            row += 2
        
        if 'tables' in parsed_content:
            for table_data in parsed_content['tables']:
                self._add_table(ws, table_data, row)
                row += len(table_data) + 2
    
    def _add_table(self, ws, table_data: List[List[str]], start_row: int):
        """Agrega tabla a Excel"""
        if not table_data:
            return
        
        # Encabezados
        header_fill = PatternFill(start_color=self.excel_config.header_bg_color,
                                 end_color=self.excel_config.header_bg_color,
                                 fill_type='solid')
        header_font = Font(bold=self.excel_config.header_bold,
                          size=self.excel_config.header_font_size,
                          color=self.excel_config.header_text_color)
        
        for col_idx, header in enumerate(table_data[0], 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = str(header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, row_data in enumerate(table_data[1:], start_row + 1):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_data
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Auto filtro
        if self.excel_config.auto_filter:
            ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(table_data[0]))}{start_row + len(table_data) - 1}"
        
        # Congelar paneles
        if self.excel_config.freeze_panes:
            ws.freeze_panes = f"A{start_row + 1}"
    
    def _add_text_content(self, ws, content: str, start_row: int):
        """Agrega contenido de texto"""
        lines = content.split('\n')
        for i, line in enumerate(lines, start_row):
            if line.strip():
                ws.cell(row=i, column=1, value=line.strip())
    
    def _auto_adjust_column_width(self, ws):
        """Ajusta automáticamente el ancho de columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_conditional_formatting(self, ws):
        """Aplica formato condicional"""
        # Buscar columnas numéricas
        for col in range(1, ws.max_column + 1):
            try:
                # Verificar si la columna tiene números
                has_numbers = False
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        has_numbers = True
                        break
                
                if has_numbers:
                    # Aplicar escala de colores
                    range_str = f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}"
                    ws.conditional_formatting.add(
                        range_str,
                        ColorScaleRule(
                            start_type='min', start_color='FFFFFF',
                            end_type='max', end_color='4CAF50'
                        )
                    )
            except:
                pass








