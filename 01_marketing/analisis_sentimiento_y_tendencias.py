#!/usr/bin/env python3
"""
Análisis de Sentimiento y Tendencias - Analiza texto y genera insights
sobre sentimientos, tendencias y patrones en contenido de marketing.
"""

import os
import re
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from datetime import datetime
from collections import Counter
import io

sns.set_style("whitegrid")

class AnalizadorSentimiento:
    """Analiza sentimiento y tendencias en texto"""
    
    def __init__(self):
        # Diccionarios de sentimiento (simplificado)
        self.palabras_positivas = {
            'excelente', 'genial', 'fantástico', 'increíble', 'perfecto',
            'bueno', 'mejor', 'éxito', 'ganar', 'crecer', 'aumentar',
            'feliz', 'satisfecho', 'recomendar', 'amor', 'me gusta'
        }
        
        self.palabras_negativas = {
            'malo', 'terrible', 'horrible', 'problema', 'error', 'fallo',
            'decepcionado', 'frustrado', 'difícil', 'complicado', 'lento',
            'caro', 'perder', 'disminuir', 'rechazar', 'no me gusta'
        }
        
        self.palabras_neutrales = {
            'normal', 'regular', 'promedio', 'estándar', 'típico'
        }
    
    def analizar_sentimiento(self, texto):
        """Analiza el sentimiento de un texto"""
        texto_lower = texto.lower()
        palabras = re.findall(r'\b\w+\b', texto_lower)
        
        positivo = sum(1 for p in palabras if p in self.palabras_positivas)
        negativo = sum(1 for p in palabras if p in self.palabras_negativas)
        neutral = sum(1 for p in palabras if p in self.palabras_neutrales)
        total = len(palabras)
        
        if total == 0:
            return {'sentimiento': 'neutral', 'score': 0.0, 'positivo': 0, 'negativo': 0, 'neutral': 0}
        
        score = (positivo - negativo) / total
        
        if score > 0.1:
            sentimiento = 'positivo'
        elif score < -0.1:
            sentimiento = 'negativo'
        else:
            sentimiento = 'neutral'
        
        return {
            'sentimiento': sentimiento,
            'score': score,
            'positivo': positivo,
            'negativo': negativo,
            'neutral': neutral,
            'total_palabras': total
        }
    
    def extraer_tendencias(self, textos):
        """Extrae tendencias y temas principales"""
        todas_palabras = []
        for texto in textos:
            palabras = re.findall(r'\b\w{4,}\b', texto.lower())
            todas_palabras.extend(palabras)
        
        # Filtrar palabras comunes
        palabras_comunes = {'este', 'esta', 'estos', 'estas', 'para', 'con', 'desde', 'hasta'}
        palabras_filtradas = [p for p in todas_palabras if p not in palabras_comunes]
        
        contador = Counter(palabras_filtradas)
        top_tendencias = contador.most_common(20)
        
        return top_tendencias
    
    def analizar_multiple(self, textos):
        """Analiza múltiples textos"""
        resultados = []
        for texto in textos:
            resultado = self.analizar_sentimiento(texto)
            resultados.append(resultado)
        
        return resultados

def crear_analisis_sentimiento_completo():
    """Crea análisis completo de sentimiento y tendencias"""
    print("📊 Creando análisis de sentimiento y tendencias...")
    
    # Textos de ejemplo (simulando contenido de marketing)
    textos_ejemplo = [
        "Nuestro producto es excelente y ha tenido un éxito increíble en el mercado. Los clientes están muy satisfechos.",
        "La campaña de marketing fue genial y generó un crecimiento significativo en las ventas.",
        "Tenemos algunos problemas con la entrega que necesitamos resolver. Los clientes están frustrados.",
        "El servicio es bueno pero podría mejorar en algunos aspectos. Es aceptable.",
        "Increíble experiencia con este producto. Lo recomendaría sin dudar. Me encanta.",
        "La calidad es terrible y el precio es muy caro. No estoy satisfecho.",
        "Excelente atención al cliente y productos de alta calidad. Muy recomendable.",
        "El proceso es complicado y lento. Necesita mejoras urgentes.",
        "Fantástico resultado. Hemos logrado todos nuestros objetivos de manera perfecta.",
        "Regular, nada especial. Cumple con lo básico pero no destaca."
    ]
    
    analizador = AnalizadorSentimiento()
    resultados = analizador.analizar_multiple(textos_ejemplo)
    tendencias = analizador.extraer_tendencias(textos_ejemplo)
    
    # Crear DataFrame
    df = pd.DataFrame(resultados)
    df['texto'] = [t[:50] + '...' for t in textos_ejemplo]
    
    # Estadísticas
    total_positivo = sum(1 for r in resultados if r['sentimiento'] == 'positivo')
    total_negativo = sum(1 for r in resultados if r['sentimiento'] == 'negativo')
    total_neutral = sum(1 for r in resultados if r['sentimiento'] == 'neutral')
    score_promedio = np.mean([r['score'] for r in resultados])
    
    # Crear visualización
    fig = plt.figure(figsize=(20, 14), facecolor='#F5F5F5')
    gs = fig.add_gridspec(3, 3, hspace=0.4, wspace=0.35)
    
    # Gráfico 1: Distribución de sentimientos
    ax1 = fig.add_subplot(gs[0, 0])
    sentimientos = ['Positivo', 'Negativo', 'Neutral']
    valores = [total_positivo, total_negativo, total_neutral]
    colores = ['#4CAF50', '#F44336', '#FF9800']
    
    bars = ax1.bar(sentimientos, valores, color=colores, alpha=0.8, 
                   edgecolor='white', linewidth=2)
    ax1.set_title('Distribución de Sentimientos', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    ax1.set_ylabel('Cantidad', fontweight='bold', fontsize=12)
    ax1.grid(axis='y', alpha=0.4, linestyle='--')
    ax1.set_facecolor('white')
    
    for bar, val in zip(bars, valores):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{val}', ha='center', va='bottom', fontweight='bold', fontsize=11)
    
    # Gráfico 2: Score promedio
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.barh(['Score Promedio'], [score_promedio], 
            color='#2196F3' if score_promedio > 0 else '#F44336', 
            alpha=0.8, edgecolor='white', linewidth=2)
    ax2.axvline(0, color='black', linestyle='--', linewidth=1)
    ax2.set_xlim(-0.5, 0.5)
    ax2.set_title('Score Promedio de Sentimiento', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    ax2.set_xlabel('Score (-1 a +1)', fontweight='bold', fontsize=12)
    ax2.grid(axis='x', alpha=0.4, linestyle='--')
    ax2.set_facecolor('white')
    ax2.text(score_promedio, 0, f'{score_promedio:.3f}', 
            ha='center' if abs(score_promedio) < 0.1 else ('right' if score_promedio < 0 else 'left'),
            va='center', fontweight='bold', fontsize=12, 
            bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
    
    # Gráfico 3: Top tendencias
    ax3 = fig.add_subplot(gs[0, 2])
    palabras, frecuencias = zip(*tendencias[:10])
    y_pos = np.arange(len(palabras))
    
    bars = ax3.barh(y_pos, frecuencias, color='#9C27B0', alpha=0.8,
                   edgecolor='white', linewidth=1)
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(palabras)
    ax3.set_title('Top 10 Tendencias/Temas', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    ax3.set_xlabel('Frecuencia', fontweight='bold', fontsize=12)
    ax3.grid(axis='x', alpha=0.4, linestyle='--')
    ax3.set_facecolor('white')
    
    # Gráfico 4: Scores individuales
    ax4 = fig.add_subplot(gs[1, :])
    indices = range(len(resultados))
    scores = [r['score'] for r in resultados]
    colores_scores = ['#4CAF50' if s > 0.1 else '#F44336' if s < -0.1 else '#FF9800' 
                     for s in scores]
    
    ax4.bar(indices, scores, color=colores_scores, alpha=0.7, 
           edgecolor='white', linewidth=1.5)
    ax4.axhline(0, color='black', linestyle='-', linewidth=1)
    ax4.axhline(0.1, color='green', linestyle='--', linewidth=1, alpha=0.5)
    ax4.axhline(-0.1, color='red', linestyle='--', linewidth=1, alpha=0.5)
    ax4.set_title('Scores de Sentimiento por Texto', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    ax4.set_xlabel('Índice del Texto', fontweight='bold', fontsize=12)
    ax4.set_ylabel('Score', fontweight='bold', fontsize=12)
    ax4.set_xticks(indices)
    ax4.grid(axis='y', alpha=0.4, linestyle='--')
    ax4.set_facecolor('white')
    
    # Gráfico 5: Análisis de palabras
    ax5 = fig.add_subplot(gs[2, 0])
    palabras_pos = sum(r['positivo'] for r in resultados)
    palabras_neg = sum(r['negativo'] for r in resultados)
    palabras_neu = sum(r['neutral'] for r in resultados)
    
    ax5.pie([palabras_pos, palabras_neg, palabras_neu], 
           labels=['Positivas', 'Negativas', 'Neutrales'],
           colors=['#4CAF50', '#F44336', '#FF9800'],
           autopct='%1.1f%%', startangle=90, textprops={'fontweight': 'bold'})
    ax5.set_title('Distribución de Palabras', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    
    # Gráfico 6: Comparativa
    ax6 = fig.add_subplot(gs[2, 1:])
    categorias = ['Positivo', 'Negativo', 'Neutral']
    valores_cat = [total_positivo, total_negativo, total_neutral]
    porcentajes = [v/sum(valores_cat)*100 for v in valores_cat]
    
    ax6.bar(categorias, valores_cat, color=colores, alpha=0.8,
           edgecolor='white', linewidth=2)
    ax6.set_title('Análisis Comparativo de Sentimientos', fontweight='bold', 
                 fontsize=14, pad=20, color='#1F4E78')
    ax6.set_ylabel('Cantidad', fontweight='bold', fontsize=12)
    ax6.grid(axis='y', alpha=0.4, linestyle='--')
    ax6.set_facecolor('white')
    
    for bar, val, pct in zip(ax6.patches, valores_cat, porcentajes):
        height = bar.get_height()
        ax6.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                f'{val} ({pct:.1f}%)', ha='center', va='bottom', 
                fontweight='bold', fontsize=11)
    
    plt.suptitle('ANÁLISIS DE SENTIMIENTO Y TENDENCIAS - MARKETING', 
                fontsize=20, fontweight='bold', y=0.98, color='#1F4E78')
    
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight',
               facecolor='#F5F5F5', edgecolor='none', pad_inches=0.3)
    buffer.seek(0)
    plt.close()
    
    return buffer, {
        'estadisticas': {
            'total_textos': len(textos_ejemplo),
            'positivo': total_positivo,
            'negativo': total_negativo,
            'neutral': total_neutral,
            'score_promedio': float(score_promedio)
        },
        'tendencias': [{'palabra': p, 'frecuencia': f} for p, f in tendencias[:20]],
        'resultados_detallados': resultados
    }

def guardar_analisis_excel(datos, archivo_excel):
    """Guarda análisis en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Hoja 1: Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    
    ws1['A1'] = "ANÁLISIS DE SENTIMIENTO - RESUMEN"
    ws1['A1'].font = Font(bold=True, size=16, color="1F4E78")
    ws1.merge_cells('A1:B1')
    
    stats = datos['estadisticas']
    ws1['A3'] = "Total Textos Analizados:"
    ws1['B3'] = stats['total_textos']
    ws1['A4'] = "Sentimientos Positivos:"
    ws1['B4'] = stats['positivo']
    ws1['A5'] = "Sentimientos Negativos:"
    ws1['B5'] = stats['negativo']
    ws1['A6'] = "Sentimientos Neutrales:"
    ws1['B6'] = stats['neutral']
    ws1['A7'] = "Score Promedio:"
    ws1['B7'] = stats['score_promedio']
    
    # Hoja 2: Tendencias
    ws2 = wb.create_sheet("Tendencias")
    ws2['A1'] = "Palabra"
    ws2['B1'] = "Frecuencia"
    ws2['A1'].font = Font(bold=True, color="FFFFFF")
    ws2['B1'].font = Font(bold=True, color="FFFFFF")
    ws2['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws2['B1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for row, tendencia in enumerate(datos['tendencias'], 2):
        ws2.cell(row=row, column=1, value=tendencia['palabra'])
        ws2.cell(row=row, column=2, value=tendencia['frecuencia'])
    
    # Hoja 3: Detalles
    ws3 = wb.create_sheet("Detalles")
    headers = ['Sentimiento', 'Score', 'Positivo', 'Negativo', 'Neutral', 'Total Palabras']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for row, resultado in enumerate(datos['resultados_detallados'], 2):
        ws3.cell(row=row, column=1, value=resultado['sentimiento'])
        ws3.cell(row=row, column=2, value=resultado['score'])
        ws3.cell(row=row, column=3, value=resultado['positivo'])
        ws3.cell(row=row, column=4, value=resultado['negativo'])
        ws3.cell(row=row, column=5, value=resultado['neutral'])
        ws3.cell(row=row, column=6, value=resultado['total_palabras'])
    
    wb.save(archivo_excel)
    print(f"✓ Análisis guardado en Excel: {archivo_excel}")

def main():
    """Función principal"""
    directorio = '/Users/adan/Documents/documentos_blatam/01_marketing'
    
    print("📊 Iniciando análisis de sentimiento y tendencias...\n")
    
    # Crear análisis
    grafico_buffer, datos = crear_analisis_sentimiento_completo()
    
    # Guardar gráfico
    archivo_imagen = os.path.join(directorio, 'ANALISIS_SENTIMIENTO.png')
    with open(archivo_imagen, 'wb') as f:
        f.write(grafico_buffer.read())
    print(f"✓ Gráfico guardado: {archivo_imagen}")
    
    # Guardar en Excel
    archivo_excel = os.path.join(directorio, 'ANALISIS_SENTIMIENTO.xlsx')
    guardar_analisis_excel(datos, archivo_excel)
    
    # Guardar JSON
    archivo_json = os.path.join(directorio, 'ANALISIS_SENTIMIENTO.json')
    with open(archivo_json, 'w', encoding='utf-8') as f:
        json.dump(datos, f, indent=2, ensure_ascii=False)
    print(f"✓ Datos JSON guardados: {archivo_json}")
    
    print("\n✅ Análisis de sentimiento completado!")
    print("📊 Incluye:")
    print("   • Distribución de sentimientos")
    print("   • Score promedio")
    print("   • Top tendencias y temas")
    print("   • Análisis detallado por texto")

if __name__ == "__main__":
    main()








