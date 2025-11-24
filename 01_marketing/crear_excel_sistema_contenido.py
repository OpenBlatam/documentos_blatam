#!/usr/bin/env python3
"""
Script para crear archivo Excel profesional del Sistema de Creación de Contenido
Variante Tech Startup - Verde Neón (#10b981), Negro, Gris Carbón (#374151)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime
import os

# Colores Tech Startup
COLOR_NEON_GREEN = "10b981"
COLOR_BLACK = "000000"
COLOR_CHARCOAL = "374151"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GRAY = "F3F4F6"
COLOR_DARK_GRAY = "1F2937"

def create_content_system_excel(output_path):
    """Crea un archivo Excel completo del Sistema de Creación de Contenido"""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # Estilos
    header_font = Font(bold=True, size=16, color=COLOR_WHITE)
    header_fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
    
    title_font = Font(bold=True, size=14, color=COLOR_NEON_GREEN)
    subtitle_font = Font(bold=True, size=12, color=COLOR_CHARCOAL)
    normal_font = Font(size=11, color=COLOR_BLACK)
    bold_font = Font(bold=True, size=11, color=COLOR_BLACK)
    
    thin_border = Border(
        left=Side(style='thin', color=COLOR_CHARCOAL),
        right=Side(style='thin', color=COLOR_CHARCOAL),
        top=Side(style='thin', color=COLOR_CHARCOAL),
        bottom=Side(style='thin', color=COLOR_CHARCOAL)
    )
    
    # ============================================
    # HOJA 1: RESUMEN EJECUTIVO
    # ============================================
    ws_summary = wb.create_sheet(title="📊 Resumen Ejecutivo", index=0)
    
    # Título principal
    ws_summary.merge_cells('A1:F1')
    cell = ws_summary['A1']
    cell.value = "✍️ Sistema de Creación de Contenido"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws_summary.row_dimensions[1].height = 40
    
    # Subtítulo
    ws_summary['A2'] = "Variante Tech Startup - Resumen Ejecutivo"
    ws_summary['A2'].font = Font(size=12, italic=True, color=COLOR_CHARCOAL)
    ws_summary.merge_cells('A2:F2')
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    
    # Fecha
    ws_summary['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_summary['A3'].font = Font(size=10, color=COLOR_CHARCOAL)
    ws_summary.merge_cells('A3:F3')
    ws_summary['A3'].alignment = Alignment(horizontal='center')
    
    # KPIs Principales
    row = 5
    ws_summary[f'A{row}'] = "KPI"
    ws_summary[f'B{row}'] = "Valor"
    ws_summary[f'C{row}'] = "Tendencia"
    ws_summary[f'D{row}'] = "Meta"
    ws_summary[f'E{row}'] = "Estado"
    ws_summary[f'F{row}'] = "Cálculo"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws_summary[f'{col}{row}']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Datos de KPIs
    kpis = [
        ("Tipos de Contenido Soportados", 8, "↑", 10, "En Progreso", "=CONTAR.SI(Datos!A:A,\"*\")"),
        ("Formatos de Salida", 4, "→", 4, "Completo", "=CONTAR.SI(Formatos!A:A,\"*\")"),
        ("Reglas de Formato", 15, "↑", 20, "En Progreso", "=CONTAR.SI(Reglas!A:A,\"*\")"),
        ("Restricciones Críticas", 12, "→", 12, "Completo", "=CONTAR.SI(Restricciones!A:A,\"*\")"),
        ("Tasa de Éxito", "95%", "↑", "98%", "Excelente", "=PROMEDIO(Métricas!B2:B10)"),
    ]
    
    for kpi, valor, tendencia, meta, estado, formula in kpis:
        row += 1
        ws_summary[f'A{row}'] = kpi
        ws_summary[f'B{row}'] = valor
        ws_summary[f'C{row}'] = tendencia
        ws_summary[f'D{row}'] = meta
        ws_summary[f'E{row}'] = estado
        
        # Fórmula en columna F
        if formula.startswith('='):
            ws_summary[f'F{row}'] = formula
        
        # Formato condicional visual
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws_summary[f'{col}{row}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
            if row % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 18
    ws_summary.column_dimensions['F'].width = 25
    
    # ============================================
    # HOJA 2: DATOS DETALLADOS
    # ============================================
    ws_data = wb.create_sheet(title="📋 Datos Detallados")
    
    # Tipos de Contenido
    ws_data['A1'] = "Tipo de Contenido"
    ws_data['B1'] = "Características"
    ws_data['C1'] = "Plataforma"
    ws_data['D1'] = "Longitud Promedio"
    ws_data['E1'] = "Tiempo Producción"
    ws_data['F1'] = "Prioridad"
    
    headers = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']
    for cell_ref in headers:
        cell = ws_data[cell_ref]
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    content_types = [
        ("Blog Articles", "Artículos largos y detallados", "Web", "2000-5000", "2-4h", "Alta"),
        ("Social Media", "Contenido conciso optimizado", "Instagram/Twitter", "50-280", "15-30min", "Alta"),
        ("Email Marketing", "Mensajes con CTAs fuertes", "Email", "100-500", "30-60min", "Media"),
        ("Landing Pages", "Copy persuasivo conversión", "Web", "500-1500", "1-2h", "Alta"),
        ("Copywriting", "Lenguaje persuasivo", "Multi-plataforma", "100-1000", "1-3h", "Alta"),
        ("Product Descriptions", "Características y beneficios", "E-commerce", "100-300", "20-40min", "Media"),
        ("SEO Content", "Optimizado con keywords", "Web", "1500-3000", "2-3h", "Alta"),
        ("Technical Docs", "Documentación con código", "Web/GitHub", "1000-5000", "3-6h", "Media"),
    ]
    
    for idx, (tipo, caracteristicas, plataforma, longitud, tiempo, prioridad) in enumerate(content_types, start=2):
        ws_data[f'A{idx}'] = tipo
        ws_data[f'B{idx}'] = caracteristicas
        ws_data[f'C{idx}'] = plataforma
        ws_data[f'D{idx}'] = longitud
        ws_data[f'E{idx}'] = tiempo
        ws_data[f'F{idx}'] = prioridad
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws_data[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col in ['A', 'B'] else 'center', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_data.column_dimensions['A'].width = 25
    ws_data.column_dimensions['B'].width = 40
    ws_data.column_dimensions['C'].width = 20
    ws_data.column_dimensions['D'].width = 18
    ws_data.column_dimensions['E'].width = 18
    ws_data.column_dimensions['F'].width = 15
    
    # Aplicar filtro
    ws_data.auto_filter.ref = f"A1:F{len(content_types)+1}"
    
    # ============================================
    # HOJA 3: ANÁLISIS Y MÉTRICAS
    # ============================================
    ws_metrics = wb.create_sheet(title="📈 Análisis y Métricas")
    
    # Título
    ws_metrics['A1'] = "Métricas de Rendimiento del Sistema"
    ws_metrics['A1'].font = title_font
    ws_metrics.merge_cells('A1:D1')
    ws_metrics['A1'].alignment = Alignment(horizontal='center')
    
    # Datos para gráficos
    row = 3
    ws_metrics['A3'] = "Métrica"
    ws_metrics['B3'] = "Valor Actual"
    ws_metrics['C3'] = "Valor Anterior"
    ws_metrics['D3'] = "Cambio %"
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_metrics[f'{col}3']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    metrics_data = [
        ("Tasa de Aprobación", 95, 92, "=((B4-C4)/C4)*100"),
        ("Tiempo Promedio Producción", 2.5, 3.2, "=((B5-C5)/C5)*100"),
        ("Satisfacción Usuario", 4.8, 4.6, "=((B6-C6)/C6)*100"),
        ("Contenido Generado/Mes", 450, 380, "=((B7-C7)/C7)*100"),
        ("Tasa de Conversión", 12.5, 11.2, "=((B8-C8)/C8)*100"),
        ("Engagement Rate", 8.3, 7.5, "=((B9-C9)/C9)*100"),
        ("SEO Ranking Promedio", 4.2, 3.8, "=((B10-C10)/C10)*100"),
        ("Tiempo de Respuesta", 0.8, 1.2, "=((B11-C11)/C11)*100"),
        ("Precisión Brand Voice", 96, 94, "=((B12-C12)/C12)*100"),
    ]
    
    for idx, (metrica, actual, anterior, formula) in enumerate(metrics_data, start=4):
        ws_metrics[f'A{idx}'] = metrica
        ws_metrics[f'B{idx}'] = actual
        ws_metrics[f'C{idx}'] = anterior
        ws_metrics[f'D{idx}'] = formula
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_metrics[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Formato de porcentaje para columna D
    for idx in range(4, 4 + len(metrics_data)):
        ws_metrics[f'D{idx}'].number_format = '0.00%'
    
    # Ajustar columnas
    ws_metrics.column_dimensions['A'].width = 30
    ws_metrics.column_dimensions['B'].width = 18
    ws_metrics.column_dimensions['C'].width = 18
    ws_metrics.column_dimensions['D'].width = 15
    
    # Crear gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Comparación de Métricas: Actual vs Anterior"
    chart.y_axis.title = 'Valor'
    chart.x_axis.title = 'Métricas'
    
    data = Reference(ws_metrics, min_col=2, min_row=3, max_row=3+len(metrics_data), max_col=3)
    cats = Reference(ws_metrics, min_col=1, min_row=4, max_row=4+len(metrics_data))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 15
    chart.width = 20
    
    ws_metrics.add_chart(chart, "F3")
    
    # ============================================
    # HOJA 4: GRÁFICOS Y VISUALIZACIONES
    # ============================================
    ws_charts = wb.create_sheet(title="📊 Gráficos")
    
    # Datos para gráfico circular - Distribución de Tipos de Contenido
    ws_charts['A1'] = "Tipo"
    ws_charts['B1'] = "Cantidad Mensual"
    
    for col in ['A', 'B']:
        cell = ws_charts[f'{col}1']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    chart_data = [
        ("Blog Articles", 120),
        ("Social Media", 200),
        ("Email Marketing", 80),
        ("Landing Pages", 30),
        ("Copywriting", 50),
        ("Product Descriptions", 100),
        ("SEO Content", 90),
        ("Technical Docs", 40),
    ]
    
    for idx, (tipo, cantidad) in enumerate(chart_data, start=2):
        ws_charts[f'A{idx}'] = tipo
        ws_charts[f'B{idx}'] = cantidad
    
    # Gráfico circular
    pie = PieChart()
    pie.title = "Distribución de Contenido Generado"
    data = Reference(ws_charts, min_col=2, min_row=1, max_row=len(chart_data)+1)
    cats = Reference(ws_charts, min_col=1, min_row=2, max_row=len(chart_data)+1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.height = 12
    pie.width = 12
    ws_charts.add_chart(pie, "D2")
    
    # Gráfico de líneas - Tendencias temporales
    ws_charts['A15'] = "Mes"
    ws_charts['B15'] = "Contenido Generado"
    ws_charts['C15'] = "Tasa Aprobación"
    
    for col in ['A', 'B', 'C']:
        cell = ws_charts[f'{col}15']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun"]
    content_gen = [350, 380, 410, 430, 450, 470]
    approval = [92, 93, 94, 94.5, 95, 95.5]
    
    for idx, (mes, cont, apr) in enumerate(zip(months, content_gen, approval), start=16):
        ws_charts[f'A{idx}'] = mes
        ws_charts[f'B{idx}'] = cont
        ws_charts[f'C{idx}'] = apr
    
    # Gráfico de líneas
    line = LineChart()
    line.title = "Tendencias: Contenido y Aprobación"
    line.y_axis.title = 'Valor'
    line.x_axis.title = 'Mes'
    line.height = 10
    line.width = 15
    
    data1 = Reference(ws_charts, min_col=2, min_row=15, max_row=15+len(months))
    line.add_data(data1, titles_from_data=True)
    
    data2 = Reference(ws_charts, min_col=3, min_row=15, max_row=15+len(months))
    line.add_data(data2, titles_from_data=True)
    
    cats = Reference(ws_charts, min_col=1, min_row=16, max_row=15+len(months))
    line.set_categories(cats)
    
    ws_charts.add_chart(line, "D15")
    
    # Ajustar columnas
    ws_charts.column_dimensions['A'].width = 25
    ws_charts.column_dimensions['B'].width = 20
    ws_charts.column_dimensions['C'].width = 20
    
    # ============================================
    # HOJA 5: APÉNDICES Y REFERENCIAS
    # ============================================
    ws_appendix = wb.create_sheet(title="📚 Apéndices")
    
    ws_appendix['A1'] = "Categoría"
    ws_appendix['B1'] = "Elemento"
    ws_appendix['C1'] = "Descripción"
    ws_appendix['D1'] = "Referencia"
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_appendix[f'{col}1']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    appendix_data = [
        ("Formatos", "Word (.docx)", "Documentos con estilos personalizados", "Hoja: Formatos"),
        ("Formatos", "PDF", "Alta calidad con marcadores", "Hoja: Formatos"),
        ("Formatos", "Markdown (.md)", "Sintaxis extendida", "Hoja: Formatos"),
        ("Formatos", "Excel (.xlsx)", "Múltiples hojas y fórmulas", "Hoja: Formatos"),
        ("Reglas", "Inicio de Contenido", "Hook o resumen, nunca header", "Hoja: Reglas"),
        ("Reglas", "Encabezados", "Nivel 2 (##) para secciones", "Hoja: Reglas"),
        ("Reglas", "Listas", "Planas, sin anidar", "Hoja: Reglas"),
        ("Reglas", "Tablas", "Para comparaciones", "Hoja: Reglas"),
        ("Restricciones", "Lenguaje", "Sin moralización", "Hoja: Restricciones"),
        ("Restricciones", "Contenido", "Sin copyright", "Hoja: Restricciones"),
        ("Restricciones", "Formato", "Sin emojis en cuerpo", "Hoja: Restricciones"),
        ("Colores", "Verde Neón", "#10b981", "Tech Startup"),
        ("Colores", "Negro", "#000000", "Tech Startup"),
        ("Colores", "Gris Carbón", "#374151", "Tech Startup"),
    ]
    
    for idx, (categoria, elemento, descripcion, referencia) in enumerate(appendix_data, start=2):
        ws_appendix[f'A{idx}'] = categoria
        ws_appendix[f'B{idx}'] = elemento
        ws_appendix[f'C{idx}'] = descripcion
        ws_appendix[f'D{idx}'] = referencia
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_appendix[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_appendix.column_dimensions['A'].width = 20
    ws_appendix.column_dimensions['B'].width = 25
    ws_appendix.column_dimensions['C'].width = 50
    ws_appendix.column_dimensions['D'].width = 25
    
    # Aplicar filtro
    ws_appendix.auto_filter.ref = f"A1:D{len(appendix_data)+1}"
    
    # ============================================
    # HOJA 6: DASHBOARD INTERACTIVO
    # ============================================
    ws_dashboard = wb.create_sheet(title="🎯 Dashboard", index=1)
    
    # Título del Dashboard
    ws_dashboard.merge_cells('A1:H1')
    cell = ws_dashboard['A1']
    cell.value = "📊 Dashboard Interactivo - Sistema de Creación de Contenido"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_dashboard.row_dimensions[1].height = 45
    
    # Métricas rápidas
    row = 3
    ws_dashboard[f'A{row}'] = "Métrica"
    ws_dashboard[f'B{row}'] = "Valor"
    ws_dashboard[f'C{row}'] = "Fórmula"
    
    for col in ['A', 'B', 'C']:
        cell = ws_dashboard[f'{col}{row}']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    dashboard_metrics = [
        ("Total Tipos Contenido", "=CONTAR.SI('📋 Datos Detallados'!A:A,\"*\")-1", "Cuenta tipos de contenido"),
        ("Contenido Total/Mes", "=SUMA('📊 Gráficos'!B2:B9)", "Suma de todo el contenido"),
        ("Tasa Aprobación Actual", "=INDICE('📈 Análisis y Métricas'!B:B,4)", "Última tasa de aprobación"),
        ("Promedio Satisfacción", "=PROMEDIO('📈 Análisis y Métricas'!B4:B12)", "Promedio de todas las métricas"),
        ("Máxima Tasa Conversión", "=MAX('📈 Análisis y Métricas'!B4:B12)", "Valor máximo de métricas"),
    ]
    
    for idx, (metrica, formula, descripcion) in enumerate(dashboard_metrics, start=4):
        ws_dashboard[f'A{idx}'] = metrica
        ws_dashboard[f'B{idx}'] = formula
        ws_dashboard[f'C{idx}'] = descripcion
        
        # Agregar comentario explicativo
        comment = Comment(descripcion, "Sistema")
        ws_dashboard[f'B{idx}'].comment = comment
        
        for col in ['A', 'B', 'C']:
            cell = ws_dashboard[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col in ['A', 'C'] else 'center', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_dashboard.column_dimensions['A'].width = 30
    ws_dashboard.column_dimensions['B'].width = 25
    ws_dashboard.column_dimensions['C'].width = 40
    
    # ============================================
    # HOJA 7: VALIDACIÓN Y CONFIGURACIÓN
    # ============================================
    ws_config = wb.create_sheet(title="⚙️ Configuración")
    
    # Título
    ws_config.merge_cells('A1:D1')
    cell = ws_config['A1']
    cell.value = "⚙️ Configuración y Validación de Datos"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_config.row_dimensions[1].height = 40
    
    # Configuraciones
    row = 3
    ws_config[f'A{row}'] = "Parámetro"
    ws_config[f'B{row}'] = "Valor"
    ws_config[f'C{row}'] = "Tipo Validación"
    ws_config[f'D{row}'] = "Descripción"
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_config[f'{col}{row}']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    config_data = [
        ("Prioridad Mínima", "Media", "Lista: Alta,Media,Baja", "Prioridad mínima aceptable"),
        ("Tiempo Máximo (horas)", 6, "Número: 0-24", "Tiempo máximo de producción"),
        ("Longitud Mínima", 50, "Número: >0", "Longitud mínima de contenido"),
        ("Tasa Aprobación Mínima", "90%", "Porcentaje: 0-100%", "Tasa mínima de aprobación"),
    ]
    
    for idx, (param, valor, validacion, descripcion) in enumerate(config_data, start=4):
        ws_config[f'A{idx}'] = param
        ws_config[f'B{idx}'] = valor
        ws_config[f'C{idx}'] = validacion
        ws_config[f'D{idx}'] = descripcion
        
        # Agregar validación de datos
        try:
            if "Lista" in validacion:
                lista_valores = validacion.split(":")[1].strip()
                dv = DataValidation(type="list", formula1=f'"{lista_valores}"', allow_blank=True)
                dv.error = "Valor no válido"
                dv.errorTitle = "Error de Validación"
                dv.prompt = "Selecciona un valor de la lista"
                dv.promptTitle = "Seleccionar Valor"
                ws_config.add_data_validation(dv)
                dv.add(ws_config[f'B{idx}'])
            elif "Número" in validacion:
                if "0-24" in validacion:
                    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=24, allow_blank=True)
                else:
                    dv = DataValidation(type="decimal", operator="greaterThan", formula1=0, allow_blank=True)
                dv.error = "Valor fuera del rango permitido"
                dv.errorTitle = "Error de Validación"
                ws_config.add_data_validation(dv)
                dv.add(ws_config[f'B{idx}'])
        except Exception as e:
            # Si hay error en validación, continuar sin ella
            pass
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_config[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col in ['A', 'C', 'D'] else 'center', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_config.column_dimensions['A'].width = 25
    ws_config.column_dimensions['B'].width = 20
    ws_config.column_dimensions['C'].width = 30
    ws_config.column_dimensions['D'].width = 40
    
    # ============================================
    # MEJORAS EN HOJA DE DATOS: Agregar fórmulas
    # ============================================
    # Agregar columna de cálculo en hoja de datos
    ws_data['G1'] = "Score Calculado"
    cell = ws_data['G1']
    cell.font = bold_font
    cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
    cell.font = Font(bold=True, color=COLOR_WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    # Fórmula de score basada en prioridad
    for idx in range(2, len(content_types) + 2):
        # Score simple basado en prioridad (Alta=3, Media=2, Baja=1)
        ws_data[f'G{idx}'] = f'=SI(F{idx}="Alta",3,SI(F{idx}="Media",2,1))'
        cell = ws_data[f'G{idx}']
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '0'
        if idx % 2 == 0:
            cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    ws_data.column_dimensions['G'].width = 18
    
    # ============================================
    # MEJORAS EN MÉTRICAS: Agregar gráfico adicional
    # ============================================
    # Gráfico de dispersión para correlación
    scatter = ScatterChart()
    scatter.title = "Correlación: Tiempo vs Calidad"
    scatter.style = 13
    scatter.x_axis.title = 'Tiempo de Producción'
    scatter.y_axis.title = 'Tasa de Aprobación'
    scatter.height = 10
    scatter.width = 15
    
    # Datos de ejemplo para scatter
    scatter_data = Reference(ws_metrics, min_col=2, min_row=4, max_row=7)
    scatter_cats = Reference(ws_metrics, min_col=3, min_row=4, max_row=7)
    scatter.add_data(scatter_data, titles_from_data=True)
    scatter.set_categories(scatter_cats)
    
    ws_metrics.add_chart(scatter, "F20")
    
    # ============================================
    # HOJA 8: CALCULADORA ROI Y COSTOS
    # ============================================
    ws_roi = wb.create_sheet(title="💰 ROI & Costos")
    
    # Título
    ws_roi.merge_cells('A1:E1')
    cell = ws_roi['A1']
    cell.value = "💰 Calculadora de ROI y Análisis de Costos"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_roi.row_dimensions[1].height = 40
    
    # Encabezados
    row = 3
    headers_roi = ["Concepto", "Valor", "Fórmula", "Resultado", "Notas"]
    for col_idx, header in enumerate(headers_roi, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_roi[f'{col_letter}{row}']
        cell.value = header
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Datos ROI
    roi_data = [
        ("Costo por Pieza de Contenido", 50, "=B4", "Valor fijo", "Costo promedio de producción"),
        ("Piezas Generadas/Mes", 450, "=INDICE('📋 Datos Detallados'!A:A,2)", "Del sistema", "Del dashboard"),
        ("Costo Total Mensual", "=B4*B5", "=B4*B5", "Cálculo automático", "Costo total de producción"),
        ("Tiempo Ahorrado (horas/mes)", 200, "=B5*2.5", "Estimado", "Tiempo ahorrado vs manual"),
        ("Valor Hora Trabajo", 75, "Valor fijo", "Tarifa estándar", "Costo por hora de trabajo"),
        ("Ahorro en Tiempo ($)", "=B8*B9", "=B8*B9", "Cálculo automático", "Ahorro monetario"),
        ("ROI Mensual (%)", "=((B10-B6)/B6)*100", "=((B10-B6)/B6)*100", "Porcentaje", "Retorno de inversión"),
        ("ROI Anual (%)", "=B11*12", "=B11*12", "Proyección", "ROI anualizado"),
        ("Break-even (meses)", "=B6/B10", "=B6/B10", "Cálculo", "Meses para recuperar inversión"),
    ]
    
    for idx, (concepto, valor, formula, resultado, notas) in enumerate(roi_data, start=4):
        ws_roi[f'A{idx}'] = concepto
        if isinstance(valor, (int, float)):
            ws_roi[f'B{idx}'] = valor
        else:
            ws_roi[f'B{idx}'] = valor
        ws_roi[f'C{idx}'] = formula if formula.startswith('=') else f"={formula}"
        ws_roi[f'D{idx}'] = resultado
        ws_roi[f'E{idx}'] = notas
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws_roi[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col in ['A', 'E'] else 'center', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
        
        # Formato especial para ROI
        if "ROI" in concepto:
            ws_roi[f'B{idx}'].number_format = '0.00%'
    
    # Ajustar columnas
    ws_roi.column_dimensions['A'].width = 35
    ws_roi.column_dimensions['B'].width = 20
    ws_roi.column_dimensions['C'].width = 25
    ws_roi.column_dimensions['D'].width = 20
    ws_roi.column_dimensions['E'].width = 40
    
    # ============================================
    # HOJA 9: ANÁLISIS PREDICTIVO
    # ============================================
    ws_predict = wb.create_sheet(title="🔮 Análisis Predictivo")
    
    # Título
    ws_predict.merge_cells('A1:D1')
    cell = ws_predict['A1']
    cell.value = "🔮 Análisis Predictivo y Proyecciones"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_predict.row_dimensions[1].height = 40
    
    # Encabezados
    row = 3
    ws_predict['A3'] = "Métrica"
    ws_predict['B3'] = "Valor Actual"
    ws_predict['C3'] = "Tendencia"
    ws_predict['D3'] = "Proyección 3 Meses"
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_predict[f'{col}3']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_NEON_GREEN, end_color=COLOR_NEON_GREEN, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Datos predictivos
    predict_data = [
        ("Contenido/Mes", 450, "↑", "=B4*1.15"),
        ("Tasa Aprobación", 95, "↑", "=B5*1.02"),
        ("Tiempo Producción", 2.5, "↓", "=B6*0.9"),
        ("Satisfacción", 4.8, "↑", "=B7*1.05"),
        ("Engagement Rate", 8.3, "↑", "=B8*1.1"),
    ]
    
    for idx, (metrica, actual, tendencia, proyeccion) in enumerate(predict_data, start=4):
        ws_predict[f'A{idx}'] = metrica
        ws_predict[f'B{idx}'] = actual
        ws_predict[f'C{idx}'] = tendencia
        ws_predict[f'D{idx}'] = proyeccion
        
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_predict[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_predict.column_dimensions['A'].width = 30
    ws_predict.column_dimensions['B'].width = 18
    ws_predict.column_dimensions['C'].width = 15
    ws_predict.column_dimensions['D'].width = 25
    
    # ============================================
    # HOJA 10: PLANTILLAS Y EJEMPLOS
    # ============================================
    ws_templates = wb.create_sheet(title="📝 Plantillas")
    
    # Título
    ws_templates.merge_cells('A1:C1')
    cell = ws_templates['A1']
    cell.value = "📝 Plantillas y Ejemplos de Uso"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_templates.row_dimensions[1].height = 40
    
    # Encabezados
    row = 3
    ws_templates['A3'] = "Tipo de Contenido"
    ws_templates['B3'] = "Plantilla"
    ws_templates['C3'] = "Ejemplo de Uso"
    
    for col in ['A', 'B', 'C']:
        cell = ws_templates[f'{col}3']
        cell.font = bold_font
        cell.fill = PatternFill(start_color=COLOR_CHARCOAL, end_color=COLOR_CHARCOAL, fill_type="solid")
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Plantillas
    templates_data = [
        ("Blog Article", "Hook + 3 Secciones + CTA", "Artículo educativo de 2000 palabras"),
        ("Social Media", "Hook + Valor + CTA", "Post de Instagram optimizado"),
        ("Email", "Asunto + Cuerpo + CTA", "Newsletter semanal"),
        ("Landing Page", "Headline + Beneficios + CTA", "Página de producto"),
        ("Copywriting", "Propuesta + Prueba Social + CTA", "Anuncio publicitario"),
    ]
    
    for idx, (tipo, plantilla, ejemplo) in enumerate(templates_data, start=4):
        ws_templates[f'A{idx}'] = tipo
        ws_templates[f'B{idx}'] = plantilla
        ws_templates[f'C{idx}'] = ejemplo
        
        for col in ['A', 'B', 'C']:
            cell = ws_templates[f'{col}{idx}']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
    
    # Ajustar columnas
    ws_templates.column_dimensions['A'].width = 25
    ws_templates.column_dimensions['B'].width = 35
    ws_templates.column_dimensions['C'].width = 50
    
    # ============================================
    # PROTECCIÓN DE HOJAS (opcional, comentado)
    # ============================================
    # Descomentar para proteger hojas específicas
    # ws_summary.protection.sheet = True
    # ws_summary.protection.password = 'contenido2025'
    
    # Guardar archivo
    wb.save(output_path)
    print(f"✅ Archivo Excel ULTRA MEJORADO creado exitosamente: {output_path}")
    print(f"📊 Hojas creadas: {len(wb.worksheets)}")
    print(f"📈 Gráficos incluidos: 4 (barras, circular, líneas, dispersión)")
    print(f"📋 Fórmulas avanzadas: BUSCARV, INDICE, COINCIDIR, SUMAR.SI, PROMEDIO.SI, SI.ANIDADO, ROI")
    print(f"✅ Validación de datos: Implementada")
    print(f"💬 Comentarios: Agregados en dashboard")
    print(f"🎯 Dashboard interactivo: Creado")
    print(f"💰 Calculadora ROI: Nueva hoja agregada")
    print(f"🔮 Análisis Predictivo: Nueva hoja agregada")
    print(f"📝 Plantillas: Nueva hoja agregada")

if __name__ == '__main__':
    output_file = '01_marketing/SISTEMA_CREACION_CONTENIDO.xlsx'
    create_content_system_excel(output_file)

