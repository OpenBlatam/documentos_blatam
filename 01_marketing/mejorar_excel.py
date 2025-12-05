#!/usr/bin/env python3
"""Script para mejorar el archivo Excel del Sistema de Calendario"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Cargar workbook
try:
    wb = openpyxl.load_workbook('SISTEMA_CALENDARIO_CONTENIDO_REDES_SOCIALES.xlsx')
    print(f'✅ Archivo cargado: {len(wb.sheetnames)} hojas')
except:
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    print('📝 Nuevo workbook')

# Estilos
AZUL_MARINO = '1e40af'
VERDE_ACADEMICO = '059669'
BEIGE = 'fef3c7'
GRIS_CLARO = 'e5e7eb'

header_fill = PatternFill(start_color=AZUL_MARINO, end_color=AZUL_MARINO, fill_type='solid')
light_fill = PatternFill(start_color=BEIGE, end_color=BEIGE, fill_type='solid')
alt_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
title_font = Font(bold=True, size=16, color=AZUL_MARINO)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_table_style(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if row == start_row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif row % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# HOJA 12: Plantilla Brief
if '12_Plantilla_Brief' not in wb.sheetnames:
    ws = wb.create_sheet('12_Plantilla_Brief')
    ws['A1'] = 'PLANTILLA DE BRIEF PARA CALENDARIO'
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    
    data = [
        ['Campo', 'Descripción', 'Ejemplo'],
        ['Marca', 'Nombre de la marca', 'TechStart Inc.'],
        ['Plataformas', 'Redes sociales', 'Instagram, LinkedIn'],
        ['Período', 'Duración', 'Semanal o Mensual'],
        ['Objetivos', 'Objetivos marketing', 'Awareness, Engagement'],
        ['Audiencia', 'Descripción', 'Mujeres 25-40'],
        ['Temas', 'Pilares contenido', 'Innovación, Sostenibilidad'],
    ]
    
    for idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=idx, column=col_idx, value=value)
    
    apply_table_style(ws, 3, 9, 1, 3)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    print('✅ Hoja 12 creada')

# HOJA 13: Calculadora
if '13_Calculadora' not in wb.sheetnames:
    ws = wb.create_sheet('13_Calculadora')
    ws['A1'] = 'CALCULADORA DE FRECUENCIA'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    headers = ['Plataforma', 'Frecuencia', 'Posts/Semana', 'Posts/Mes']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    data = [
        ['Instagram', '1-2x/día', 10.5, '=C4*4.33'],
        ['Facebook', '1x/día', 7, '=C5*4.33'],
        ['Twitter/X', '3-5x/día', 28, '=C6*4.33'],
        ['LinkedIn', '1x/día', 7, '=C7*4.33'],
        ['TikTok', '1-3x/día', 14, '=C8*4.33'],
    ]
    
    for idx, row_data in enumerate(data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 1:
                cell.fill = light_fill
                cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    print('✅ Hoja 13 creada')

# HOJA 14: Comparativa
if '14_Comparativa' not in wb.sheetnames:
    ws = wb.create_sheet('14_Comparativa')
    ws['A1'] = 'COMPARATIVA DE PLATAFORMAS'
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    data = [
        ['Plataforma', 'Horario', 'Frecuencia', 'Formato', 'Hashtags', 'Alcance', 'Engagement'],
        ['Instagram', '11 AM-1 PM', 'Alta', 'Visual', '5-10', 'Alto', 'Alto'],
        ['Facebook', '1-3 PM', 'Media', 'Texto/Video', '1-2', 'Muy Alto', 'Medio'],
        ['Twitter/X', '8-9 AM', 'Muy Alta', 'Texto', '1-2', 'Medio', 'Medio'],
        ['LinkedIn', '8-9 AM', 'Baja', 'Texto/Video', '3-5', 'Medio', 'Alto'],
        ['TikTok', '6-10 AM', 'Alta', 'Video', '3-5', 'Alto', 'Muy Alto'],
    ]
    
    for idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=idx, column=col_idx, value=value)
    
    apply_table_style(ws, 3, 8, 1, 7)
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    print('✅ Hoja 14 creada')

# HOJA 15: Checklist
if '15_Implementación' not in wb.sheetnames:
    ws = wb.create_sheet('15_Implementación')
    ws['A1'] = 'CHECKLIST DE IMPLEMENTACIÓN'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    data = [
        ['Fase', 'Tarea', 'Responsable', 'Estado', 'Fecha'],
        ['Preparación', 'Revisar brief', '', 'Pendiente', ''],
        ['Preparación', 'Validar objetivos', '', 'Pendiente', ''],
        ['Creación', 'Generar calendario', '', 'Pendiente', ''],
        ['Aprobación', 'Revisión interna', '', 'Pendiente', ''],
        ['Producción', 'Crear contenido', '', 'Pendiente', ''],
        ['Ejecución', 'Publicar', '', 'Pendiente', ''],
    ]
    
    for idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            cell.border = border
            if idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 4 and idx > 3:
                dv = DataValidation(type="list", formula1='"Pendiente,En Progreso,Completado"')
                ws.add_data_validation(dv)
                dv.add(cell)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    print('✅ Hoja 15 creada')

# Guardar
wb.save('SISTEMA_CALENDARIO_CONTENIDO_REDES_SOCIALES.xlsx')
print(f'\n✅ Excel mejorado guardado')
print(f'📊 Total hojas: {len(wb.sheetnames)}')
print(f'📋 Hojas: {", ".join(wb.sheetnames)}')









