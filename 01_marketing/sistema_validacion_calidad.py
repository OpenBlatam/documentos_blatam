#!/usr/bin/env python3
"""
Sistema de Validación y Control de Calidad - Valida documentos generados
y verifica que cumplan con estándares de calidad.
"""

import os
import json
from datetime import datetime
from pathlib import Path
import subprocess
import sys

class ValidadorCalidad:
    """Valida calidad de documentos generados"""
    
    def __init__(self):
        self.directorio = '/Users/adan/Documents/documentos_blatam/01_marketing'
        self.estandares = {
            'word': {
                'tamaño_min_kb': 50,
                'tamaño_max_mb': 10,
                'extensiones': ['.docx']
            },
            'excel': {
                'tamaño_min_kb': 100,
                'tamaño_max_mb': 20,
                'extensiones': ['.xlsx'],
                'hojas_minimas': 1
            },
            'powerpoint': {
                'tamaño_min_kb': 200,
                'tamaño_max_mb': 50,
                'extensiones': ['.pptx']
            },
            'pdf': {
                'tamaño_min_kb': 100,
                'tamaño_max_mb': 15,
                'extensiones': ['.pdf']
            },
            'html': {
                'tamaño_min_kb': 10,
                'tamaño_max_mb': 5,
                'extensiones': ['.html']
            },
            'png': {
                'tamaño_min_kb': 50,
                'tamaño_max_mb': 10,
                'extensiones': ['.png'],
                'resolucion_minima': 1000  # píxeles
            }
        }
    
    def validar_archivo(self, archivo_path):
        """Valida un archivo individual"""
        resultado = {
            'archivo': os.path.basename(archivo_path),
            'ruta': archivo_path,
            'valido': True,
            'errores': [],
            'advertencias': [],
            'tamaño_kb': 0,
            'fecha_modificacion': None
        }
        
        if not os.path.exists(archivo_path):
            resultado['valido'] = False
            resultado['errores'].append('Archivo no existe')
            return resultado
        
        # Obtener tamaño
        tamaño_bytes = os.path.getsize(archivo_path)
        tamaño_kb = tamaño_bytes / 1024
        tamaño_mb = tamaño_kb / 1024
        resultado['tamaño_kb'] = tamaño_kb
        
        # Obtener fecha
        fecha_mod = datetime.fromtimestamp(os.path.getmtime(archivo_path))
        resultado['fecha_modificacion'] = fecha_mod.isoformat()
        
        # Determinar tipo
        extension = Path(archivo_path).suffix.lower()
        tipo = None
        
        for tipo_doc, estandares in self.estandares.items():
            if extension in estandares['extensiones']:
                tipo = tipo_doc
                break
        
        if not tipo:
            resultado['advertencias'].append(f'Extensión desconocida: {extension}')
            return resultado
        
        # Validar según estándares
        estandares_tipo = self.estandares[tipo]
        
        # Validar tamaño mínimo
        if tamaño_kb < estandares_tipo['tamaño_min_kb']:
            resultado['valido'] = False
            resultado['errores'].append(
                f'Tamaño muy pequeño: {tamaño_kb:.1f} KB '
                f'(mínimo: {estandares_tipo["tamaño_min_kb"]} KB)'
            )
        
        # Validar tamaño máximo
        if tamaño_mb > estandares_tipo['tamaño_max_mb']:
            resultado['advertencias'].append(
                f'Tamaño muy grande: {tamaño_mb:.2f} MB '
                f'(máximo recomendado: {estandares_tipo["tamaño_max_mb"]} MB)'
            )
        
        # Validaciones específicas por tipo
        if tipo == 'excel':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(archivo_path, read_only=True)
                num_hojas = len(wb.sheetnames)
                if num_hojas < estandares_tipo.get('hojas_minimas', 1):
                    resultado['advertencias'].append(
                        f'Pocas hojas: {num_hojas} (mínimo recomendado: '
                        f'{estandares_tipo["hojas_minimas"]})'
                    )
                wb.close()
            except Exception as e:
                resultado['advertencias'].append(f'No se pudo validar estructura Excel: {e}')
        
        return resultado
    
    def validar_directorio(self, directorio=None):
        """Valida todos los archivos en un directorio"""
        if directorio is None:
            directorio = self.directorio
        
        extensiones_validas = []
        for estandares in self.estandares.values():
            extensiones_validas.extend(estandares['extensiones'])
        
        resultados = []
        archivos_encontrados = []
        
        for ext in extensiones_validas:
            archivos = list(Path(directorio).glob(f'*{ext}'))
            archivos_encontrados.extend(archivos)
        
        print(f"🔍 Validando {len(archivos_encontrados)} archivos...\n")
        
        for archivo in archivos_encontrados:
            resultado = self.validar_archivo(str(archivo))
            resultados.append(resultado)
            
            estado = "✅" if resultado['valido'] else "❌"
            print(f"{estado} {resultado['archivo']}")
            if resultado['errores']:
                for error in resultado['errores']:
                    print(f"   ❌ Error: {error}")
            if resultado['advertencias']:
                for advertencia in resultado['advertencias']:
                    print(f"   ⚠️  Advertencia: {advertencia}")
        
        return resultados
    
    def generar_reporte(self, resultados, archivo_salida=None):
        """Genera reporte de validación"""
        if archivo_salida is None:
            archivo_salida = os.path.join(
                self.directorio, 
                f'REPORTE_VALIDACION_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
            )
        
        total = len(resultados)
        validos = sum(1 for r in resultados if r['valido'])
        invalidos = total - validos
        con_advertencias = sum(1 for r in resultados if r['advertencias'])
        
        reporte = {
            'fecha': datetime.now().isoformat(),
            'resumen': {
                'total_archivos': total,
                'archivos_validos': validos,
                'archivos_invalidos': invalidos,
                'archivos_con_advertencias': con_advertencias,
                'porcentaje_valido': (validos / total * 100) if total > 0 else 0
            },
            'resultados_detallados': resultados
        }
        
        with open(archivo_salida, 'w', encoding='utf-8') as f:
            json.dump(reporte, f, indent=2, ensure_ascii=False)
        
        # También generar reporte en texto
        archivo_txt = archivo_salida.replace('.json', '.txt')
        with open(archivo_txt, 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("REPORTE DE VALIDACIÓN DE CALIDAD\n")
            f.write("="*70 + "\n\n")
            f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
            f.write("RESUMEN\n")
            f.write("-"*70 + "\n")
            f.write(f"Total de archivos: {total}\n")
            f.write(f"Archivos válidos: {validos} ({validos/total*100:.1f}%)\n")
            f.write(f"Archivos inválidos: {invalidos} ({invalidos/total*100:.1f}%)\n")
            f.write(f"Archivos con advertencias: {con_advertencias}\n\n")
            f.write("DETALLES\n")
            f.write("-"*70 + "\n\n")
            
            for resultado in resultados:
                estado = "✅ VÁLIDO" if resultado['valido'] else "❌ INVÁLIDO"
                f.write(f"{estado} - {resultado['archivo']}\n")
                f.write(f"   Tamaño: {resultado['tamaño_kb']:.1f} KB\n")
                if resultado['errores']:
                    for error in resultado['errores']:
                        f.write(f"   ❌ {error}\n")
                if resultado['advertencias']:
                    for advertencia in resultado['advertencias']:
                        f.write(f"   ⚠️  {advertencia}\n")
                f.write("\n")
        
        print(f"\n📄 Reporte guardado:")
        print(f"   • {archivo_salida}")
        print(f"   • {archivo_txt}")
        
        return reporte

def main():
    """Función principal"""
    validador = ValidadorCalidad()
    
    print("🔍 SISTEMA DE VALIDACIÓN DE CALIDAD\n")
    print("="*70)
    
    # Validar directorio
    resultados = validador.validar_directorio()
    
    # Generar reporte
    print("\n📊 Generando reporte...")
    reporte = validador.generar_reporte(resultados)
    
    # Resumen final
    print("\n" + "="*70)
    print("RESUMEN DE VALIDACIÓN")
    print("="*70)
    print(f"Total archivos: {reporte['resumen']['total_archivos']}")
    print(f"✅ Válidos: {reporte['resumen']['archivos_validos']} "
          f"({reporte['resumen']['porcentaje_valido']:.1f}%)")
    print(f"❌ Inválidos: {reporte['resumen']['archivos_invalidos']}")
    print(f"⚠️  Con advertencias: {reporte['resumen']['archivos_con_advertencias']}")
    print("="*70)

if __name__ == "__main__":
    main()








