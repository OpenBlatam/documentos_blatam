import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_enhanced_excel(filename):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    workbook = writer.book

    # --- Styles ---
    header_font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid') # Navy Blue
    
    cell_font = Font(name='Segoe UI', size=11)
    currency_format = '$#,##0'
    percent_format = '0.0%'
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), 
                         right=Side(style='thin', color='D9D9D9'), 
                         top=Side(style='thin', color='D9D9D9'), 
                         bottom=Side(style='thin', color='D9D9D9'))

    def apply_professional_style(ws):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = left_align

        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 8, 50)

    # --- 1. Dashboard Summary ---
    ws_dash = workbook.create_sheet("Dashboard")
    ws_dash.sheet_view.showGridLines = False
    
    ws_dash['B2'] = "VENTURA CAPITAL - EXECUTIVE DASHBOARD"
    ws_dash['B2'].font = Font(name='Segoe UI', size=20, bold=True, color='203764')
    
    # Key Metrics Cards
    metrics = [
        ("Pre-money Valuation", "$8.0M", "Series A Target"),
        ("Ask Amount", "$2.0M", "20% Equity"),
        ("ARR (2026 Proj.)", "$3.6M", "12x Multiple"),
        ("LTV/CAC", "9:1", "Unit Economics")
    ]
    
    row_start = 4
    col_start = 2
    for title, value, sub in metrics:
        cell = ws_dash.cell(row=row_start, column=col_start)
        cell.value = title
        cell.font = Font(bold=True, color='7F7F7F')
        
        val_cell = ws_dash.cell(row=row_start+1, column=col_start)
        val_cell.value = value
        val_cell.font = Font(size=18, bold=True, color='203764')
        
        sub_cell = ws_dash.cell(row=row_start+2, column=col_start)
        sub_cell.value = sub
        sub_cell.font = Font(size=9, italic=True, color='A5A5A5')
        
        # Border for card look
        for r in range(row_start, row_start+3):
            ws_dash.cell(row=r, column=col_start).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws_dash.cell(row=r, column=col_start).alignment = Alignment(horizontal='center')
        
        col_start += 2 # Spacing

    # --- 2. Proyecciones Financieras (Enhanced) ---
    # Transposed data for better charting
    fin_data = {
        "Year": [2024, 2025, 2026],
        "ARR": [139000, 1460000, 3600000],
        "Revenue": [100000, 1200000, 3000000], # Approx
        "EBITDA": [-865000, -350000, 1600000]
    }
    df_fin = pd.DataFrame(fin_data)
    df_fin.to_excel(writer, sheet_name='Proyecciones', index=False, startrow=1)
    ws_fin = writer.sheets['Proyecciones']
    apply_professional_style(ws_fin)
    
    # Format Numbers
    for row in ws_fin.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = currency_format

    # Add Line Chart for ARR & Revenue
    chart_fin = LineChart()
    chart_fin.title = "Crecimiento Proyectado (ARR vs Revenue)"
    chart_fin.style = 12
    chart_fin.y_axis.title = "USD"
    chart_fin.x_axis.title = "Year"
    
    data = Reference(ws_fin, min_col=2, min_row=1, max_col=3, max_row=4)
    cats = Reference(ws_fin, min_col=1, min_row=2, max_row=4)
    chart_fin.add_data(data, titles_from_data=True)
    chart_fin.set_categories(cats)
    ws_fin.add_chart(chart_fin, "F2")

    # --- 3. Cap Table (Enhanced with Pie Chart) ---
    cap_data = [
        ["Shareholder", "Shares", "Ownership %", "Value ($)"],
        ["Founders", 8000000, 0.64, 6400000],
        ["ESOP", 2000000, 0.16, 1600000],
        ["Investors (Series A)", 2500000, 0.20, 2000000],
    ]
    
    df_cap = pd.DataFrame(cap_data[1:], columns=cap_data[0])
    df_cap.to_excel(writer, sheet_name='Cap Table', index=False, startrow=1)
    ws_cap = writer.sheets['Cap Table']
    apply_professional_style(ws_cap)
    
    # Percent formatting
    for cell in ws_cap['C']:
        cell.number_format = percent_format
    # Currency formatting
    for cell in ws_cap['D']:
        cell.number_format = currency_format

    # Pie Chart
    pie = PieChart()
    pie.title = "Estructura Accionaria (Post-Money)"
    labels = Reference(ws_cap, min_col=1, min_row=2, max_row=4)
    data = Reference(ws_cap, min_col=3, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws_cap.add_chart(pie, "F2")

    # --- 4. Due Diligence Checklist (Interactive) ---
    dd_headers = ["Categoría", "Documento / Requisito", "Prioridad", "Status", "Notas"]
    dd_items = [
        ["Legal", "Acta Constitutiva", "Alta", "Completado", "Archivo en Data Room"],
        ["Legal", "Poderes Notariales", "Alta", "Completado", ""],
        ["Legal", "Registro de Accionistas", "Alta", "Pendiente", "Actualizar libro"],
        ["Financiero", "Estados Financieros 2023", "Alta", "Completado", "Auditados"],
        ["Financiero", "Proyecciones 2024-2026", "Media", "En Revisión", "Ajustar EBITDA"],
        ["Técnico", "Arquitectura de Software", "Alta", "Completado", ""],
        ["Técnico", "Reporte de Pentesting", "Media", "Pendiente", "Programado Oct"],
        ["Mercado", "Análisis de Competencia", "Baja", "En Revisión", "Agregar nuevos players"]
    ]
    
    df_dd = pd.DataFrame(dd_items, columns=dd_headers)
    df_dd.to_excel(writer, sheet_name='Due Diligence Tracker', index=False, startrow=1)
    ws_dd = writer.sheets['Due Diligence Tracker']
    apply_professional_style(ws_dd)
    
    # Add Data Validation / Conditional Formatting
    # Green for Completado, Yellow for En Revisión, Red for Pendiente
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green_font = Font(color='006100')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    yellow_font = Font(color='9C6500')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_font = Font(color='9C0006')

    ws_dd.conditional_formatting.add('D2:D100', CellIsRule(operator='equal', formula=['"Completado"'], stopIfTrue=True, fill=green_fill, font=green_font))
    ws_dd.conditional_formatting.add('D2:D100', CellIsRule(operator='equal', formula=['"En Revisión"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
    ws_dd.conditional_formatting.add('D2:D100', CellIsRule(operator='equal', formula=['"Pendiente"'], stopIfTrue=True, fill=red_fill, font=red_font))

    writer.close()
    print(f"Enhanced Excel created: {filename}")

if __name__ == "__main__":
    create_enhanced_excel("Ventura_Capital_Herramientas_V2.xlsx")


